import pandas as pd
import requests
import urllib3
from datetime import datetime, timedelta
import time
import os
import tempfile
import win32com.client
import openpyxl
from openpyxl.styles import PatternFill
import concurrent.futures
import logging
import traceback

# Ensure charts generate in the background without opening UI windows
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ─── LOGGING SETUP ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
INPUT_EXCEL_PATH    = r'C:\path\to\your\input.xlsx'
SHEETS_TO_PROCESS   = ['Sheet1', 'Sheet2']   # or ['all'] for all sheets
LOGSOURCE_COLUMN    = 'log source name'
IP_COLUMN           = 'IP'
IN_QRADAR_COLUMN    = 'In Qradar?'

QRADAR_HOST         = 'https://your-qradar-host'
QRADAR_USERNAME     = 'your-username'
QRADAR_PASSWORD     = 'your-password'
VERIFY_SSL          = False

DRAFT_OUTPUT_PATH        = os.path.join(os.path.dirname(INPUT_EXCEL_PATH), 'inactive_and_errors.xlsx')
ACTIVITY_THRESHOLD_DAYS  = 7
REQUEST_TIMEOUT          = 30
MAX_WORKERS              = 10

EXPECTED_LS_TYPES = ['Microsoft Security', 'Linux OS']

# ─── GROUP THRESHOLD CONFIGURATION ────────────────────────────────────────────
GROUP_COLUMN     = None
GROUP_THRESHOLDS = {}

# ─── END CONFIGURATION ─────────────────────────────────────────────────────────

MIN_TIMESTAMP = 0
MAX_TIMESTAMP = 2147483647

LOG_SOURCE_TYPES_CACHE = {}

_MAPI_PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"


# ─── GROUP THRESHOLD RESOLVER ──────────────────────────────────────────────────
def resolve_threshold(group_name=None):
    if not GROUP_COLUMN:
        return ACTIVITY_THRESHOLD_DAYS
    if not group_name or str(group_name).strip().lower() in ('nan', 'none', '', 'null'):
        return ACTIVITY_THRESHOLD_DAYS
    group_clean = str(group_name).strip().lower()
    for key, value in GROUP_THRESHOLDS.items():
        if str(key).strip().lower() == group_clean:
            return value
    return ACTIVITY_THRESHOLD_DAYS


# ─── CHART STATS HELPER ────────────────────────────────────────────────────────
def _chart_stats(stats_dict):
    d = dict(stats_dict)
    d['Active'] = d.get('Active', 0) + d.pop('Inferred', 0)
    return d


def test_qradar_connection(qradar_host, username, password):
    print("🔗 Testing QRadar connection...")
    qradar_host = qradar_host.rstrip('/')
    endpoint    = f"{qradar_host}/api/help/versions"
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0'}
        )
        if resp.status_code == 200:
            print("✅ QRadar connection successful!")
            return True
        elif resp.status_code == 401:
            print("❌ Authentication failed! Check username/password.")
            return False
        else:
            print(f"⚠️ Unexpected response: {resp.status_code}")
            return False
    except requests.exceptions.Timeout:
        print("❌ Connection timed out.")
        return False
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Connection error: {e}")
        return False
    except Exception as e:
        logger.error("Unexpected error during connection test:\n%s", traceback.format_exc())
        print(f"❌ Connection failed: {e}")
        return False


def fetch_log_source_types(qradar_host, username, password):
    print("📥 Fetching Log Source Types Dictionary into memory...")
    qradar_host = qradar_host.rstrip('/')
    endpoint    = f"{qradar_host}/api/config/event_sources/log_source_management/log_source_types"
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0'}
        )
        if resp.status_code == 200:
            for t in resp.json():
                ls_id   = t.get('id')
                ls_name = t.get('name')
                if ls_id is not None and ls_name is not None:
                    LOG_SOURCE_TYPES_CACHE[ls_id] = ls_name
            print(f"✅ Cached {len(LOG_SOURCE_TYPES_CACHE)} Log Source Types.")
        else:
            print(f"⚠️ Failed to fetch Log Source Types. API returned {resp.status_code}.")
    except requests.exceptions.Timeout:
        print("❌ Timed out while fetching Log Source Types.")
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Connection error while fetching Log Source Types: {e}")
    except Exception as e:
        logger.error("Unexpected error in fetch_log_source_types:\n%s", traceback.format_exc())
        print(f"❌ Error fetching Log Source Types: {e}")


def _empty_details():
    return {
        'qradar_id':             'N/A',
        'enabled':               'Unknown',
        'last_seen':             'N/A',
        'activity_status':       'Not Found',
        'days_since_last_event': None,
        'actual_name':           'N/A',
        'ls_type':               'N/A',
        'is_older_expected':     False
    }


def safe_timestamp_conversion(timestamp_ms, threshold=None):
    effective_threshold = threshold if threshold is not None else ACTIVITY_THRESHOLD_DAYS
    if not timestamp_ms:
        return 'No events recorded', 'No Activity', None
    try:
        if isinstance(timestamp_ms, float):
            timestamp_ms = int(timestamp_ms)
        if timestamp_ms > 4102444800:
            timestamp_seconds = timestamp_ms / 1000.0
        else:
            timestamp_seconds = timestamp_ms
        if timestamp_seconds <= MIN_TIMESTAMP or timestamp_seconds > MAX_TIMESTAMP:
            return f'Invalid timestamp: {timestamp_ms}', 'Unknown', None
        last_event_datetime   = datetime.fromtimestamp(timestamp_seconds)
        last_seen             = last_event_datetime.strftime('%Y-%m-%d %H:%M:%S')
        time_diff             = datetime.now() - last_event_datetime
        days_since_last_event = time_diff.days
        threshold_time        = datetime.now() - timedelta(days=effective_threshold)
        activity_status       = 'Active' if last_event_datetime > threshold_time else 'Inactive'
        return last_seen, activity_status, days_since_last_event
    except Exception:
        logger.error("Timestamp conversion error:\n%s", traceback.format_exc())
        return f'Invalid timestamp: {timestamp_ms}', 'Unknown', None


def get_log_source_details(qradar_host, username, password, identifier, is_ip=False, threshold=None):
    clean_identifier = str(identifier).replace('"', '').replace("'", "").strip()
    if is_ip:
        query_filter = f'protocol_parameters contains value="{clean_identifier}" or name ilike "%{clean_identifier}%"'
    else:
        query_filter = f'name ilike "%{clean_identifier}%"'

    ls_endpoint = f"{qradar_host.rstrip('/')}/api/config/event_sources/log_source_management/log_sources"

    try:
        resp = requests.get(
            ls_endpoint,
            params={'filter': query_filter},
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0'}
        )
        if resp.status_code != 200:
            return {'status': f'API Error {resp.status_code}', **_empty_details()}

        ls_data = resp.json()
        if not ls_data:
            return {'status': 'Not Found', **_empty_details()}

        valid_sources = []
        if is_ip:
            for src in ls_data:
                params    = src.get('protocol_parameters', [])
                in_params = any(p.get('value') == clean_identifier for p in params)
                in_name   = clean_identifier.lower() in str(src.get('name', '')).lower()
                if in_params or in_name:
                    valid_sources.append(src)
        else:
            valid_sources = ls_data

        if not valid_sources:
            return {'status': 'Not Found', **_empty_details()}

        expected_sources   = []
        unexpected_sources = []

        for src in valid_sources:
            type_name      = LOG_SOURCE_TYPES_CACHE.get(src.get('type_id'), "")
            api_name_clean = str(type_name).lower()
            is_match = any(
                all(w in api_name_clean for w in str(exp).lower().split())
                for exp in EXPECTED_LS_TYPES
            )
            if is_match:
                expected_sources.append(src)
            else:
                unexpected_sources.append(src)

        def get_best_source(source_list):
            if not source_list:
                return None
            enabled  = [s for s in source_list if s.get('enabled') is True]
            disabled = [s for s in source_list if s.get('enabled') is False]
            enabled.sort(key=lambda x: x.get('last_event_time') or 0, reverse=True)
            disabled.sort(key=lambda x: x.get('last_event_time') or 0, reverse=True)
            return enabled[0] if enabled else disabled[0]

        found_source      = None
        is_older_expected = False
        absolute_max_time = max([s.get('last_event_time') or 0 for s in valid_sources])

        if expected_sources:
            found_source = get_best_source(expected_sources)
            if (found_source.get('last_event_time') or 0) < absolute_max_time:
                is_older_expected = True
        else:
            found_source = get_best_source(unexpected_sources)

        ls_id        = found_source.get('id')
        ls_name      = found_source.get('name', identifier)
        type_id      = found_source.get('type_id')
        ls_type_name = LOG_SOURCE_TYPES_CACHE.get(type_id, f"Unknown Type ID: {type_id}")

        last_event_time_ms = found_source.get('last_event_time')
        last_seen, activity_status, days_since_last_event = safe_timestamp_conversion(
            last_event_time_ms, threshold=threshold
        )

        enabled_str = 'Yes' if found_source.get('enabled', False) else 'No'

        return {
            'status':                'Found',
            'qradar_id':             str(ls_id) if ls_id is not None else '',
            'actual_name':           ls_name,
            'ls_type':               ls_type_name,
            'enabled':               enabled_str,
            'last_seen':             last_seen,
            'activity_status':       activity_status,
            'days_since_last_event': days_since_last_event,
            'is_older_expected':     is_older_expected
        }

    except requests.exceptions.Timeout:
        logger.warning("Request timed out for identifier: %s", identifier)
        return {'status': 'Timeout', **_empty_details()}
    except requests.exceptions.ConnectionError as e:
        logger.error("Connection error for identifier %s: %s", identifier, e)
        return {'status': 'Connection Error', **_empty_details()}
    except Exception as e:
        logger.error("Unexpected error for identifier %s:\n%s", identifier, traceback.format_exc())
        return {'status': f'Error: {str(e)[:50]}', **_empty_details()}


def process_single_row(idx, name_val, ip_val, qradar_host, username, password, group_val=None):
    if name_val and str(name_val).lower() in ['nan', 'none', '', 'null']: name_val = None
    if ip_val   and str(ip_val).lower()   in ['nan', 'none', '', 'null']: ip_val   = None

    effective_threshold = resolve_threshold(group_val)
    details             = None
    search_method       = "None"

    if name_val:
        details = get_log_source_details(
            qradar_host, username, password, name_val,
            is_ip=False, threshold=effective_threshold
        )
        if details['status'] == 'Found':
            search_method = "Name"

    if (not details or details['status'] != 'Found') and ip_val:
        details = get_log_source_details(
            qradar_host, username, password, ip_val,
            is_ip=True, threshold=effective_threshold
        )
        if details['status'] == 'Found':
            search_method = "IP"

    if not details:
        details = {'status': 'Empty/Invalid', **_empty_details()}

    return idx, name_val, details, search_method


def process_sheet(df, sheet_name, qradar_host, username, password,
                  logsource_column, ip_column, in_qradar_col):
    print(f"\n{'='*60}")
    print(f"📋 Processing Sheet: {sheet_name}")
    print(f"{'='*60}")

    if not df.empty:
        df.columns = df.columns.str.strip()

    required_columns = [in_qradar_col, logsource_column, ip_column]
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing columns {missing} in sheet '{sheet_name}'. Skipping.")
        return df

    cols_to_init = {
        'status':                'object',
        'qradar_id':             'object',
        'enabled':               'object',
        'last_seen':             'object',
        'activity_status':       'object',
        'days_since_last_event': 'float64',
        'remarks':               'object',
        'QRadar Actual Name':    'object',
        'Log Source Type':       'object',
        'Is Older Expected':     'bool'
    }
    for col, dtype in cols_to_init.items():
        if col in df.columns:
            df[col] = None
        else:
            df[col] = pd.Series(dtype=dtype)

    row_index_map = {
        df_idx: excel_row
        for excel_row, df_idx in enumerate(df.index, start=2)
    }

    in_qradar_series = df[in_qradar_col].astype(str).str.lower()
    process_mask     = in_qradar_series.str.contains("yes",                 na=False)
    pending_mask     = in_qradar_series.str.contains("pending-maintenance", na=False)

    rows_to_process = df[process_mask]
    total_rows      = len(df)
    target_count    = len(rows_to_process)
    pending_count   = pending_mask.sum()

    group_feature_active = bool(GROUP_COLUMN and GROUP_COLUMN in df.columns)
    if GROUP_COLUMN and not group_feature_active:
        print(f"⚠️  GROUP_COLUMN '{GROUP_COLUMN}' not found — using global threshold.")

    print(f"📊 Total: {total_rows} | To Scan: {target_count} | Pending: {pending_count}")

    skipped_mask = ~(process_mask | pending_mask)
    df.loc[skipped_mask, 'remarks'] = "Skipped (Not Yes or Pending)"

    if pending_count > 0:
        df.loc[pending_mask, 'status']          = 'Pending-Maintenance'
        df.loc[pending_mask, 'activity_status'] = 'Pending-Maintenance'
        df.loc[pending_mask, 'remarks']         = 'Pending Maintenance (Not Scanned)'
        df.loc[pending_mask, 'last_seen']       = 'N/A'

    if target_count == 0:
        return df

    processed_count = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(
                process_single_row,
                idx,
                str(row[logsource_column]).strip(),
                str(row[ip_column]).strip(),
                qradar_host,
                username,
                password,
                str(row[GROUP_COLUMN]).strip() if group_feature_active else None
            ): idx for idx, row in rows_to_process.iterrows()
        }

        for future in concurrent.futures.as_completed(futures):
            processed_count += 1
            try:
                idx, name_val, details, search_method = future.result()
            except Exception as worker_exc:
                original_idx = futures[future]
                logger.error("Worker crashed for row %s:\n%s", original_idx, traceback.format_exc())
                print(f"\n⚠️  [{processed_count}/{target_count}] Worker crashed: {worker_exc}")
                df.at[original_idx, 'status']          = 'Worker Error'
                df.at[original_idx, 'remarks']         = f'Thread exception: {str(worker_exc)[:80]}'
                df.at[original_idx, 'activity_status'] = 'Error'
                df.at[original_idx, 'last_seen']       = 'N/A'
                continue

            print(f"\n🔹 [{processed_count}/{target_count}] {name_val or 'Unknown'} -> {details['status']}")

            df.at[idx, 'QRadar Actual Name'] = details['actual_name']
            df.at[idx, 'Log Source Type']    = details['ls_type']
            df.at[idx, 'Is Older Expected']  = details.get('is_older_expected', False)

            if details['status'] == 'Found':
                df.at[idx, 'qradar_id']             = details['qradar_id']
                df.at[idx, 'enabled']               = details['enabled']
                df.at[idx, 'last_seen']             = details['last_seen']
                df.at[idx, 'days_since_last_event'] = details['days_since_last_event']

                base_remark = f"Found by {search_method}"
                if details.get('is_older_expected'):
                    base_remark += " | ⚠️ Bypassed newer unexpected source"
                    print(f"      🚨 Bypassed newer unexpected log source!")

                if details['enabled'] == 'No':
                    df.at[idx, 'status']          = 'Found'
                    df.at[idx, 'remarks']         = "Disabled on QRadar"
                    df.at[idx, 'activity_status'] = "Disabled"
                    print(f"      ⚪ Status: Disabled")
                else:
                    df.at[idx, 'status']          = 'Found'
                    df.at[idx, 'remarks']         = base_remark
                    df.at[idx, 'activity_status'] = details['activity_status']
                    print(f"      📊 Activity: {details['activity_status']}")
                    print(f"      📅 Last Event: {details['last_seen']} ({details['days_since_last_event']} days ago)")
            else:
                status_val = details['status']
                remark_val = f"❌ {status_val}"
                act_val    = "Error" if "Error" in status_val or "Timeout" in status_val else "Not Found"

                if name_val and "AP" in name_val:
                    status_val = "Inferred"
                    remark_val = "Under WLC (Inferred)"
                    act_val    = "Inferred"
                    print("      ℹ️  Inferred as WLC")
                elif name_val and "FW" in name_val:
                    status_val = "Inferred"
                    remark_val = "Under Forti (Inferred)"
                    act_val    = "Inferred"
                    print("      ℹ️  Inferred as FortiGate")
                else:
                    print(f"      ❌ Result: {status_val.upper()}")

                df.at[idx, 'status']                = status_val
                df.at[idx, 'remarks']               = remark_val
                df.at[idx, 'activity_status']       = act_val
                df.at[idx, 'last_seen']             = "N/A"
                df.at[idx, 'days_since_last_event'] = None

    return df


# ══════════════════════════════════════════════════════════════════════════════
#  VISUAL / AESTHETIC SECTION — all changes below here are purely cosmetic
# ══════════════════════════════════════════════════════════════════════════════

# ─── COLOUR PALETTE ───────────────────────────────────────────────────────────
_P = {
    'bg0':          '#07030f',   # deepest background
    'bg1':          '#0e0820',   # main background
    'bg2':          '#130b28',   # card surface
    'bg3':          '#1a1035',   # elevated / header panels
    'border':       '#2d1b5e',   # panel borders
    'border_soft':  '#1e1240',   # subtle row dividers
    'purple_deep':  '#3b1f7a',   # deep accent fill
    'purple_mid':   '#6d28d9',   # brand purple
    'purple_bright':'#8b5cf6',   # bright violet
    'purple_light': '#a78bfa',   # lavender
    'purple_pale':  '#c4b5fd',   # text highlight
    'text_primary': '#e2d9f3',   # primary body text
    'text_secondary':'#9381b5',  # secondary labels
    'text_muted':   '#5a4a80',   # muted / disabled
    # Status colours
    'green':        '#10b981',
    'red':          '#ef4444',
    'orange':       '#f59e0b',
    'gray':         '#6b7280',
    'cyan':         '#06b6d4',
    'blue':         '#3b82f6',
}

# ─── STATUS METADATA ──────────────────────────────────────────────────────────
_STATUS_META = {
    'Inactive':    {'bg': '#1a0810', 'accent': _P['red'],    'badge_bg': '#7f1d1d', 'label': 'INACTIVE',    'icon': '●'},
    'No Activity': {'bg': '#1a0810', 'accent': _P['red'],    'badge_bg': '#7f1d1d', 'label': 'NO ACTIVITY', 'icon': '●'},
    'Not Found':   {'bg': '#0f0b1e', 'accent': _P['gray'],   'badge_bg': '#374151', 'label': 'NOT FOUND',   'icon': '◌'},
    'Error':       {'bg': '#150c04', 'accent': _P['orange'],  'badge_bg': '#78350f', 'label': 'API ERROR',   'icon': '▲'},
}

def _get_status_meta(activity_status):
    s = str(activity_status).strip()
    for key, meta in _STATUS_META.items():
        if key.lower() in s.lower():
            return meta
    return {
        'bg': '#0f0b1e', 'accent': _P['gray'],
        'badge_bg': '#374151', 'label': s.upper()[:12], 'icon': '◌'
    }


# ─── DONUT CHART GENERATOR ────────────────────────────────────────────────────
def generate_pie_chart(data_dict, title, prefix='qradar_chart'):
    """
    Renders a donut chart on a deep-purple background that matches the email
    aesthetic. Inferred is already merged into Active by _chart_stats() before
    this function is called — no logic change here.
    """
    filtered_data = {k: v for k, v in data_dict.items() if v > 0}
    if not filtered_data:
        return None

    labels = list(filtered_data.keys())
    sizes  = list(filtered_data.values())

    # Cohesive palette that reads well on dark purple
    color_map = {
        'Active':              '#10b981',   # emerald
        'Inactive':            '#ef4444',   # red
        'Not Found':           '#6b7280',   # slate
        'API Errors':          '#f59e0b',   # amber
        'Disabled':            '#06b6d4',   # cyan
        'Inferred':            '#8b5cf6',   # violet (fallback if ever shown)
        'Pending-Maintenance': '#3b82f6',   # blue
    }
    colors = [color_map.get(lbl, '#a78bfa') for lbl in labels]

    bg_color = '#0e0820'

    fig, ax = plt.subplots(figsize=(5, 3.8), facecolor=bg_color)
    ax.set_facecolor(bg_color)

    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=None,           # legend handles labels instead
        colors=colors,
        autopct=lambda pct: f'{pct:.1f}%' if pct > 4 else '',
        startangle=140,
        wedgeprops={'edgecolor': bg_color, 'linewidth': 2.5, 'width': 0.52},
        pctdistance=0.78,
        textprops={'fontsize': 7.5}
    )

    for at in autotexts:
        at.set_color('#f0eaff')
        at.set_fontweight('bold')
        at.set_fontsize(7.5)

    # Centre hole annotation — total count
    total = sum(sizes)
    ax.text(0, 0.08, str(total), ha='center', va='center',
            fontsize=18, fontweight='bold', color='#c4b5fd')
    ax.text(0, -0.22, 'TOTAL', ha='center', va='center',
            fontsize=6.5, color='#6d5a9a', fontweight='600',
            fontfamily='monospace')

    # Legend on the right
    legend_patches = [
        mpatches.Patch(color=color_map.get(lbl, '#a78bfa'),
                       label=f'{lbl}  {v}')
        for lbl, v in zip(labels, sizes)
    ]
    leg = ax.legend(
        handles=legend_patches,
        loc='center left',
        bbox_to_anchor=(1.02, 0.5),
        fontsize=7.5,
        frameon=False,
        labelcolor='#a78bfa',
        handlelength=1.2,
        handleheight=1.0,
        borderpad=0.6,
        labelspacing=0.65,
    )
    for text in leg.get_texts():
        text.set_color('#c4b5fd')

    ax.set_title(title, color='#a78bfa', fontsize=9.5, fontweight='700',
                 pad=12, fontfamily='monospace')

    ax.axis('equal')
    plt.tight_layout(pad=0.4)

    tmp      = tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix=f'{prefix}_')
    filepath = tmp.name
    tmp.close()
    plt.savefig(filepath, bbox_inches='tight', dpi=110,
                facecolor=bg_color, edgecolor='none')
    plt.close()
    return filepath


# ─── OUTLOOK DRAFT HELPER ─────────────────────────────────────────────────────
def create_html_outlook_draft(attachment_path, subject, html_body, image_paths):
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail    = outlook.CreateItem(0)
        mail.Subject = subject

        if os.path.exists(attachment_path):
            mail.Attachments.Add(attachment_path)

        for cid, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                attachment = mail.Attachments.Add(img_path)
                attachment.PropertyAccessor.SetProperty(_MAPI_PR_ATTACH_CONTENT_ID, cid)

        mail.HTMLBody = html_body
        mail.Display()
        print(f"\n✉️  Email draft created successfully.")

    except Exception as e:
        logger.error("Failed to create Outlook draft:\n%s", traceback.format_exc())
        print(f"\n❌ Failed to create Outlook draft: {e}")

    finally:
        for cid, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception as cleanup_err:
                    print(f"⚠️ Could not delete temp image {img_path}: {cleanup_err}")


# ─── ACTIONABLE TABLE ─────────────────────────────────────────────────────────
def _build_actionable_table(report_df, logsource_col, ip_col):
    """
    Inline HTML table of actionable rows — purple dark theme.
    Columns: Hostname · IP · QRadar ID · Last Event · Status
    """
    if report_df is None or len(report_df) == 0:
        return (
            f'<p style="color:{_P["text_muted"]};font-size:11px;'
            f'font-style:italic;padding:10px 2px;">No actionable items for this sheet.</p>'
        )

    P = _P
    rows_html = ''
    for i, (_, row) in enumerate(report_df.iterrows()):
        hostname   = str(row.get(logsource_col, 'N/A') or 'N/A')
        ip_val     = str(row.get(ip_col,        'N/A') or 'N/A')
        qradar_id  = str(row.get('qradar_id',   'N/A') or 'N/A')
        last_seen  = str(row.get('last_seen',   'N/A') or 'N/A')
        act_status = str(row.get('activity_status', 'Unknown') or 'Unknown')

        meta   = _get_status_meta(act_status)
        row_bg = P['bg2'] if i % 2 == 0 else P['bg1']

        days = row.get('days_since_last_event')
        if pd.notna(days) and days is not None:
            try:
                d = int(days)
                days_str = (
                    f'<span style="color:{P["text_muted"]};font-size:9px;'
                    f'display:block;margin-top:2px;font-family:monospace;">'
                    f'{"today" if d == 0 else f"{d}d ago"}</span>'
                )
            except Exception:
                days_str = ''
        else:
            days_str = ''

        hostname_display = hostname if len(hostname) <= 40 else hostname[:38] + '…'

        # Left accent stripe colour based on severity
        left_color = meta['accent']

        rows_html += f"""
        <tr style="background:{row_bg};border-left:3px solid {left_color}20;">
          <td style="padding:9px 14px;border-bottom:1px solid {P['border_soft']};
                     font-size:11px;color:{P['text_primary']};max-width:230px;">
            <span title="{hostname}" style="font-family:monospace;">
              {hostname_display}
            </span>
          </td>
          <td style="padding:9px 14px;border-bottom:1px solid {P['border_soft']};
                     font-size:11px;color:{P['text_secondary']};text-align:center;
                     white-space:nowrap;font-family:monospace;">{ip_val}</td>
          <td style="padding:9px 14px;border-bottom:1px solid {P['border_soft']};
                     font-size:11px;color:{P['purple_light']};text-align:center;
                     white-space:nowrap;font-family:monospace;">{qradar_id}</td>
          <td style="padding:9px 14px;border-bottom:1px solid {P['border_soft']};
                     font-size:11px;color:{P['text_secondary']};text-align:center;">
            {last_seen}{days_str}
          </td>
          <td style="padding:9px 14px;border-bottom:1px solid {P['border_soft']};
                     text-align:center;">
            <span style="
              background:{meta['badge_bg']};
              color:#f0eaff;
              font-size:9px;
              font-weight:700;
              padding:3px 10px;
              border-radius:4px;
              letter-spacing:0.6px;
              white-space:nowrap;
              font-family:monospace;
            ">{meta['icon']}&nbsp;{meta['label']}</span>
          </td>
        </tr>"""

    P = _P
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;margin-top:10px;border-radius:6px;overflow:hidden;">
      <thead>
        <tr style="background:{P['bg3']};">
          <th style="padding:8px 14px;text-align:left;font-size:9px;
                     color:{P['purple_light']};font-weight:700;
                     text-transform:uppercase;letter-spacing:1.2px;
                     border-bottom:2px solid {P['purple_deep']};
                     font-family:monospace;">Hostname / Log Source</th>
          <th style="padding:8px 14px;text-align:center;font-size:9px;
                     color:{P['purple_light']};font-weight:700;
                     text-transform:uppercase;letter-spacing:1.2px;
                     border-bottom:2px solid {P['purple_deep']};
                     font-family:monospace;">IP Address</th>
          <th style="padding:8px 14px;text-align:center;font-size:9px;
                     color:{P['purple_light']};font-weight:700;
                     text-transform:uppercase;letter-spacing:1.2px;
                     border-bottom:2px solid {P['purple_deep']};
                     font-family:monospace;">QRadar ID</th>
          <th style="padding:8px 14px;text-align:center;font-size:9px;
                     color:{P['purple_light']};font-weight:700;
                     text-transform:uppercase;letter-spacing:1.2px;
                     border-bottom:2px solid {P['purple_deep']};
                     font-family:monospace;">Last Event</th>
          <th style="padding:8px 14px;text-align:center;font-size:9px;
                     color:{P['purple_light']};font-weight:700;
                     text-transform:uppercase;letter-spacing:1.2px;
                     border-bottom:2px solid {P['purple_deep']};
                     font-family:monospace;">Status</th>
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
    </table>"""


# ─── FULL EMAIL HTML ───────────────────────────────────────────────────────────
def _build_email_html(global_stats, sheet_stats, total_issues,
                      images_to_embed, report_frames,
                      logsource_col, ip_col):
    """
    Builds the complete HTML email — deep purple QRadar Inventory Validation theme.

    Visual changes vs previous version:
      • Full deep-purple palette, no white/light backgrounds
      • Inferred folded into the Active metric card (no separate Inferred card)
      • Title: QRadar Inventory Validation Report
      • Donut charts replace old flat pies
      • Refined monospace data typography throughout
      • Per-sheet blocks with purple gradient headers
    """
    P        = _P
    run_time = datetime.now().strftime('%d %b %Y  ·  %H:%M:%S')

    # Active = physical active + inferred (display only; logic unchanged)
    active_display = global_stats['Active'] + global_stats['Inferred']

    total_scanned = sum(global_stats.values()) - global_stats['Pending-Maintenance']

    # ── Severity badge ──
    if total_issues == 0:
        badge_bg, badge_label = '#065f46', '✔ &nbsp;ALL CLEAR'
    elif total_issues <= 10:
        badge_bg, badge_label = '#78350f', f'⚠ &nbsp;{total_issues} ISSUES'
    else:
        badge_bg, badge_label = '#7f1d1d', f'⚠ &nbsp;{total_issues} ISSUES'

    # ── Metric card builder ──
    def metric_card(label, value, accent, sub=''):
        sub_html = (
            f'<div style="font-size:9px;color:{P["text_muted"]};'
            f'margin-top:4px;font-family:monospace;">{sub}</div>'
        ) if sub else ''
        return f"""
        <td style="padding:5px 4px;">
          <div style="
            background:{P['bg2']};
            border:1px solid {P['border']};
            border-top:3px solid {accent};
            border-radius:6px;
            padding:14px 16px;
            min-width:95px;
            text-align:center;
          ">
            <div style="font-size:28px;font-weight:800;color:{accent};
                        letter-spacing:-1.5px;line-height:1;
                        font-family:monospace;">{value}</div>
            <div style="font-size:9px;color:{P['text_secondary']};margin-top:6px;
                        text-transform:uppercase;letter-spacing:1px;">{label}</div>
            {sub_html}
          </div>
        </td>"""

    # Inferred is folded into Active — no separate card
    overall_cards = (
        metric_card('Active',         active_display,                  P['green'],  'incl. inferred') +
        metric_card('Inactive',       global_stats['Inactive'],        P['red'])    +
        metric_card('Not Found',      global_stats['Not Found'],       P['gray'])   +
        metric_card('Disabled',       global_stats['Disabled'],        P['cyan'])   +
        metric_card('API Errors',     global_stats['API Errors'],      P['orange']) +
        metric_card('Pending Maint.', global_stats['Pending-Maintenance'], P['blue'])
    )

    # ── Per-sheet section builder ──
    def stat_pill(label, value, color):
        if value == 0:
            return ''
        return (
            f'<span style="display:inline-block;background:{color}18;'
            f'border:1px solid {color}50;color:{color};font-size:9.5px;'
            f'font-weight:700;padding:3px 11px;border-radius:3px;'
            f'margin:3px 4px 3px 0;letter-spacing:0.4px;font-family:monospace;">'
            f'{label}&nbsp;&nbsp;{value}</span>'
        )

    sheet_blocks = ''
    for sheet_name, counts in sheet_stats.items():
        cid         = f"chart_{sheet_name.replace(' ', '_')}"
        sheet_total = sum(counts.values()) - counts['Pending-Maintenance']
        issue_count = counts['Inactive'] + counts['Not Found'] + counts['API Errors']

        pills = (
            stat_pill('ACTIVE',   counts['Active'] + counts['Inferred'], P['green'])  +
            stat_pill('INACTIVE', counts['Inactive'],                    P['red'])    +
            stat_pill('NOT FOUND',counts['Not Found'],                   P['gray'])   +
            stat_pill('DISABLED', counts['Disabled'],                    P['cyan'])   +
            stat_pill('ERRORS',   counts['API Errors'],                  P['orange']) +
            stat_pill('PENDING',  counts['Pending-Maintenance'],         P['blue'])
        )

        chart_html = (
            f'<img src="cid:{cid}" '
            f'style="display:block;max-width:340px;margin:14px auto 0;border-radius:4px;">'
        ) if cid in images_to_embed else ''

        actionable_html = _build_actionable_table(
            report_frames.get(sheet_name), logsource_col, ip_col
        )

        if issue_count > 0:
            hdr_gradient = f'linear-gradient(90deg,{P["purple_deep"]}cc,{P["bg3"]})'
            hdr_border   = P['purple_mid']
            issue_badge  = (
                f'<span style="background:{P["red"]}22;color:{P["red"]};'
                f'font-size:9px;font-weight:700;padding:4px 13px;'
                f'border-radius:3px;border:1px solid {P["red"]}44;'
                f'font-family:monospace;letter-spacing:0.5px;">'
                f'{issue_count} ISSUE{"S" if issue_count != 1 else ""}</span>'
            )
        else:
            hdr_gradient = f'linear-gradient(90deg,#0a2a1a,{P["bg3"]})'
            hdr_border   = P['green']
            issue_badge  = (
                f'<span style="background:{P["green"]}18;color:{P["green"]};'
                f'font-size:9px;font-weight:700;padding:4px 13px;'
                f'border-radius:3px;border:1px solid {P["green"]}44;'
                f'font-family:monospace;letter-spacing:0.5px;">✔ CLEAR</span>'
            )

        sheet_blocks += f"""
        <div style="
          background:{P['bg2']};
          border:1px solid {P['border']};
          border-radius:8px;
          margin-bottom:20px;
          overflow:hidden;
        ">
          <!-- Sheet header bar -->
          <div style="
            background:{hdr_gradient};
            border-left:4px solid {hdr_border};
            padding:13px 18px;
          ">
            <table width="100%" cellpadding="0" cellspacing="0"><tr>
              <td>
                <span style="font-size:13px;font-weight:700;
                             color:{P['text_primary']};letter-spacing:0.3px;">
                  {sheet_name}
                </span>
                <span style="font-size:10px;color:{P['text_muted']};
                             margin-left:12px;font-family:monospace;">
                  {sheet_total} sources scanned
                </span>
              </td>
              <td align="right">{issue_badge}</td>
            </tr></table>
          </div>

          <!-- Stat pills + chart -->
          <div style="padding:14px 18px 12px;">
            <div style="margin-bottom:8px;">{pills}</div>
            {chart_html}
          </div>

          <!-- Actionable table -->
          <div style="
            border-top:1px solid {P['border']};
            padding:12px 18px 16px;
          ">
            <div style="
              font-size:9px;
              color:{P['purple_bright']};
              text-transform:uppercase;
              letter-spacing:2px;
              font-family:monospace;
              margin-bottom:4px;
            ">Requires Attention</div>
            <div style="font-size:10px;color:{P['text_muted']};margin-bottom:6px;">
              Inactive · Not Found · API Errors — full dataset in attached Excel.
            </div>
            {actionable_html}
          </div>
        </div>"""

    # ── Overall chart ──
    overall_chart_html = (
        f'<img src="cid:overall_chart" '
        f'style="display:block;max-width:400px;margin:14px auto 6px;border-radius:4px;">'
    ) if 'overall_chart' in images_to_embed else ''

    # ── Full HTML ──
    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:{P['bg0']};
             font-family:'Segoe UI',Helvetica,Arial,sans-serif;">

<table width="100%" cellpadding="0" cellspacing="0"
       style="background:{P['bg0']};padding:28px 0;">
<tr><td align="center">
<table width="720" cellpadding="0" cellspacing="0"
       style="max-width:720px;width:100%;">

  <!-- ══ HEADER ══════════════════════════════════════════════════ -->
  <tr>
    <td style="
      background:linear-gradient(135deg,{P['bg3']} 0%,#1f1045 55%,{P['purple_deep']} 100%);
      border-radius:10px 10px 0 0;
      padding:30px 34px 26px;
      border-bottom:2px solid {P['purple_mid']};
    ">
      <table width="100%" cellpadding="0" cellspacing="0"><tr>
        <td>
          <div style="
            font-size:9px;
            color:{P['purple_light']};
            letter-spacing:3.5px;
            text-transform:uppercase;
            margin-bottom:10px;
            font-family:monospace;
          ">QRadar · Inventory Validation</div>
          <div style="
            font-size:22px;
            font-weight:800;
            color:{P['text_primary']};
            letter-spacing:-0.5px;
            line-height:1.3;
          ">Log Source<br>Validation Report</div>
          <div style="
            margin-top:10px;
            font-size:11px;
            color:{P['text_muted']};
            font-family:monospace;
            letter-spacing:0.3px;
          ">{run_time}</div>
        </td>
        <td align="right" valign="top">
          <div style="
            background:{badge_bg};
            color:#f0eaff;
            font-size:11px;
            font-weight:700;
            padding:9px 20px;
            border-radius:4px;
            letter-spacing:1px;
            white-space:nowrap;
            font-family:monospace;
          ">{badge_label}</div>
        </td>
      </tr></table>
    </td>
  </tr>

  <!-- ══ ACTION BANNER ══════════════════════════════════════════ -->
  <tr>
    <td style="
      background:#1a0810;
      border-left:4px solid {P['red']};
      padding:12px 22px;
    ">
      <span style="color:{P['red']};font-size:12px;font-weight:700;
                   font-family:monospace;letter-spacing:0.5px;">
        ACTION REQUIRED &nbsp;·&nbsp;
      </span>
      <span style="color:{P['text_secondary']};font-size:12px;">
        {total_issues} issues require attention.
        Attached Excel contains <strong style="color:{P['text_primary']};">actionable items only</strong>.
      </span>
    </td>
  </tr>

  <!-- ══ OVERALL METRICS ════════════════════════════════════════ -->
  <tr>
    <td style="background:{P['bg1']};padding:24px 24px 16px;">
      <div style="
        font-size:9px;
        color:{P['purple_light']};
        text-transform:uppercase;
        letter-spacing:2.5px;
        margin-bottom:16px;
        font-family:monospace;
      ">Overall Summary &nbsp;·&nbsp; {total_scanned} Sources Validated</div>
      <table cellpadding="0" cellspacing="0">
        <tr>{overall_cards}</tr>
      </table>
    </td>
  </tr>

  <!-- ══ OVERALL DONUT CHART ════════════════════════════════════ -->
  <tr>
    <td style="background:{P['bg1']};padding:4px 24px 26px;text-align:center;">
      <div style="
        font-size:9px;
        color:{P['purple_light']};
        text-transform:uppercase;
        letter-spacing:2.5px;
        margin-bottom:8px;
        font-family:monospace;
      ">Inventory Health Distribution</div>
      {overall_chart_html}
    </td>
  </tr>

  <!-- ══ DIVIDER ════════════════════════════════════════════════ -->
  <tr>
    <td style="padding:0;">
      <div style="height:1px;background:linear-gradient(90deg,
        {P['bg0']},{P['purple_mid']} 35%,{P['purple_mid']} 65%,{P['bg0']});"></div>
    </td>
  </tr>

  <!-- ══ PER-SHEET BREAKDOWN ════════════════════════════════════ -->
  <tr>
    <td style="background:{P['bg1']};padding:24px 24px 20px;">
      <div style="
        font-size:9px;
        color:{P['purple_light']};
        text-transform:uppercase;
        letter-spacing:2.5px;
        margin-bottom:4px;
        font-family:monospace;
      ">Breakdown by Sheet</div>
      <div style="font-size:10px;color:{P['text_muted']};margin-bottom:18px;">
        Each section shows health metrics, distribution chart,
        and a table of sources requiring remediation.
      </div>
      {sheet_blocks}
    </td>
  </tr>

  <!-- ══ FOOTER ════════════════════════════════════════════════ -->
  <tr>
    <td style="
      background:{P['bg3']};
      border-radius:0 0 10px 10px;
      padding:16px 34px;
      text-align:center;
      border-top:1px solid {P['border']};
    ">
      <div style="
        font-size:9px;
        color:{P['text_muted']};
        letter-spacing:0.8px;
        font-family:monospace;
      ">QRadar Inventory Validation &nbsp;·&nbsp; Auto-generated {run_time}</div>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  END OF VISUAL SECTION — logic below is unchanged
# ══════════════════════════════════════════════════════════════════════════════

def filter_and_email(processed_sheets_only, draft_path,
                     logsource_col=LOGSOURCE_COLUMN, ip_col=IP_COLUMN):
    report_frames   = {}
    sheet_stats     = {}
    images_to_embed = {}

    global_stats = {
        'Active': 0, 'Inactive': 0, 'Not Found': 0,
        'API Errors': 0, 'Disabled': 0, 'Inferred': 0,
        'Pending-Maintenance': 0
    }

    for name, df in processed_sheets_only.items():
        if 'status' not in df.columns:
            continue
        processed_df = df[df['status'].notna()].copy()
        if len(processed_df) == 0:
            continue

        active_count    = len(processed_df[
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'Active')
        ])
        mask_inactive   = (
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'].isin(['Inactive', 'No Activity']))
        )
        inactive_count  = mask_inactive.sum()
        disabled_count  = len(processed_df[
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'Disabled')
        ])
        inferred_count  = len(processed_df[processed_df['status'] == 'Inferred'])
        mask_not_found  = processed_df['status'] == 'Not Found'
        not_found_count = mask_not_found.sum()
        mask_error      = processed_df['status'].astype(str).str.startswith('API Error', na=False)
        error_count     = mask_error.sum()
        pending_count   = len(processed_df[processed_df['status'] == 'Pending-Maintenance'])

        sheet_counts = {
            'Active': active_count, 'Inactive': inactive_count,
            'Not Found': not_found_count, 'API Errors': error_count,
            'Disabled': disabled_count, 'Inferred': inferred_count,
            'Pending-Maintenance': pending_count
        }

        sheet_stats[name] = sheet_counts
        for k in global_stats:
            global_stats[k] += sheet_counts[k]

        mask_report = mask_inactive | mask_not_found | mask_error

        if mask_report.any():
            sub = processed_df[mask_report].copy()

            def build_inactive_remark(row):
                days = row.get('days_since_last_event')
                if pd.notna(days):
                    return f'Inactive - No events in last {int(days)} days'
                return 'Inactive - No events recorded'

            for idx in sub[mask_inactive.loc[sub.index]].index:
                sub.at[idx, 'remarks'] = build_inactive_remark(sub.loc[idx])

            report_frames[name] = sub

    if not report_frames:
        print("✅ No Actionable Issues detected; skipping email.")
        return

    # ── Save actionable Excel ──
    try:
        with pd.ExcelWriter(draft_path, engine='openpyxl') as writer:
            for sheet_name, df in report_frames.items():
                out_df = df.copy()
                if 'Is Older Expected' in out_df.columns:
                    out_df = out_df.drop(columns=['Is Older Expected'])
                out_df.to_excel(writer, sheet_name=sheet_name, index=False)
    except PermissionError:
        print(f"❌ Could not save report to '{draft_path}'. Is the file open?")
        return

    # ── Generate charts ──
    overall_path = generate_pie_chart(
        _chart_stats(global_stats), "Overall Inventory Status",
        prefix='qradar_overall'
    )
    if overall_path:
        images_to_embed["overall_chart"] = overall_path

    for name, counts in sheet_stats.items():
        cid        = f"chart_{name.replace(' ', '_')}"
        chart_path = generate_pie_chart(
            _chart_stats(counts), f"{name} — Status",
            prefix=f'qradar_{name}'
        )
        if chart_path:
            images_to_embed[cid] = chart_path

    total_issues = (global_stats['Inactive'] +
                    global_stats['Not Found'] +
                    global_stats['API Errors'])

    html_body = _build_email_html(
        global_stats, sheet_stats, total_issues,
        images_to_embed, report_frames,
        logsource_col, ip_col
    )

    subject = (f"QRadar Inventory Validation — "
               f"{total_issues} Source{'s' if total_issues != 1 else ''} Require Attention")
    create_html_outlook_draft(draft_path, subject, html_body, images_to_embed)


def save_surgical_updates_to_excel(filepath, processed_dataframes):
    try:
        wb = openpyxl.load_workbook(filepath)

        red_fill    = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        cols_to_update = [
            'status', 'qradar_id', 'enabled', 'last_seen', 'activity_status',
            'days_since_last_event', 'remarks', 'QRadar Actual Name', 'Log Source Type'
        ]

        for sheet_name, df in processed_dataframes.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws         = wb[sheet_name]
            header_row = 1
            col_map    = {}

            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_row, column=col_idx).value
                if cell_val is not None:
                    col_map[str(cell_val).strip()] = col_idx

            for c in cols_to_update:
                if c not in col_map:
                    new_col_idx    = ws.max_column + 1
                    col_map[c]     = new_col_idx
                    ws.cell(row=header_row, column=new_col_idx).value = c

            row_index_map = {
                df_idx: excel_row
                for excel_row, df_idx in enumerate(df.index, start=2)
            }

            for idx, row in df.iterrows():
                excel_row         = row_index_map[idx]
                is_older_expected = row.get('Is Older Expected', False)

                for c in cols_to_update:
                    val = row[c]
                    if pd.isna(val):
                        val = ""

                    target_cell       = ws.cell(row=excel_row, column=col_map[c])
                    target_cell.value = val

                    if c == 'Log Source Type' and val not in ("", "N/A"):
                        is_match       = False
                        api_name_clean = str(val).lower()
                        for exp in EXPECTED_LS_TYPES:
                            exp_words = str(exp).lower().split()
                            if all(w in api_name_clean for w in exp_words):
                                is_match = True
                                break

                        if is_older_expected:
                            target_cell.fill = red_fill
                        elif not is_match:
                            target_cell.fill = yellow_fill

        wb.save(filepath)
        print("✅ Original Excel file updated surgically.")

    except PermissionError:
        print(f"❌ Permission Denied! '{filepath}' is open. Close it and re-run.")
    except Exception as e:
        logger.error("Failed to surgically save Excel:\n%s", traceback.format_exc())
        print(f"❌ Failed to surgically save Excel: {e}")


def main():
    if not VERIFY_SSL:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print("🚀 Starting QRadar Log Source Checker (Multi-threaded & Surgical)...")

    if GROUP_COLUMN:
        print(f"🏷️  Group Threshold: ENABLED  "
              f"(column: '{GROUP_COLUMN}', {len(GROUP_THRESHOLDS)} threshold(s))")
    else:
        print(f"🏷️  Group Threshold: DISABLED  (global: {ACTIVITY_THRESHOLD_DAYS}d)")

    if not test_qradar_connection(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD):
        return

    fetch_log_source_types(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD)

    print(f"\n📖 Reading Excel: {INPUT_EXCEL_PATH}")
    try:
        all_sheets = pd.read_excel(INPUT_EXCEL_PATH, sheet_name=None)
    except Exception as e:
        logger.error("Failed to read Excel:\n%s", traceback.format_exc())
        print(f"❌ Failed to read Excel: {e}")
        return

    if [s.lower() for s in SHEETS_TO_PROCESS] == ['all']:
        to_process = list(all_sheets.keys())
    else:
        to_process = SHEETS_TO_PROCESS

    for sheet in to_process:
        if sheet not in all_sheets:
            print(f"⚠️  Sheet '{sheet}' not found. Available: {list(all_sheets.keys())}")

    for sheet in to_process:
        if sheet in all_sheets:
            all_sheets[sheet] = process_sheet(
                all_sheets[sheet], sheet,
                QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD,
                LOGSOURCE_COLUMN, IP_COLUMN, IN_QRADAR_COLUMN
            )

    print(f"\n💾 Saving updates to original Excel...")
    processed_sheets_only = {k: v for k, v in all_sheets.items() if k in to_process}

    save_surgical_updates_to_excel(INPUT_EXCEL_PATH, processed_sheets_only)
    filter_and_email(
        processed_sheets_only, DRAFT_OUTPUT_PATH,
        logsource_col=LOGSOURCE_COLUMN, ip_col=IP_COLUMN
    )

    print(f"\n✅ Completed!")


if __name__ == '__main__':
    main()

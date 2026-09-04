"""
QRadar Rule Effectiveness Auditor
==================================
Cross-references every ENABLED correlation rule against offense history
across three lookback windows — 1 / 3 / 6 months — to answer the question
that actually matters to a SOC: are our rules working, and will they fire
when the conditions they were built for actually happen?

Looking at a single 30-day window (the old approach) can't tell the
difference between a rule that's *always* been unused and a rule that fired
every week for a year and then suddenly went quiet last month. Those are
very different problems — one is stale content, the other could be a
silent detection gap (a log source stopped feeding it, a field mapping
changed, a reference set emptied, an upstream rule it depends on broke).
Comparing the three windows side by side is how you tell them apart.

Per rule, this produces:
  · Dead — Never Fired   Zero offenses across the FULL 6-month window.
                          Strongest signal a rule is obsolete or was never
                          wired up correctly. Candidate for review/removal.
  · Recently Silent       Fired at some point in the 6-month window, but
                          nothing at all in the last 90 days. The "used to
                          work, might be quietly broken" signal — investigate
                          before assuming it's just low frequency.
  · Highly Active         Firing well above the expected volume for its
                          window. NOT automatically bad — a prompt to
                          confirm the volume is genuine and thresholds /
                          suppression are tuned as intended.
  · Active                Contributing at a normal, expected rate.

Zero AQL. All data from REST API:
  GET /api/analytics/rules
  GET /api/siem/offenses

Output → five-sheet Excel workbook + an Outlook HTML email draft with a
         1/3/6-month comparison view up top and a full switchable section
         per lookback window underneath.

A NOTE ON RULE DESCRIPTIONS
----------------------------
QRadar's public REST API does not reliably expose a rule's full boolean
test logic (the AND/OR conditions built in the Rule Wizard) as a plain-text
field in most versions. This script pulls whatever descriptive metadata IS
commonly available (name, type, owner, creation/modification dates, and a
notes/description field if your environment populates one) and is written
defensively so it degrades to a clear placeholder rather than guessing. For
the full trigger logic behind a specific rule, cross-reference it in the
QRadar Console (Offenses → Rules → select rule → Actions → Export/View).

A NOTE ON "SWITCHING VIEWS" IN THE EMAIL
------------------------------------------
Outlook's HTML rendering engine (Word) strips JavaScript and doesn't
support the CSS tricks (":target", modern selectors) that make click-to-
hide-and-show tabs work on the web. Building the email as "hidden until
clicked" tabs would risk a recipient opening it in desktop Outlook and
seeing a blank card. Instead, all three windows are always fully visible,
stacked with clear section headers, with a side-by-side comparison table up
top and quick-jump links to each section — this works in every mail client,
Outlook included, while still giving you the three views to compare.
"""

import os
import csv
import json
import time
import tempfile
import traceback
import logging
from datetime import datetime, timedelta

import requests
import urllib3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import win32com.client

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# ─── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  ←  edit this block only
# ══════════════════════════════════════════════════════════════════════════════

QRADAR_HOST     = 'https://your-qradar-host'
QRADAR_USERNAME = 'your-username'
QRADAR_PASSWORD = 'your-password'
VERIFY_SSL      = False          # locally hosted → safe to leave False

# ── Lookback windows ──────────────────────────────────────────────────────────
# The auditor builds THREE views over the same rule set — short / medium /
# long — so a rule that looks merely "quiet" in isolation stands out as a
# real trend once you can see its history.
#
# high_activity_threshold scales with window length (held at a constant
# rate of ~1.7 offenses/day) so "Highly Active" means the same thing in
# every view, instead of the 6-month window flagging everything as active
# purely because it's had more time to accumulate offenses.
LOOKBACK_PERIODS = [
    {'key': '1_month', 'label': '1 Month',  'days': 30,  'high_activity_threshold': 50},
    {'key': '3_month', 'label': '3 Months', 'days': 90,  'high_activity_threshold': 150},
    {'key': '6_month', 'label': '6 Months', 'days': 180, 'high_activity_threshold': 300},
]

# A rule with zero offense contributions across the FULL (longest) lookback
# window is classified "Dead". Kept at 0 — any activity at all means it's
# not truly dead, just possibly rare.
RULE_DEAD_THRESHOLD = 0

# ── "Newly Triggered" window (the Progress signal) ────────────────────────────
# A rule is "Newly Triggered" when EVERY offense it has contributed to,
# anywhere in our full visibility window (the longest LOOKBACK_PERIODS
# window above), falls inside the last NEWLY_TRIGGERED_WINDOW_DAYS days —
# i.e. it fired recently and has zero history before that, as far as we can
# see. This is the ONLY thing the email's Progress section shows.
#
# Independent of the 1/3/6-month views above on purpose, so you can dial in
# exactly what "recent" means for this specific question (e.g. 7 days to
# catch this week's changes) without disturbing the Dead/Highly Active
# thresholds. It MUST be smaller than the longest LOOKBACK_PERIODS window
# (currently 180 days) — otherwise there's no "before" left to compare
# against and everything with any activity would qualify. The script warns
# at runtime if this isn't the case.
NEWLY_TRIGGERED_WINDOW_DAYS = 7

# ── Output ────────────────────────────────────────────────────────────────────
OUTPUT_DIR = r'C:\path\to\your\output'   # folder where Excel file is saved

# ── Run-history tracking (for the "recovered since last audit" insight) ──────
# The ONLY extra moving part this adds is one small JSON file — no database,
# no second spreadsheet to maintain by hand. Each run writes a tiny snapshot
# (rule_id -> status) here; the NEXT run reads it back to work out which
# previously Dead / Recently Silent rules have since started firing again.
# Set to False for zero extra files — everything else still works, you just
# lose the run-over-run "recovered" comparison (the within-run "Newly Active"
# insight below needs no state file at all and always works).
ENABLE_RUN_HISTORY = True
STATE_FILE = os.path.join(OUTPUT_DIR, '.rule_effectiveness_state.json')  # don't delete between runs

# ── Optional hand-written trigger-logic notes ─────────────────────────────────
# QRadar's REST API does not expose a rule's full boolean test logic in most
# versions (see the module docstring) — run inspect_rule_fields.py once to
# confirm what your environment actually returns before assuming otherwise.
# This file is a manual, OPTIONAL top-up: {"<rule_id>": "plain-English trigger
# description"}. You only need to fill it in for rules that actually need
# attention (your Dead / Recently Silent list), not the whole rule base — the
# script merges it in automatically and everything else falls back to
# whatever the API gives us, or the honest "not available" placeholder if
# nothing does. Missing file, or a rule_id missing from it, is fine.
RULE_NOTES_OVERLAY_FILE = os.path.join(OUTPUT_DIR, 'rule_notes_overlay.json')

# ── Manually-tested rules log (drives "Investigated Rules" in the email) ─────
# This is YOUR record of manual triage — not anything QRadar reports. Add one
# row per rule the moment you've manually confirmed it fires as expected
# (via Atomic Red Team, synthetic log injection, whatever method), either by
# editing this CSV directly or with the mark_rule_tested.py helper (which
# also prevents duplicate rows for the same rule_id). Created automatically
# on first run if missing. Columns: rule_id, tested_date, notes.
TESTED_RULES_LOG_FILE = os.path.join(OUTPUT_DIR, 'tested_rules_log.csv')

# ── API / networking ──────────────────────────────────────────────────────────
REQUEST_TIMEOUT  = 30    # seconds per HTTP call
MAX_RETRIES      = 3     # attempts before giving up (exponential backoff)
RETRY_DELAY_BASE = 1.5   # seconds — waits: 1.5s → 3s → 6s
API_PAGE_SIZE    = 9999  # max items per page (QRadar Range header, 0-indexed)
MAX_PAGES        = 100   # safety cap on pagination loops (100 x 10000 = 1M items)
QRADAR_API_VERSION = '14.0'   # sent as the 'Version' header on every request — bump
                               # this if your deployment is on a materially different
                               # QRadar release and you hit unexpected 4xx errors

# ══════════════════════════════════════════════════════════════════════════════
#  END CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

_MAPI_PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"

OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, 'qradar_rule_effectiveness.xlsx')

# Status labels used throughout (kept as constants so every sheet, chart, and
# table agree on the exact string — a typo here would silently break a filter).
STATUS_DEAD             = 'Dead — Never Fired'
STATUS_RECENTLY_SILENT  = 'Recently Silent'
STATUS_HIGHLY_ACTIVE    = 'Highly Active'
STATUS_ACTIVE           = 'Active'

# Per-period statuses (used for the three switchable views) are the same
# idea but scoped to a single window, so "Dead" here just means "0 offenses
# in THIS window" rather than the stronger 6-month claim above.
PERIOD_STATUS_DEAD          = 'Dead'
PERIOD_STATUS_HIGHLY_ACTIVE = 'Highly Active'
PERIOD_STATUS_ACTIVE        = 'Active'

# A rule tagged "newly triggered" fired ONLY within NEWLY_TRIGGERED_WINDOW_DAYS
# and had zero contributions anywhere else in the full lookback window — i.e.
# as far as we can see, it just started working. Computed fresh from a single
# run's offense counts; no history file needed.
LABEL_NEWLY_TRIGGERED = f'Newly Triggered (last {NEWLY_TRIGGERED_WINDOW_DAYS}d, never before)'

PERIOD_NOTE = {
    PERIOD_STATUS_DEAD:          'No offenses in this window — review whether the rule is still needed.',
    PERIOD_STATUS_HIGHLY_ACTIVE: 'Firing well above typical volume for this window — confirm thresholds are tuned as intended.',
    PERIOD_STATUS_ACTIVE:        'Contributing normally.',
}


# ─── SHARED HTTP HELPERS ───────────────────────────────────────────────────────

def _api_get_page(path, range_start, range_end, params=None, label='request'):
    """
    Fetches a single page of a QRadar list endpoint using an explicit Range
    header. Returns (items, total) where total is parsed from the
    Content-Range response header when QRadar provides one, else None.

    Handles exponential-backoff retry on Timeout / ConnectionError and clean
    propagation of 401 / other HTTP errors.
    """
    url     = f"{QRADAR_HOST.rstrip('/')}{path}"
    headers = {
        'Accept':  'application/json',
        'Version': QRADAR_API_VERSION,
        'Range':   f'items={range_start}-{range_end}',
    }
    last_err = None

    for attempt in range(MAX_RETRIES):
        if attempt > 0:
            wait = RETRY_DELAY_BASE * (2 ** (attempt - 1))
            logger.warning("Retry %d/%d for '%s' — waiting %.1fs after %s",
                           attempt, MAX_RETRIES - 1, label, wait,
                           type(last_err).__name__)
            time.sleep(wait)
        try:
            resp = requests.get(
                url,
                params=params,
                auth=(QRADAR_USERNAME, QRADAR_PASSWORD),
                verify=VERIFY_SSL,
                timeout=REQUEST_TIMEOUT,
                headers=headers,
            )
            if resp.status_code in (200, 206):
                total = None
                cr = resp.headers.get('Content-Range', '')
                if cr:
                    try:
                        total = int(cr.split('/')[-1].strip())
                    except Exception:
                        pass
                return resp.json(), total

            elif resp.status_code == 401:
                raise RuntimeError("Authentication failed — check QRADAR_USERNAME / PASSWORD")
            else:
                last_err = RuntimeError(f"HTTP {resp.status_code}")
                logger.warning("HTTP %d for '%s' (attempt %d/%d)",
                               resp.status_code, label, attempt + 1, MAX_RETRIES)

        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as exc:
            last_err = exc
            logger.warning("%s on attempt %d for '%s'",
                           type(exc).__name__, attempt + 1, label)
        except RuntimeError:
            raise
        except Exception:
            logger.error("Non-retriable error for '%s':\n%s", label, traceback.format_exc())
            raise

    raise RuntimeError(f"All {MAX_RETRIES} attempts failed for '{label}': {last_err}")


def _api_get(path, params=None, label='request'):
    """Single-page GET — fine for small endpoints. Warns (but doesn't
    auto-paginate) if the server reports more items than fit in one page."""
    items, total = _api_get_page(path, 0, API_PAGE_SIZE, params=params, label=label)
    if total is not None and total > API_PAGE_SIZE + 1:
        logger.warning(
            "Pagination cap hit for '%s': %d items on server, only %d fetched. "
            "Use _api_get_all for this endpoint.", label, total, API_PAGE_SIZE + 1
        )
    return items


def _api_get_all(path, params=None, label='request', page_size=None, max_pages=MAX_PAGES):
    """
    Loops _api_get_page until every item is fetched. This matters once the
    offense lookback stretches to 3-6 months — a busy SOC can generate more
    offenses in that span than fit in a single page, and silently
    truncating that data would skew the whole rule-effectiveness picture
    toward "everything looks dead."
    """
    page_size = page_size or (API_PAGE_SIZE + 1)
    all_items = []
    start = 0
    for page_num in range(max_pages):
        end = start + page_size - 1
        items, total = _api_get_page(
            path, start, end, params=params,
            label=f'{label} (items {start}-{end})'
        )
        if not items:
            break
        all_items.extend(items)
        if total is not None and len(all_items) >= total:
            break
        if len(items) < page_size:
            break
        start += page_size
    else:
        logger.warning(
            "Hit max_pages=%d for '%s' — data may be incomplete. Raise MAX_PAGES "
            "if your deployment has more history than that.", max_pages, label
        )
    return all_items


# ─── CONNECTION TEST ──────────────────────────────────────────────────────────

def validate_config():
    """
    Catches the mistakes people actually make editing the config block,
    before they turn into a confusing failure five minutes into a run.
    Returns True if it's safe to proceed.
    """
    problems = []

    if QRADAR_HOST.rstrip('/') in ('https://your-qradar-host', ''):
        problems.append("QRADAR_HOST is still the placeholder — set it to your real console URL.")
    if QRADAR_USERNAME in ('your-username', '') or QRADAR_PASSWORD in ('your-password', ''):
        problems.append("QRADAR_USERNAME / QRADAR_PASSWORD look like the placeholders.")
    if OUTPUT_DIR == r'C:\path\to\your\output':
        problems.append("OUTPUT_DIR is still the placeholder path.")

    days = [p['days'] for p in LOOKBACK_PERIODS]
    if len(LOOKBACK_PERIODS) != 3:
        problems.append(f"LOOKBACK_PERIODS has {len(LOOKBACK_PERIODS)} entries — the auditor "
                         f"assumes exactly 3 (short/mid/long) throughout.")
    elif days != sorted(days) or len(set(days)) != 3:
        problems.append(f"LOOKBACK_PERIODS days must be strictly increasing (got {days}).")

    thresholds = [p['high_activity_threshold'] for p in LOOKBACK_PERIODS]
    if thresholds != sorted(thresholds):
        problems.append(f"high_activity_threshold should increase with window length (got {thresholds}) "
                         f"— otherwise 'Highly Active' won't mean a consistent rate across windows.")

    if NEWLY_TRIGGERED_WINDOW_DAYS >= LOOKBACK_PERIODS[-1]['days']:
        problems.append(f"NEWLY_TRIGGERED_WINDOW_DAYS ({NEWLY_TRIGGERED_WINDOW_DAYS}) must be smaller "
                         f"than the longest lookback window ({LOOKBACK_PERIODS[-1]['days']} days).")
    if NEWLY_TRIGGERED_WINDOW_DAYS <= 0:
        problems.append("NEWLY_TRIGGERED_WINDOW_DAYS must be a positive number of days.")

    if problems:
        print("⚠️  Configuration issues found:")
        for p in problems:
            print(f"   - {p}")
        print()

    # Placeholder host/creds are fatal (nothing downstream will work);
    # everything else is a warning we can still run with.
    fatal = any('placeholder' in p and 'QRADAR_' in p for p in problems)
    return not fatal


def test_connection():
    print("🔗 Testing QRadar connection...")
    try:
        result = _api_get('/api/help/versions', label='connection test')
        if result:
            print("✅ QRadar connection successful!")
            return True
        print("⚠️  Unexpected empty response from /api/help/versions")
        return False
    except RuntimeError as e:
        print(f"❌ {e}")
        return False
    except Exception as e:
        print(f"❌ Connection failed: {e}")
        return False


# ─── DATA FETCHERS ─────────────────────────────────────────────────────────────

def fetch_all_rules():
    """
    Returns list of enabled correlation rule objects. Building blocks are
    excluded — they never generate offenses and would pollute the dead-rule
    analysis with hundreds of false positives.
    """
    print("📥 Fetching correlation rules...")
    try:
        data = _api_get_all('/api/analytics/rules', label='analytics rules')
        rules = [
            r for r in data
            if r.get('enabled', False) is True
            and 'BB' not in str(r.get('type', '')).upper()
            and 'BUILDING_BLOCK' not in str(r.get('type', '')).upper()
        ]
        print(f"   ✅ {len(rules)} enabled correlation rules (of {len(data)} total).")
        return rules
    except Exception as e:
        print(f"   ❌ Failed to fetch rules: {e}")
        return []


def fetch_offenses_for_lookback(days):
    """
    Fetches every offense with start_time within the last `days` days,
    fully paginated. Called ONCE for the longest configured window — the
    shorter 1/3-month views are derived by slicing this same list in memory
    (see filter_offenses_by_days), so QRadar is only queried once no matter
    how many lookback windows are configured.
    """
    print(f"📥 Fetching offenses from the last {days} days (paginated)...")
    cutoff_ms  = int((datetime.now() - timedelta(days=days)).timestamp() * 1000)
    filter_str = f'start_time >= {cutoff_ms}'
    fields_str = 'id,rules,start_time,magnitude,status,severity'
    try:
        data = _api_get_all(
            '/api/siem/offenses',
            params={'filter': filter_str, 'fields': fields_str},
            label='offenses'
        )
        print(f"   ✅ {len(data)} offenses fetched.")
        return data
    except Exception as e:
        print(f"   ⚠️  Filtered offense fetch failed ({e}). Trying unfiltered...")
        try:
            data = _api_get_all('/api/siem/offenses', label='offenses (unfiltered)')
            data = [o for o in data if (o.get('start_time') or 0) >= cutoff_ms]
            print(f"   ✅ {len(data)} offenses after in-memory filter.")
            return data
        except Exception as e2:
            print(f"   ❌ Offense fetch failed entirely: {e2} — rule analysis will show every rule as having zero contributions.")
            return []


def filter_offenses_by_days(offenses, days):
    """Slices an already-fetched offense list down to the last `days` days."""
    cutoff_ms = int((datetime.now() - timedelta(days=days)).timestamp() * 1000)
    return [o for o in offenses if (o.get('start_time') or 0) >= cutoff_ms]


# ─── RULE METADATA HELPERS ─────────────────────────────────────────────────────

def _extract_rule_ids(offense):
    """
    Extracts contributing rule IDs from an offense object.
    Handles: list of dicts ({'id': X, 'type': Y}), list of ints, or missing field.
    """
    ids = []
    for r in offense.get('rules', []):
        if isinstance(r, dict):
            rid = r.get('id')
        elif isinstance(r, (int, str)):
            rid = r
        else:
            rid = None
        if rid is not None:
            try:
                ids.append(int(rid))
            except (TypeError, ValueError):
                pass
    return ids


def _html_escape(text):
    """
    Every rule name / description / note that lands in the email comes from
    QRadar or a human-edited file, not from us — none of it is guaranteed
    to be HTML-safe. A rule named "Alerts < 5 & > 1" or a CSV note with a
    stray "<" would otherwise corrupt the email's markup. Cheap insurance,
    always applied right before a value is dropped into an f-string of HTML.
    """
    if text is None:
        return ''
    text = str(text)
    return (text.replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;'))


def _csv_formula_safe(text):
    """
    Defangs a value before it's written into a cell that Excel/CSV might
    interpret as a formula (a leading =, +, -, @, tab, or CR can trigger
    "CSV/Excel formula injection" if the file is ever opened by someone
    other than its author). These notes come from a human-edited file, so
    the risk here is low, but it costs nothing to prefix defensively.
    """
    if text is None:
        return ''
    text = str(text)
    if text and text[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + text
    return text


def _format_epoch_ms(val):
    """QRadar timestamps are epoch milliseconds. Returns a readable date, or '—' if absent/invalid."""
    if not val:
        return '—'
    try:
        return datetime.fromtimestamp(int(val) / 1000).strftime('%Y-%m-%d')
    except (TypeError, ValueError, OSError, OverflowError):
        return '—'


def _get_rule_description(rule):
    """
    Best-effort human-readable description of what the rule does / how it
    can be triggered.

    IMPORTANT: QRadar's public REST API does not reliably expose the full
    boolean test-stack (the AND/OR conditions built in the Rule Wizard) as
    plain text in most versions. This checks the few field names that
    sometimes carry notes in different QRadar versions, and falls back to a
    placeholder pointing back to the console rather than guessing.
    """
    for field in ('notes', 'description', 'rule_description', 'comment'):
        val = rule.get(field)
        if val and isinstance(val, str) and val.strip():
            return val.strip()
    return 'Not exposed via API — review in QRadar Console (Offenses → Rules) for full trigger logic.'


def _load_notes_overlay():
    """Loads the OPTIONAL hand-written rule_id -> description overlay. Missing
    file (the common case until you've documented a few rules) is not an
    error — just means nothing gets overridden."""
    if not os.path.exists(RULE_NOTES_OVERLAY_FILE):
        return {}
    try:
        with open(RULE_NOTES_OVERLAY_FILE, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.warning("Could not read rule notes overlay (%s) — ignoring it this run.", e)
        return {}


# ─── MANUALLY-TESTED RULES LOG ──────────────────────────────────────────────
# Everything above this point is derived from QRadar's own data. This part
# is the one place the auditor tracks something QRadar can't tell it: which
# rules a human has actually gone and tested.

def _ensure_tested_rules_log_exists():
    """Creates the CSV with just a header row on first run. Safe to call
    every run — it's a no-op once the file exists."""
    if os.path.exists(TESTED_RULES_LOG_FILE):
        return
    try:
        os.makedirs(os.path.dirname(TESTED_RULES_LOG_FILE), exist_ok=True)
        with open(TESTED_RULES_LOG_FILE, 'w', newline='') as f:
            csv.writer(f).writerow(['rule_id', 'tested_date', 'notes'])
        print(f"   📝 Created {TESTED_RULES_LOG_FILE} — add a row (or run "
              f"mark_rule_tested.py) each time you manually confirm a rule fires.")
    except Exception as e:
        logger.warning("Could not create tested rules log (%s).", e)


def load_tested_rules_log():
    """
    Reads the manually-maintained tested-rules log. De-duplicated by
    rule_id — if the same rule was logged more than once (e.g. re-tested
    after a fix), the LAST row for that rule_id wins, but it only ever
    counts once toward the total.
    """
    _ensure_tested_rules_log_exists()
    entries = {}
    try:
        with open(TESTED_RULES_LOG_FILE, 'r', newline='') as f:
            for row in csv.DictReader(f):
                rid = str(row.get('rule_id', '')).strip()
                if not rid:
                    continue
                entries[rid] = {
                    'rule_id':     rid,
                    'tested_date': (row.get('tested_date') or '').strip(),
                    'notes':       (row.get('notes') or '').strip(),
                }
    except Exception as e:
        logger.warning("Could not read tested rules log (%s) — treating as empty.", e)
    return entries


def build_tested_rules_view(master_df, tested_entries):
    """Joins the tested-rules log with CURRENT rule metadata (name/type),
    so a rename in QRadar doesn't leave the log showing a stale name. A
    rule_id that no longer exists is kept (the testing work still happened)
    but flagged as no longer found."""
    cols = ['rule_id', 'rule_name', 'rule_type', 'tested_date', 'notes']
    if not tested_entries:
        return pd.DataFrame(columns=cols)

    lookup = {}
    if not master_df.empty:
        for _, r in master_df.iterrows():
            lookup[str(int(r['rule_id']))] = (r['rule_name'], r['rule_type'])

    rows = []
    for rid, entry in tested_entries.items():
        name, rtype = lookup.get(rid, (f'(rule no longer found — ID {rid})', 'Unknown'))
        rows.append({
            'rule_id': rid, 'rule_name': name, 'rule_type': rtype,
            'tested_date': entry['tested_date'], 'notes': _csv_formula_safe(entry['notes']),
        })
    df = pd.DataFrame(rows)

    def _parse_date(s):
        for fmt in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f'):
            try:
                return datetime.strptime(s, fmt)
            except (ValueError, TypeError):
                continue
        return datetime.min

    df['_sort'] = df['tested_date'].apply(_parse_date)
    df = df.sort_values('_sort', ascending=False).drop(columns=['_sort']).reset_index(drop=True)
    return df


def _classify_single_period(count, high_activity_threshold):
    """Independent Dead/Highly Active/Active classification using ONLY this window's count."""
    if count <= RULE_DEAD_THRESHOLD:
        return PERIOD_STATUS_DEAD
    if count >= high_activity_threshold:
        return PERIOD_STATUS_HIGHLY_ACTIVE
    return PERIOD_STATUS_ACTIVE


def _recommendation_for_status(status):
    return {
        STATUS_DEAD: (
            f"No offense contributions across the full {LOOKBACK_PERIODS[-1]['days']}-day "
            "lookback. Review the rule's logic and confirm it's still needed — "
            "disable it if it's obsolete."
        ),
        STATUS_RECENTLY_SILENT: (
            "Fired in the past but nothing recently. This can mean the condition "
            "genuinely hasn't occurred — or that something upstream broke (a log "
            "source stopped feeding it, a field mapping changed, a reference set "
            "emptied). Worth a quick check rather than assuming it's just quiet."
        ),
        STATUS_HIGHLY_ACTIVE: (
            "Firing well above typical volume for its window. Confirm this reflects "
            "genuine, expected activity and that thresholds/suppression are tuned "
            "as intended."
        ),
        STATUS_ACTIVE: 'Contributing normally across all lookback windows. No action required.',
    }.get(status, '')


# ─── RULE ANALYSIS ────────────────────────────────────────────────────────────

def build_master_rule_table(rules, offenses_all):
    """
    Builds one row per enabled rule with:
      - descriptive metadata (name, type, origin, owner, dates, description)
      - offense counts for each configured lookback period
      - an INDEPENDENT Dead/Highly Active/Active status per period
        (status_1_month, status_3_month, status_6_month — drives the three
        switchable email views)
      - one COMPOSITE 'overall_status' that reasons across periods together
        (drives the Excel workbook and the "needs investigation" callout) —
        this is what catches a rule that fired steadily for months and then
        went quiet, which looks merely "quiet" in any single-period view
        but is a much stronger signal once you can see the history.
    """
    period_offenses = {p['key']: filter_offenses_by_days(offenses_all, p['days'])
                        for p in LOOKBACK_PERIODS}

    period_counts = {}
    for key, offs in period_offenses.items():
        counts = {}
        for o in offs:
            for rid in _extract_rule_ids(o):
                counts[rid] = counts.get(rid, 0) + 1
        period_counts[key] = counts

    # Independent "recent window" for the Newly Triggered signal — deliberately
    # NOT tied to LOOKBACK_PERIODS[0], so NEWLY_TRIGGERED_WINDOW_DAYS can be
    # tuned on its own (e.g. 7 days) without disturbing the Dead/Highly Active
    # per-period thresholds.
    newly_window_offenses = filter_offenses_by_days(offenses_all, NEWLY_TRIGGERED_WINDOW_DAYS)
    newly_window_counts = {}
    for o in newly_window_offenses:
        for rid in _extract_rule_ids(o):
            newly_window_counts[rid] = newly_window_counts.get(rid, 0) + 1

    notes_overlay = _load_notes_overlay()   # optional, hand-written — see config block

    rows = []
    for rule in rules:
        rid = rule.get('id')
        overlay_note = notes_overlay.get(str(rid))
        row = {
            'rule_id':     rid,
            'rule_name':   _csv_formula_safe(rule.get('name', f'Rule {rid}')),
            'rule_type':   rule.get('type', 'UNKNOWN'),
            'origin':      rule.get('origin', 'UNKNOWN'),
            'owner':       rule.get('owner', 'Unknown'),
            'created':     _format_epoch_ms(rule.get('creation_date')),
            'modified':    _format_epoch_ms(rule.get('modification_date')),
            'description': _csv_formula_safe(overlay_note.strip() if overlay_note and overlay_note.strip() else _get_rule_description(rule)),
            'offenses_newly_window': newly_window_counts.get(rid, 0),
        }
        for p in LOOKBACK_PERIODS:
            count = period_counts[p['key']].get(rid, 0)
            row[f"offenses_{p['key']}"] = count
            row[f"status_{p['key']}"]   = _classify_single_period(count, p['high_activity_threshold'])
        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    longest_key  = LOOKBACK_PERIODS[-1]['key']    # e.g. 6_month — defines "ever fired"
    mid_key      = LOOKBACK_PERIODS[1]['key']     # e.g. 3_month — defines "still recently active"
    shortest_key = LOOKBACK_PERIODS[0]['key']     # e.g. 1_month — defines "currently highly active"

    def _overall(r):
        if r[f'offenses_{longest_key}'] <= RULE_DEAD_THRESHOLD:
            return STATUS_DEAD
        if r[f'offenses_{mid_key}'] <= RULE_DEAD_THRESHOLD:
            return STATUS_RECENTLY_SILENT
        if r[f'status_{shortest_key}'] == PERIOD_STATUS_HIGHLY_ACTIVE:
            return STATUS_HIGHLY_ACTIVE
        return STATUS_ACTIVE

    df['overall_status']  = df.apply(_overall, axis=1)
    df['recommendation']  = df['overall_status'].map(_recommendation_for_status)

    # "Triggered in the last NEWLY_TRIGGERED_WINDOW_DAYS days and never
    # before that" — every offense this rule has EVER contributed to, within
    # our full visibility window, falls inside the recent window. The recent
    # window is always a subset of the full window, so offenses_6m ==
    # offenses_newly_window means nothing happened outside it.
    df['newly_triggered'] = (
        (df[f'offenses_{longest_key}'] == df['offenses_newly_window'])
        & (df['offenses_newly_window'] > 0)
    )
    return df


# ─── RUN-HISTORY TRACKING (recovered-since-last-audit) ─────────────────────────
# This is the one piece of state the auditor keeps between runs, and it's
# deliberately tiny: a JSON dict of {rule_id: last-seen status}. It is NOT a
# database and it isn't meant to be browsed by anyone — it exists purely so
# the next run can answer "did any of the rules we flagged last time start
# working again?" without you having to remember or re-track anything.

def load_previous_state():
    """
    Returns the previous run's {rule_id: {...}} snapshot, or None if there
    isn't one yet (first run, or history tracking was off last time). None
    is a distinct signal from {} — it means "no baseline to compare against"
    rather than "compared, and nothing matched."
    """
    if not os.path.exists(STATE_FILE):
        return None
    try:
        with open(STATE_FILE, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.warning("Could not read previous state file (%s) — treating this as the first run.", e)
        return None


def save_current_state(master_df, run_timestamp):
    """Persists a minimal per-rule snapshot for the NEXT run to compare against."""
    state = {}
    for _, r in master_df.iterrows():
        state[str(int(r['rule_id']))] = {
            'rule_name':      r['rule_name'],
            'overall_status': r['overall_status'],
            'run_timestamp':  run_timestamp,
        }
    try:
        os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
        with open(STATE_FILE, 'w') as f:
            json.dump(state, f, indent=2)
    except Exception as e:
        logger.warning("Could not write state file (%s) — the 'recovered since last audit' "
                        "comparison will be unavailable next run.", e)


def find_recovered_rules(master_df, previous_state):
    """
    Rules that were Dead or Recently Silent as of the PREVIOUS run and are
    now Active / Highly Active / no-longer-flagged — i.e. they've started
    firing again since we last called them out. This is the "did the fix
    actually work" signal, and it's the one that needs the state file: it's
    inherently a comparison across two points in time, not something a
    single run's data can answer on its own.
    """
    if not previous_state or master_df.empty:
        return pd.DataFrame()

    flagged_before = {STATUS_DEAD, STATUS_RECENTLY_SILENT}
    rows = []
    for _, r in master_df.iterrows():
        rid  = str(int(r['rule_id']))
        prev = previous_state.get(rid)
        if prev and prev.get('overall_status') in flagged_before and r['overall_status'] not in flagged_before:
            rows.append({
                'rule_id':          int(r['rule_id']),
                'rule_name':        r['rule_name'],
                'rule_type':        r['rule_type'],
                'previous_status':  prev.get('overall_status'),
                'current_status':   r['overall_status'],
                'offenses_recent':  int(r[f"offenses_{LOOKBACK_PERIODS[0]['key']}"]),
                'previous_run':     prev.get('run_timestamp', 'unknown'),
            })
    return pd.DataFrame(rows)


# ─── CHART ────────────────────────────────────────────────────────────────────

_BG = '#0a0618'   # dark background baked into the chart PNG


def generate_rule_trend_chart(master_df):
    """
    Grouped bar chart: Dead / Highly Active / Active rule counts, one group
    per lookback window. This is the single visual answer to "switching
    between views" — rather than hiding two windows at a time, it puts all
    three side by side so the trend (rules going dead vs. staying active as
    the window widens) is visible in one glance.
    """
    if master_df.empty:
        return None

    categories = [PERIOD_STATUS_DEAD, PERIOD_STATUS_HIGHLY_ACTIVE, PERIOD_STATUS_ACTIVE]
    colours    = {PERIOD_STATUS_DEAD: '#ef4444', PERIOD_STATUS_HIGHLY_ACTIVE: '#3b82f6', PERIOD_STATUS_ACTIVE: '#10b981'}

    period_labels = [p['label'] for p in LOOKBACK_PERIODS]
    counts_by_cat = {cat: [] for cat in categories}
    for p in LOOKBACK_PERIODS:
        vc = master_df[f"status_{p['key']}"].value_counts()
        for cat in categories:
            counts_by_cat[cat].append(int(vc.get(cat, 0)))

    fig, ax = plt.subplots(figsize=(6.5, 4), facecolor=_BG)
    ax.set_facecolor(_BG)

    x_positions = list(range(len(period_labels)))
    width   = 0.24
    offsets = [-width, 0, width]

    max_val = 1
    for i, cat in enumerate(categories):
        positions = [xi + offsets[i] for xi in x_positions]
        bars = ax.bar(positions, counts_by_cat[cat], width=width,
                       color=colours[cat], label=cat, edgecolor=_BG, linewidth=1)
        for bar, val in zip(bars, counts_by_cat[cat]):
            max_val = max(max_val, val)
            if val > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                        str(val), ha='center', va='bottom', color='#c4b5fd',
                        fontsize=8, fontfamily='monospace')

    ax.set_ylim(0, max_val * 1.32)
    ax.set_xticks(x_positions)
    ax.set_xticklabels(period_labels, color='#c4b5fd', fontsize=9, fontfamily='monospace')
    ax.set_ylabel('Rule count', color='#7c6fa0', fontsize=8, fontfamily='monospace')
    ax.set_title('Rule Status Across Lookback Windows', color='#a78bfa',
                 fontsize=10, fontweight='700', fontfamily='monospace', pad=12)
    ax.tick_params(colors='#c4b5fd', labelsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor('#3a2d6a')
        spine.set_linewidth(0.5)
    ax.legend(fontsize=8, frameon=False, labelcolor='#c4b5fd', loc='upper left')

    plt.tight_layout(pad=0.8)
    tmp  = tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix='qr_trend_')
    path = tmp.name
    tmp.close()
    plt.savefig(path, bbox_inches='tight', dpi=110, facecolor=_BG, edgecolor='none')
    plt.close(fig)
    return path


# ─── EXCEL REPORT ─────────────────────────────────────────────────────────────

_FILLS = {
    'red':    PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
    'orange': PatternFill(start_color='FFBF47', end_color='FFBF47', fill_type='solid'),
    'green':  PatternFill(start_color='A8E6CF', end_color='A8E6CF', fill_type='solid'),
    'blue':   PatternFill(start_color='74B9FF', end_color='74B9FF', fill_type='solid'),
    'header': PatternFill(start_color='2D2257', end_color='2D2257', fill_type='solid'),
}
_HDR_FONT = Font(bold=True, color='E8E0FF', size=10)
_BOLD     = Font(bold=True, size=10)
_CENTRE   = Alignment(horizontal='center', vertical='center')
_LEFT     = Alignment(horizontal='left',   vertical='center')
_WRAP     = Alignment(horizontal='left',   vertical='top', wrap_text=True)


def _status_fill_key(status):
    return {
        STATUS_DEAD:            'red',
        STATUS_RECENTLY_SILENT: 'orange',
        STATUS_HIGHLY_ACTIVE:   'blue',
        STATUS_ACTIVE:          'green',
    }.get(status, 'green')


def _write_sheet_header(ws, columns, col_widths):
    ws.append(columns)
    for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
        cell.fill      = _FILLS['header']
        cell.font      = _HDR_FONT
        cell.alignment = _CENTRE
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20


def _write_summary_sheet(wb, master_df):
    ws = wb.active
    ws.title = 'Executive Summary'

    ws['A1'] = 'QRadar Rule Effectiveness Report'
    ws['A1'].font = Font(bold=True, size=16, color='2D2257')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10, color='7C6FA0')
    ws['A3'] = f"Lookback windows: {', '.join(p['label'] for p in LOOKBACK_PERIODS)}"
    ws['A3'].font = Font(size=10, color='7C6FA0')

    ws['A5'] = 'Total enabled correlation rules analyzed:'
    ws['B5'] = int(len(master_df))
    ws['B5'].font = _BOLD

    row = 7
    ws.cell(row=row, column=1, value='Status').fill = _FILLS['header']
    ws.cell(row=row, column=1).font = _HDR_FONT
    for i, p in enumerate(LOOKBACK_PERIODS, start=2):
        c = ws.cell(row=row, column=i, value=p['label'])
        c.fill = _FILLS['header']
        c.font = _HDR_FONT
        c.alignment = _CENTRE
    row += 1

    for status, fill_key in [
        (PERIOD_STATUS_DEAD, 'red'),
        (PERIOD_STATUS_HIGHLY_ACTIVE, 'blue'),
        (PERIOD_STATUS_ACTIVE, 'green'),
    ]:
        ws.cell(row=row, column=1, value=status).font = _BOLD
        for i, p in enumerate(LOOKBACK_PERIODS, start=2):
            count = int((master_df[f"status_{p['key']}"] == status).sum())
            c = ws.cell(row=row, column=i, value=count)
            c.alignment = _CENTRE
            c.fill = _FILLS[fill_key]
        row += 1

    row += 1
    silent_count = int((master_df['overall_status'] == STATUS_RECENTLY_SILENT).sum())
    dead_count   = int((master_df['overall_status'] == STATUS_DEAD).sum())
    ws.cell(row=row, column=1, value='Dead — never fired in 6 months:').font = _BOLD
    ws.cell(row=row, column=2, value=dead_count).font = _BOLD
    row += 1
    ws.cell(row=row, column=1, value='Recently silent — fired historically, quiet now:').font = _BOLD
    ws.cell(row=row, column=2, value=silent_count).font = _BOLD
    row += 2

    note = (
        'Note: rule descriptions reflect whatever metadata QRadar exposes via REST '
        'API for each rule. Full boolean trigger logic (the AND/OR test stack built '
        'in the Rule Wizard) is not reliably exposed via this API in most QRadar '
        'versions — cross-reference the Rules console for full logic wherever a '
        'description is marked as unavailable.'
    )
    ws.cell(row=row, column=1, value=note)
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color='7C6FA0')
    ws.cell(row=row, column=1).alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 48

    for i, w in enumerate([46, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_master_sheet(wb, master_df):
    ws = wb.create_sheet('Rule Effectiveness (All Rules)')

    period_headers = [f"Offenses ({p['label']})" for p in LOOKBACK_PERIODS]
    cols = (['Rule ID', 'Rule Name', 'Type', 'Origin', 'Owner', 'Created', 'Modified']
            + period_headers
            + ['Overall Status', 'Newly Triggered?', 'Recommendation', 'Description / Trigger Logic'])
    widths = ([9, 40, 14, 10, 14, 12, 12]
              + [16] * len(period_headers)
              + [20, 12, 46, 60])
    _write_sheet_header(ws, cols, widths)

    if master_df.empty:
        return

    status_col_idx = cols.index('Overall Status') + 1
    desc_col_idx   = cols.index('Description / Trigger Logic') + 1

    for _, r in master_df.iterrows():
        values = [int(r['rule_id']), str(r['rule_name']), str(r['rule_type']),
                  str(r['origin']), str(r['owner']), str(r['created']), str(r['modified'])]
        values += [int(r[f"offenses_{p['key']}"]) for p in LOOKBACK_PERIODS]
        values += [str(r['overall_status']), ('Yes' if r.get('newly_triggered') else 'No'),
                   str(r['recommendation']), str(r['description'])]
        ws.append(values)

        r_idx = ws.max_row
        status_cell = ws.cell(row=r_idx, column=status_col_idx)
        status_cell.font = _BOLD
        status_cell.fill = _FILLS[_status_fill_key(r['overall_status'])]

        ws.cell(row=r_idx, column=desc_col_idx).alignment = _WRAP

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"


def _write_filtered_sheet(wb, master_df, status_value, sheet_title, sort_col, ascending):
    ws = wb.create_sheet(sheet_title)

    cols = (['Rule ID', 'Rule Name', 'Type', 'Origin', 'Owner', 'Created']
            + [f"Offenses ({p['label']})" for p in LOOKBACK_PERIODS]
            + ['Recommendation', 'Description / Trigger Logic'])
    widths = [9, 42, 14, 10, 14, 12] + [16] * len(LOOKBACK_PERIODS) + [46, 60]
    _write_sheet_header(ws, cols, widths)

    if master_df.empty:
        ws.append(['No rule data available.'])
        return

    filtered = master_df[master_df['overall_status'] == status_value]
    if sort_col in filtered.columns:
        filtered = filtered.sort_values(sort_col, ascending=ascending)

    if filtered.empty:
        ws.append(['No rules currently fall into this category — nothing to review.'])
        return

    desc_col_idx = cols.index('Description / Trigger Logic') + 1

    for _, r in filtered.iterrows():
        values = [int(r['rule_id']), str(r['rule_name']), str(r['rule_type']),
                  str(r['origin']), str(r['owner']), str(r['created'])]
        values += [int(r[f"offenses_{p['key']}"]) for p in LOOKBACK_PERIODS]
        values += [str(r['recommendation']), str(r['description'])]
        ws.append(values)
        ws.cell(row=ws.max_row, column=desc_col_idx).alignment = _WRAP

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"


def _write_progress_sheet(wb, master_df, recovered_df, has_baseline):
    """
    Two stacked blocks on one sheet — deliberately not two more tabs, since
    both lists are usually short:
      · Newly Triggered — THE headline signal (matches the email exactly):
        fired only in the last NEWLY_TRIGGERED_WINDOW_DAYS days, nothing
        before that anywhere in the full lookback window. Needs nothing but
        this run's data.
      · Recovered Since Last Audit — a secondary, run-over-run comparison
        (needs the state file / has_baseline). Kept here for anyone who
        wants it; the email no longer surfaces it to keep that view to the
        one clear question you asked for.
    """
    ws = wb.create_sheet('Newly Triggered & Recovered')
    row = 1

    ws.cell(row=row, column=1,
            value=f'Newly Triggered — Last {NEWLY_TRIGGERED_WINDOW_DAYS} Days, Never Before').font = \
        Font(bold=True, size=12, color='065F46')
    row += 2
    newly = master_df[master_df['newly_triggered']] if 'newly_triggered' in master_df.columns else master_df.iloc[0:0]
    if newly.empty:
        ws.cell(row=row, column=1, value='No rules currently fall into this category.')
        ws.cell(row=row, column=1).font = Font(italic=True, size=10, color='7C6FA0')
        row += 2
    else:
        cols2 = ['Rule ID', 'Rule Name', 'Type', f'Offenses (last {NEWLY_TRIGGERED_WINDOW_DAYS}d)']
        for i, c in enumerate(cols2, start=1):
            cell = ws.cell(row=row, column=i, value=c)
            cell.fill, cell.font, cell.alignment = _FILLS['header'], _HDR_FONT, _CENTRE
        row += 1
        for _, r in newly.iterrows():
            ws.cell(row=row, column=1, value=int(r['rule_id']))
            ws.cell(row=row, column=2, value=str(r['rule_name']))
            ws.cell(row=row, column=3, value=str(r['rule_type']))
            ws.cell(row=row, column=4, value=int(r['offenses_newly_window']))
            row += 1
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Recovered Since Last Audit (secondary — run-over-run)').font = \
        Font(bold=True, size=12, color='7C6FA0')
    row += 2
    if not has_baseline:
        ws.cell(row=row, column=1,
                value='No previous audit on record yet — this run establishes the baseline '
                      'that future runs will compare against.')
        ws.cell(row=row, column=1).font = Font(italic=True, size=10, color='7C6FA0')
    elif recovered_df.empty:
        ws.cell(row=row, column=1, value='No rules have recovered since the last audit.')
        ws.cell(row=row, column=1).font = Font(italic=True, size=10, color='7C6FA0')
    else:
        cols = ['Rule ID', 'Rule Name', 'Type', 'Was', 'Now', 'Offenses (recent window)', 'Compared Against Run']
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=i, value=c)
            cell.fill, cell.font, cell.alignment = _FILLS['header'], _HDR_FONT, _CENTRE
        row += 1
        for _, r in recovered_df.iterrows():
            ws.cell(row=row, column=1, value=int(r['rule_id']))
            ws.cell(row=row, column=2, value=str(r['rule_name']))
            ws.cell(row=row, column=3, value=str(r['rule_type']))
            was_cell = ws.cell(row=row, column=4, value=str(r['previous_status']))
            was_cell.fill, was_cell.font = _FILLS['red'], _BOLD
            now_cell = ws.cell(row=row, column=5, value=str(r['current_status']))
            now_cell.fill, now_cell.font = _FILLS['green'], _BOLD
            ws.cell(row=row, column=6, value=int(r['offenses_recent']))
            ws.cell(row=row, column=7, value=str(r['previous_run']))
            row += 1

    for i, w in enumerate([10, 42, 14, 18, 18, 22, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_tested_rules_sheet(wb, tested_view_df, total_rules):
    """The human-maintained log, laid out as a small report rather than a
    plain table so 'how much have we covered' is visible at a glance."""
    ws = wb.create_sheet('Rules Tested & Triggered')
    investigated = len(tested_view_df)
    pct = round(investigated / total_rules * 100) if total_rules else 0

    ws['A1'] = 'Rules Tested & Triggered'
    ws['A1'].font = Font(bold=True, size=14, color='2D2257')
    ws['A2'] = f'{investigated} of {total_rules} rules manually validated so far ({pct}%)'
    ws['A2'].font = Font(bold=True, size=11, color='065F46')
    ws['A3'] = ('Maintained by hand — add a row to tested_rules_log.csv (or run '
                'mark_rule_tested.py) each time you manually confirm a rule fires as '
                'expected. Entries persist across runs and are de-duplicated by rule ID, '
                'so re-testing a rule updates its date rather than adding a second row.')
    ws['A3'].font = Font(italic=True, size=9, color='7C6FA0')
    ws['A3'].alignment = _WRAP
    ws.merge_cells('A3:E3')
    ws.row_dimensions[3].height = 32

    header_row = 5
    cols = ['Rule ID', 'Rule Name', 'Type', 'Tested Date', 'Notes']
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=i, value=c)
        cell.fill, cell.font, cell.alignment = _FILLS['header'], _HDR_FONT, _CENTRE
    for i, w in enumerate([10, 42, 14, 14, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if tested_view_df.empty:
        ws.cell(row=header_row + 1, column=1, value='No rules logged as tested yet.')
        ws.cell(row=header_row + 1, column=1).font = Font(italic=True, size=10, color='7C6FA0')
        return

    row = header_row
    for _, r in tested_view_df.iterrows():
        row += 1
        try:
            rid_val = int(r['rule_id'])
        except (TypeError, ValueError):
            rid_val = str(r['rule_id'])
        ws.cell(row=row, column=1, value=rid_val)
        ws.cell(row=row, column=2, value=str(r['rule_name']))
        ws.cell(row=row, column=3, value=str(r['rule_type']))
        ws.cell(row=row, column=4, value=str(r['tested_date']))
        ws.cell(row=row, column=5, value=str(r['notes'])).alignment = _WRAP

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(cols))}{ws.max_row}"


def save_excel_report(master_df, recovered_df, has_baseline, tested_view_df, path):
    """
    Saves the seven-sheet workbook:
      1. Executive Summary
      2. Rule Effectiveness (All Rules) — the sortable master table
      3. Dead Rules — never fired in 6 months, oldest first
      4. Recently Silent — fired historically, quiet in the last 90 days
      5. Highly Active Rules — firing well above typical volume
      6. Newly Triggered & Recovered — the automatic progress/insight sheet
      7. Rules Tested & Triggered — your manual triage log

    Returns the path actually written to (may differ from `path` if it was
    locked and we fell back to a timestamped filename), or None on total
    failure. A locked output file used to mean the whole run's analysis was
    thrown away — now it just lands next to the usual file instead.
    """
    try:
        wb = openpyxl.Workbook()

        if master_df.empty:
            ws = wb.active
            ws.title = 'No Data'
            ws['A1'] = 'No enabled rules or offense data were available to analyze.'
            ws['A1'].font = _BOLD
        else:
            _write_summary_sheet(wb, master_df)
            _write_master_sheet(wb, master_df)
            _write_filtered_sheet(wb, master_df, STATUS_DEAD, 'Dead Rules',
                                   sort_col='created', ascending=True)
            _write_filtered_sheet(wb, master_df, STATUS_RECENTLY_SILENT, 'Recently Silent',
                                   sort_col=f"offenses_{LOOKBACK_PERIODS[-1]['key']}", ascending=False)
            _write_filtered_sheet(wb, master_df, STATUS_HIGHLY_ACTIVE, 'Highly Active Rules',
                                   sort_col=f"offenses_{LOOKBACK_PERIODS[0]['key']}", ascending=False)
            _write_progress_sheet(wb, master_df, recovered_df, has_baseline)
            _write_tested_rules_sheet(wb, tested_view_df, len(master_df))

        try:
            wb.save(path)
            print(f"✅ Excel saved → {path}")
            return path
        except PermissionError:
            fallback = path.replace('.xlsx', f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            print(f"⚠️  '{path}' is open elsewhere — saving to '{fallback}' instead so this "
                  f"run isn't lost. Close the original file to have future runs overwrite it again.")
            wb.save(fallback)
            print(f"✅ Excel saved → {fallback}")
            return fallback

    except Exception as e:
        logger.error("Excel save failed:\n%s", traceback.format_exc())
        print(f"❌ Excel save error: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL HTML
# ══════════════════════════════════════════════════════════════════════════════

_C = {
    'purple':   '#9b72f5',
    'violet':   '#c4b5fd',
    'lavender': '#a78bfa',
    'dim':      '#7c6fa0',
    'green':    '#10b981',
    'red':      '#ef4444',
    'orange':   '#f97316',
    'amber':    '#f59e0b',
    'gray':     '#8b9ab0',
    'cyan':     '#06b6d4',
    'blue':     '#3b82f6',
    'badge_red':   '#7f1d1d',
    'badge_amber': '#78350f',
    'badge_green': '#065f46',
    'badge_gray':  '#2d3748',
    'badge_blue':  '#1e3a5f',
}
_CARD_BG = '#150f33'   # metric-card / bar-track background — one step lighter than the page bg


def _build_highly_active_table_html(master_df):
    """
    The email's ONE Highly Active section. Consolidated rather than
    repeated per window — the 1/3/6-month columns give the same comparison
    the old three separate sections did, without three near-identical
    tables. Ranked by the current (shortest-window) count since that's what
    "highly active right now" actually means.
    """
    C = _C
    if master_df.empty:
        return f'<p style="color:{C["dim"]};font-size:11px;font-family:monospace;padding:8px 0;">No rule data available.</p>'

    shortest_key = LOOKBACK_PERIODS[0]['key']
    hi = master_df[master_df[f'status_{shortest_key}'] == PERIOD_STATUS_HIGHLY_ACTIVE]
    if hi.empty:
        return (f'<p style="color:{C["green"]};font-size:11px;font-weight:700;font-family:monospace;'
                f'padding:8px 0;">✔ No rules are highly active right now.</p>')

    hi = hi.sort_values(f'offenses_{shortest_key}', ascending=False)
    rows_html = ''
    _rb = f'border-bottom:1px solid {C["dim"]}25;'
    for _, row in hi.iterrows():
        name = _html_escape(row['rule_name'])
        name_short = name[:40] + '…' if len(name) > 40 else name
        rtype = _html_escape(row['rule_type'])
        cells = ''.join(
            f'<td style="padding:7px 10px;{_rb}font-size:10px;color:{C["dim"]};'
            f'font-family:monospace;text-align:center;">{int(row[f"offenses_{p["key"]}"])}</td>'
            for p in LOOKBACK_PERIODS
        )
        rows_html += f"""
        <tr>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['violet']};font-family:monospace;" title="{name}">{name_short}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['dim']};font-family:monospace;text-align:center;">{rtype}</td>
          {cells}
        </tr>"""

    period_headers = ''.join(
        f'<th style="padding:6px 10px;text-align:center;font-size:9px;color:{C["blue"]};font-weight:700;'
        f'text-transform:uppercase;letter-spacing:1px;font-family:monospace;'
        f'border-top:2px solid {C["blue"]};border-bottom:1px solid {C["blue"]}50;">{p["label"]}</th>'
        for p in LOOKBACK_PERIODS
    )
    _hdr = f'border-top:2px solid {C["blue"]};border-bottom:1px solid {C["blue"]}50;'
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;margin-top:8px;">
      <thead><tr>
        <th style="padding:6px 10px;text-align:left;font-size:9px;color:{C['blue']};font-weight:700;
                   text-transform:uppercase;letter-spacing:1.2px;font-family:monospace;{_hdr}">Rule Name</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['blue']};font-weight:700;
                   text-transform:uppercase;letter-spacing:1.2px;font-family:monospace;{_hdr}">Type</th>
        {period_headers}
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>"""


def _build_newly_triggered_table_html(master_df):
    """
    The email's Progress section: ONLY rules that triggered in the last
    NEWLY_TRIGGERED_WINDOW_DAYS days and never before that (as far as the
    full lookback window lets us see). The run-over-run "recovered since
    last audit" comparison still runs, but lives in the Excel workbook only
    now, to keep this view to the one clear question you asked for.
    """
    C = _C
    if master_df.empty or 'newly_triggered' not in master_df.columns:
        return f'<p style="color:{C["dim"]};font-size:11px;font-family:monospace;padding:8px 0;">No rule data available.</p>'

    newly = master_df[master_df['newly_triggered']]
    if newly.empty:
        return (f'<p style="color:{C["green"]};font-size:11px;font-weight:700;font-family:monospace;'
                f'padding:8px 0;">✔ Nothing newly triggered in the last {NEWLY_TRIGGERED_WINDOW_DAYS} '
                f'day(s) — no movement to report this run.</p>')

    newly = newly.sort_values('offenses_newly_window', ascending=False)
    rows_html = ''
    _rb = f'border-bottom:1px solid {C["dim"]}25;'
    for _, row in newly.iterrows():
        name = _html_escape(row['rule_name'])
        name_short = name[:48] + '…' if len(name) > 48 else name
        rtype = _html_escape(row['rule_type'])
        rows_html += f"""
        <tr>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['violet']};font-family:monospace;" title="{name}">{name_short}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['dim']};font-family:monospace;text-align:center;">{rtype}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['green']};font-family:monospace;text-align:center;font-weight:700;">{int(row['offenses_newly_window'])}</td>
        </tr>"""

    _hdr = f'border-top:2px solid {C["green"]};border-bottom:1px solid {C["green"]}50;'
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;margin-top:8px;">
      <thead><tr>
        <th style="padding:6px 10px;text-align:left;font-size:9px;color:{C['green']};font-weight:700;
                   text-transform:uppercase;letter-spacing:1.4px;font-family:monospace;{_hdr}">Rule Name</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['green']};font-weight:700;
                   text-transform:uppercase;letter-spacing:1.4px;font-family:monospace;{_hdr}">Type</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['green']};font-weight:700;
                   text-transform:uppercase;letter-spacing:1.4px;font-family:monospace;{_hdr}">Offenses (last {NEWLY_TRIGGERED_WINDOW_DAYS}d)</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>"""


def _coverage_bar_html(current, total, color):
    """Email-safe progress bar — two <td>s with percentage widths, no CSS
    gradients or box-shadow, so it renders the same in Outlook as anywhere
    else."""
    C = _C
    total_safe = max(total, 1)
    pct = min(100, round(current / total_safe * 100))
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;margin-top:10px;">
      <tr>
        <td width="{pct}%" style="background:{color};height:7px;font-size:1px;line-height:1px;">&nbsp;</td>
        <td width="{100 - pct}%" style="background:{_CARD_BG};height:7px;font-size:1px;line-height:1px;">&nbsp;</td>
      </tr>
    </table>
    <div style="font-size:9px;color:{C['dim']};margin-top:4px;font-family:monospace;text-align:right;">
      {current} of {total} rules manually verified ({pct}%)
    </div>"""


def build_email_html(master_df, recovered_df, has_baseline, tested_view_df, chart_cid):
    """Assembles the full HTML email body."""
    C        = _C
    run_time = datetime.now().strftime('%d %b %Y  ·  %H:%M:%S')

    if master_df.empty:
        return f"""<!DOCTYPE html><html><body bgcolor="{_BG}" style="margin:0;padding:0;background-color:{_BG};font-family:monospace;">
        <table width="100%" bgcolor="{_BG}" style="background-color:{_BG};"><tr><td style="padding:24px;">
        <h2 style="color:{C['purple']};">QRadar Rule Effectiveness Report</h2>
        <p style="color:{C['dim']};">No enabled rules or offense data were available at
        {run_time}. Check the console output for connection or permission errors.</p>
        </td></tr></table></body></html>"""

    total_rules   = len(master_df)
    hi_1m_total   = int((master_df[f"status_{LOOKBACK_PERIODS[0]['key']}"] == PERIOD_STATUS_HIGHLY_ACTIVE).sum())
    newly_triggered_total = int(master_df['newly_triggered'].sum())
    investigated_total    = len(tested_view_df) if tested_view_df is not None else 0

    if newly_triggered_total > 0:
        hdr_bg, hdr_txt = C['badge_green'], f'✔  {newly_triggered_total} NEWLY TRIGGERED'
    elif hi_1m_total > 0:
        hdr_bg, hdr_txt = C['badge_blue'], f'▲  {hi_1m_total} HIGHLY ACTIVE'
    else:
        hdr_bg, hdr_txt = C['badge_green'], '✔  STEADY STATE'

    def badge(bg, txt):
        return (f'<span style="background:{bg};color:#f0eaff;font-size:10px;font-weight:700;'
                f'padding:4px 12px;border-radius:3px;letter-spacing:0.8px;'
                f'font-family:monospace;white-space:nowrap;">{txt}</span>')

    def metric_card(label, value, color, note=''):
        note_html = (f'<div style="font-size:9px;color:{C["dim"]};margin-top:4px;'
                     f'font-family:monospace;">{note}</div>') if note else ''
        return (f'<td width="25%" style="padding:4px;">'
                f'<table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">'
                f'<tr><td style="background:{_CARD_BG};border:1px solid {color}55;border-radius:10px;'
                f'padding:14px 8px;text-align:center;">'
                f'<div style="font-size:26px;font-weight:800;color:{color};line-height:1;'
                f'font-family:monospace;letter-spacing:-1px;">{value}</div>'
                f'<div style="font-size:9px;color:{C["dim"]};margin-top:6px;'
                f'text-transform:uppercase;letter-spacing:1px;">{label}</div>{note_html}'
                f'</td></tr></table></td>')

    headline_metrics = (
        metric_card('Total Rules', total_rules, C['purple'])
        + metric_card('Highly Active (1mo)', hi_1m_total, C['blue'])
        + metric_card(f'Newly Triggered ({NEWLY_TRIGGERED_WINDOW_DAYS}d)', newly_triggered_total, C['green'])
        + metric_card('Investigated Rules', investigated_total, C['cyan'])
    )
    coverage_bar = _coverage_bar_html(investigated_total, total_rules, C['cyan'])

    chart_html = (f'<img src="cid:{chart_cid}" alt="Rule status trend chart" '
                  f'style="display:block;max-width:100%;margin:14px auto 0;border-radius:8px;">') if chart_cid else ''
    progress_html      = _build_newly_triggered_table_html(master_df)
    highly_active_html = _build_highly_active_table_html(master_df)

    def section_header(icon, title, color, subtitle):
        return f"""
  <tr><td style="padding:24px 0 4px;border-top:2px solid {color}60;">
    <span style="font-size:13px;font-weight:700;color:{color};font-family:monospace;">{icon} {title}</span>
    <div style="font-size:10px;color:{C['dim']};margin-top:4px;">{subtitle}</div>
  </td></tr>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="color-scheme" content="dark">
<meta name="supported-color-schemes" content="dark">
</head>
<body bgcolor="{_BG}" style="margin:0;padding:0;background-color:{_BG};font-family:'Segoe UI',Helvetica,Arial,sans-serif;">
<table width="100%" bgcolor="{_BG}" cellpadding="0" cellspacing="0" style="background-color:{_BG};padding:24px 0;">
<tr><td align="center">
<table width="660" cellpadding="0" cellspacing="0" style="max-width:660px;width:100%;background-color:{_BG};">

  <!-- ══ MASTHEAD ══ -->
  <tr><td style="padding:4px 0 14px;border-bottom:3px solid {C['purple']};">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>
      <td>
        <div style="font-size:9px;color:{C['dim']};letter-spacing:3px;
                    text-transform:uppercase;font-family:monospace;margin-bottom:8px;">
          QRadar &nbsp;·&nbsp; Weekly Intelligence Report
        </div>
        <div style="font-size:24px;font-weight:800;color:{C['violet']};
                    letter-spacing:-0.5px;line-height:1.2;">
          Rule Effectiveness Auditor
        </div>
        <div style="margin-top:8px;font-size:11px;color:{C['dim']};font-family:monospace;">{run_time}</div>
      </td>
      <td align="right" valign="top">{badge(hdr_bg, hdr_txt)}</td>
    </tr></table>
  </td></tr>

  <!-- ══ HEADLINE METRICS ══ -->
  <tr><td style="padding:20px 0 4px;">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>{headline_metrics}</tr></table>
  </td></tr>
  <tr><td style="padding:4px 4px 8px;">{coverage_bar}</td></tr>

  <!-- ══ TREND CHART — unchanged 1/3/6-month comparison ══ -->
  <tr><td style="padding:16px 0 2px;">
    <span style="font-size:9px;color:{C['purple']};text-transform:uppercase;letter-spacing:2px;
                 font-family:monospace;font-weight:700;">1 / 3 / 6 Month Trend</span>
  </td></tr>
  <tr><td style="padding:0 0 8px;text-align:center;">{chart_html}</td></tr>

  {section_header('✔', 'Progress — Newly Triggered', C['green'],
                   f"Fired in the last {NEWLY_TRIGGERED_WINDOW_DAYS} day(s) and never before that, "
                   f"anywhere in the full lookback window.")}
  <tr><td style="padding:0 0 8px;">{progress_html}</td></tr>

  {section_header('▲', 'Highly Active Rules', C['blue'],
                   "Firing well above typical volume for their window right now — worth confirming "
                   "the volume is genuine and thresholds are tuned as intended.")}
  <tr><td style="padding:0 0 20px;">{highly_active_html}</td></tr>

  <!-- ══ FOOTER ══ -->
  <tr><td style="padding:16px 0 20px;border-top:1px solid {C['purple']}30;">
    <div style="font-size:9px;color:{C['dim']};font-family:monospace;letter-spacing:0.5px;line-height:1.6;">
      QRadar Rule Effectiveness Auditor &nbsp;·&nbsp; Auto-generated {run_time}<br>
      Dead &amp; Recently Silent rules, the full Highly Active history, and the tested-rules log
      are all in the attached Excel workbook — this email intentionally shows only what's moving.
    </div>
  </td></tr>

</table>
</td></tr>
</table>
</body></html>"""


# ─── OUTLOOK DRAFT ────────────────────────────────────────────────────────────

def create_outlook_draft(excel_path, subject, html_body, images, high_importance=False):
    """
    Creates an Outlook draft with embedded PNG chart(s) and Excel attachment.
    images = {'cid_key': '/path/to/file.png', ...}
    """
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail    = outlook.CreateItem(0)
        mail.Subject = subject
        if high_importance:
            mail.Importance = 2   # olImportanceHigh

        if excel_path and os.path.exists(excel_path):
            mail.Attachments.Add(excel_path)

        for cid, img_path in images.items():
            if img_path and os.path.exists(img_path):
                att = mail.Attachments.Add(img_path)
                att.PropertyAccessor.SetProperty(_MAPI_PR_ATTACH_CONTENT_ID, cid)

        mail.HTMLBody = html_body
        mail.Display()
        print("\n✉️  Outlook draft created.")

    except Exception as e:
        logger.error("Outlook draft failed:\n%s", traceback.format_exc())
        print(f"\n❌ Outlook draft error: {e}")

    finally:
        for cid, img_path in images.items():
            if img_path and os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception:
                    pass


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if not VERIFY_SSL:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print("=" * 62)
    print("  QRadar Rule Effectiveness Auditor")
    print("=" * 62)
    print(f"  Host            : {QRADAR_HOST}")
    print(f"  Lookback windows: {', '.join(p['label'] for p in LOOKBACK_PERIODS)}")
    print(f"  Newly-triggered : last {NEWLY_TRIGGERED_WINDOW_DAYS} day(s), never before")
    print(f"  Retry config    : {MAX_RETRIES} attempts, {RETRY_DELAY_BASE}s base backoff")
    print("=" * 62)

    if not validate_config():
        print("❌ Fix the configuration issues above before running.")
        return

    if not test_connection():
        return

    rules = fetch_all_rules()
    if not rules:
        print("\n❌ No enabled rules returned — check credentials, host URL, and permissions.")
        return

    max_days = max(p['days'] for p in LOOKBACK_PERIODS)
    offenses_all = fetch_offenses_for_lookback(max_days)
    if not offenses_all:
        print("\n⚠️  WARNING: no offenses were retrieved for the lookback window. This could "
              "mean the environment genuinely has none, but it could also mean the offense "
              "fetch failed silently — check the warnings above before treating every rule "
              "below as 'Dead'.")

    print("\n🔍 Building rule effectiveness table across all lookback windows...")
    master_df = build_master_rule_table(rules, offenses_all)

    if master_df.empty:
        print("\n❌ No rule data to report.")
        return

    for p in LOOKBACK_PERIODS:
        vc = master_df[f"status_{p['key']}"].value_counts()
        print(f"   {p['label']:>9}: Dead={vc.get(PERIOD_STATUS_DEAD, 0):<4} "
              f"Highly Active={vc.get(PERIOD_STATUS_HIGHLY_ACTIVE, 0):<4} "
              f"Active={vc.get(PERIOD_STATUS_ACTIVE, 0)}")

    silent_count = int((master_df['overall_status'] == STATUS_RECENTLY_SILENT).sum())
    print(f"\n   ⚠️  {silent_count} rule(s) fired historically but have gone quiet recently.")

    placeholder_desc = 'Not exposed via API'
    undocumented = int(master_df['description'].str.contains(placeholder_desc, na=False).sum())
    if undocumented:
        scope = "all" if undocumented == len(master_df) else f"{undocumented} of {len(master_df)}"
        print(f"\n   ℹ️  Trigger-logic descriptions are unavailable via the API for {scope} rule(s). "
              f"Run inspect_rule_fields.py once to confirm what your environment actually returns, "
              f"and consider populating {RULE_NOTES_OVERLAY_FILE} for your Dead / Recently Silent rules.")

    newly_triggered_count = int(master_df['newly_triggered'].sum())
    print(f"\n   ↑ {newly_triggered_count} rule(s) triggered in the last {NEWLY_TRIGGERED_WINDOW_DAYS} "
          f"day(s) and never before that.")

    # ── Run-history comparison (recovered since last audit — Excel only) ─────
    previous_state = load_previous_state() if ENABLE_RUN_HISTORY else None
    has_baseline   = previous_state is not None
    recovered_df   = find_recovered_rules(master_df, previous_state) if has_baseline else pd.DataFrame()

    if not ENABLE_RUN_HISTORY:
        print("   ℹ️  Run-history tracking is off (ENABLE_RUN_HISTORY=False) — "
              "'recovered since last audit' (Excel only) will be skipped this run.")
    elif not has_baseline:
        print("   ℹ️  No previous audit on record — this run establishes the baseline for "
              "the Excel workbook's 'recovered since last audit' comparison.")
    else:
        print(f"   ℹ️  {len(recovered_df)} rule(s) recovered since the last audit "
              f"(see the Excel workbook — this run's email focuses on Newly Triggered only).")

    if ENABLE_RUN_HISTORY:
        save_current_state(master_df, datetime.now().isoformat())

    # ── Manually-tested rules log ─────────────────────────────────────────────
    tested_entries  = load_tested_rules_log()
    tested_view_df  = build_tested_rules_view(master_df, tested_entries)
    investigated_count = len(tested_view_df)
    print(f"   📋 {investigated_count} of {len(master_df)} rules manually verified so far "
          f"(tested_rules_log.csv).")

    print(f"\n💾 Saving Excel report → {OUTPUT_EXCEL}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    saved_excel_path = save_excel_report(master_df, recovered_df, has_baseline, tested_view_df, OUTPUT_EXCEL)

    print("\n📊 Generating trend chart...")
    chart_path = generate_rule_trend_chart(master_df)
    if chart_path:
        print("   ✅ Trend chart generated.")
    else:
        print("   ⚠️  Trend chart skipped (no rule data).")

    print("\n✉️  Building email draft...")
    html_body = build_email_html(master_df, recovered_df, has_baseline, tested_view_df,
                                  chart_cid='rule_trend' if chart_path else None)

    dead_total = int((master_df['overall_status'] == STATUS_DEAD).sum())
    subject = (
        f"QRadar Rule Effectiveness — {newly_triggered_count} newly triggered, "
        f"{investigated_count} investigated, {dead_total} dead"
    )
    # Flag the email urgent only when there's a lot of fresh signal to look at —
    # not for the steady-state case, so "High" actually means something.
    high_importance = (newly_triggered_count + silent_count) > 10

    images = {'rule_trend': chart_path} if chart_path else {}
    create_outlook_draft(saved_excel_path, subject, html_body, images, high_importance=high_importance)
    print("\n✅ Done!")


if __name__ == '__main__':
    main()

import pandas as pd
import requests
import urllib3
from datetime import datetime, timedelta
import time
import os
import win32com.client
import openpyxl
from openpyxl.styles import PatternFill
import concurrent.futures

# Ensure charts generate in the background without opening UI windows
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt 

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
INPUT_EXCEL_PATH = r'C:\path\to\your\input.xlsx'
SHEETS_TO_PROCESS = ['Sheet1', 'Sheet2']  # or ['all'] for all sheets
LOGSOURCE_COLUMN = 'log source name'
IP_COLUMN = 'IP'
IN_QRADAR_COLUMN = 'In Qradar?'  # Defines if it should be scanned, skipped, or pended

QRADAR_HOST = 'https://your-qradar-host'
QRADAR_USERNAME = 'your-username'
QRADAR_PASSWORD = 'your-password'
VERIFY_SSL = False

DRAFT_OUTPUT_PATH = os.path.join(os.path.dirname(INPUT_EXCEL_PATH), 'inactive_and_errors.xlsx')
ACTIVITY_THRESHOLD_DAYS = 7  # Consider log source inactive if no events in X days
REQUEST_TIMEOUT = 30
MAX_WORKERS = 10  # Number of simultaneous API requests to QRadar

# Expected Log Source Types (Fuzzy Matching Applied)
# You only need to put the core keywords here now. 
# "Microsoft Security" will catch "Microsoft Windows Security Event Log".
EXPECTED_LS_TYPES = ['Microsoft Security', 'Linux OS']

# ─── END CONFIGURATION ─────────────────────────────────────────────────────────

# Valid timestamp range for QRadar API responses
MIN_TIMESTAMP = 0
MAX_TIMESTAMP = 2147483647

# Global Cache to prevent API DDOS when resolving Log Source Type IDs
LOG_SOURCE_TYPES_CACHE = {}


def test_qradar_connection(qradar_host, username, password):
    """
    Test QRadar connection and validate credentials before starting the heavy processing.
    """
    print("🔗 Testing QRadar connection...")
    qradar_host = qradar_host.rstrip('/')
    endpoint = f"{qradar_host}/api/help/versions"
    
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={
                'Accept': 'application/json', 
                'Version': '14.0'
            }
        )
        
        if resp.status_code == 200:
            print("✅ QRadar connection successful!")
            return True
        elif resp.status_code == 401:
            print("❌ Authentication failed! Check username/password.")
            return False
        else:
            print(f"⚠️ Unexpected response: {resp.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ Connection failed: {e}")
        return False


def fetch_log_source_types(qradar_host, username, password):
    """
    PRE-FETCH CACHE: Downloads the master dictionary of Log Source Type IDs 
    to their string Names so worker threads don't have to constantly query the API.
    """
    print("📥 Fetching Log Source Types Dictionary into memory...")
    qradar_host = qradar_host.rstrip('/')
    endpoint = f"{qradar_host}/api/config/event_sources/log_source_management/log_source_types"
    
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={
                'Accept': 'application/json', 
                'Version': '14.0'
            }
        )
        
        if resp.status_code == 200:
            types_data = resp.json()
            for t in types_data:
                ls_id = t.get('id')
                ls_name = t.get('name')
                if ls_id is not None and ls_name is not None:
                    LOG_SOURCE_TYPES_CACHE[ls_id] = ls_name
            print(f"✅ Successfully cached {len(LOG_SOURCE_TYPES_CACHE)} Log Source Types.")
        else:
            print(f"⚠️ Failed to fetch Log Source Types. API returned {resp.status_code}.")
            
    except Exception as e:
        print(f"❌ Error fetching Log Source Types: {e}")


def _empty_details():
    """
    Return an empty details structure to ensure dictionary keys 
    always exist even if the log source is not found.
    """
    return {
        'qradar_id': 'N/A',
        'enabled': 'Unknown',
        'last_seen': 'N/A',
        'activity_status': 'Not Found',
        'days_since_last_event': None,
        'actual_name': 'N/A',
        'ls_type': 'N/A',
        'is_older_expected': False  # Flag to trigger the RED highlight in Excel
    }


def safe_timestamp_conversion(timestamp_ms):
    """
    Safely convert the QRadar epoch timestamp to a human-readable datetime string
    and calculate how many days have passed since the last event.
    """
    if not timestamp_ms:
        return 'No events recorded', 'No Activity', None
    
    try:
        if isinstance(timestamp_ms, float):
            timestamp_ms = int(timestamp_ms)
        
        # Convert milliseconds to seconds if needed
        if timestamp_ms > 4102444800:
            timestamp_seconds = timestamp_ms / 1000.0
        else:
            timestamp_seconds = timestamp_ms
        
        # Validate timestamp boundaries
        if timestamp_seconds <= MIN_TIMESTAMP or timestamp_seconds > MAX_TIMESTAMP:
            return f'Invalid timestamp: {timestamp_ms}', 'Unknown', None
        
        last_event_datetime = datetime.fromtimestamp(timestamp_seconds)
        last_seen = last_event_datetime.strftime('%Y-%m-%d %H:%M:%S')
        
        time_diff = datetime.now() - last_event_datetime
        days_since_last_event = time_diff.days
        
        threshold_time = datetime.now() - timedelta(days=ACTIVITY_THRESHOLD_DAYS)
        
        # Determine activity status based on the defined threshold
        if last_event_datetime > threshold_time:
            activity_status = 'Active'
        else:
            activity_status = 'Inactive'
            
        return last_seen, activity_status, days_since_last_event
        
    except Exception as e:
        return f'Invalid timestamp: {timestamp_ms}', 'Unknown', None


def get_log_source_details(qradar_host, username, password, identifier, is_ip=False):
    """
    Get log source details directly from the QRadar API.
    Implements FUZZY MATCHING and ABSOLUTE TYPE-PRIORITY OVERRIDE to guarantee 
    expected log source types are chosen over anything else.
    """
    clean_identifier = str(identifier).replace('"', '').replace("'", "").strip()
    
    if is_ip:
        query_filter = f'protocol_parameters contains value="{clean_identifier}" or name ilike "%{clean_identifier}%"'
    else:
        query_filter = f'name ilike "%{clean_identifier}%"'
    
    ls_endpoint = f"{qradar_host.rstrip('/')}/api/config/event_sources/log_source_management/log_sources"

    try:
        resp = requests.get(
            ls_endpoint,
            params={'filter': query_filter},
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0'}
        )
        
        if resp.status_code != 200:
            return {'status': f'API Error {resp.status_code}', **_empty_details()}

        ls_data = resp.json()
        
        if not ls_data:
            return {'status': 'Not Found', **_empty_details()}

        # ─── 1. PRE-FILTER VALID MATCHES ───
        valid_sources = []
        if is_ip:
            for src in ls_data:
                params = src.get('protocol_parameters', [])
                in_params = any(p.get('value') == clean_identifier for p in params)
                in_name = clean_identifier.lower() in str(src.get('name', '')).lower()
                if in_params or in_name:
                    valid_sources.append(src)
        else:
            valid_sources = ls_data

        if not valid_sources:
            return {'status': 'Not Found', **_empty_details()}

        # ─── 2. FUZZY MATCH & TYPE SEPARATION ───
        expected_sources = []
        unexpected_sources = []
        
        for src in valid_sources:
            type_name = LOG_SOURCE_TYPES_CACHE.get(src.get('type_id'), "")
            api_name_clean = str(type_name).lower()
            
            is_match = False
            for exp in EXPECTED_LS_TYPES:
                # Split user config into words (e.g. "Microsoft Security" -> ["microsoft", "security"])
                exp_words = str(exp).lower().split()
                # If ALL keywords exist anywhere in the QRadar Type Name, it's a match!
                if all(w in api_name_clean for w in exp_words):
                    is_match = True
                    break
                    
            if is_match:
                expected_sources.append(src)
            else:
                unexpected_sources.append(src)

        # Helper to sort buckets: Enabled First -> Newest First
        def get_best_source(source_list):
            if not source_list: return None
            enabled = [s for s in source_list if s.get('enabled') is True]
            disabled = [s for s in source_list if s.get('enabled') is False]
            enabled.sort(key=lambda x: x.get('last_event_time') or 0, reverse=True)
            disabled.sort(key=lambda x: x.get('last_event_time') or 0, reverse=True)
            return enabled[0] if enabled else disabled[0]

        # ─── 3. ABSOLUTE TYPE-PRIORITY OVERRIDE ───
        found_source = None
        is_older_expected = False
        
        # Calculate the absolute max timestamp across ALL matched sources to determine if we passed up a newer one
        absolute_max_time = max([s.get('last_event_time') or 0 for s in valid_sources])
        
        if expected_sources:
            # NO MATTER WHAT: If an expected type exists, lock onto it.
            found_source = get_best_source(expected_sources)
            
            # Did we bypass a newer WinCollect to pick this older MSEL? Flag it RED.
            if (found_source.get('last_event_time') or 0) < absolute_max_time:
                is_older_expected = True
        else:
            # Fallback: Pick the best unexpected source (Will be flagged YELLOW in Excel)
            found_source = get_best_source(unexpected_sources)

        # ─── 4. EXTRACT FINAL DETAILS ───
        ls_id = found_source.get('id')
        ls_name = found_source.get('name', identifier)
        type_id = found_source.get('type_id')
        
        ls_type_name = LOG_SOURCE_TYPES_CACHE.get(type_id, f"Unknown Type ID: {type_id}")
        
        last_event_time_ms = found_source.get('last_event_time')
        last_seen, activity_status, days_since_last_event = safe_timestamp_conversion(last_event_time_ms)
        
        enabled_str = 'Yes' if found_source.get('enabled', False) else 'No'
            
        return {
            'status': 'Found',
            'qradar_id': str(ls_id) if ls_id is not None else '',
            'actual_name': ls_name,
            'ls_type': ls_type_name,
            'enabled': enabled_str,
            'last_seen': last_seen,
            'activity_status': activity_status,
            'days_since_last_event': days_since_last_event,
            'is_older_expected': is_older_expected
        }

    except Exception as e:
        return {'status': f'Error: {str(e)[:50]}...', **_empty_details()}


def process_single_row(idx, name_val, ip_val, qradar_host, username, password):
    """
    WORKER FUNCTION FOR THREADING: 
    This runs independently inside a thread pool.
    """
    if name_val and str(name_val).lower() in ['nan', 'none', '', 'null']: name_val = None
    if ip_val and str(ip_val).lower() in ['nan', 'none', '', 'null']: ip_val = None
    
    details = None
    search_method = "None"
    
    # 1. Primary Search by Name
    if name_val:
        details = get_log_source_details(qradar_host, username, password, name_val, is_ip=False)
        if details['status'] == 'Found':
            search_method = "Name"
    
    # 2. Fallback Search by IP 
    if (not details or details['status'] != 'Found') and ip_val:
        details = get_log_source_details(qradar_host, username, password, ip_val, is_ip=True)
        if details['status'] == 'Found':
            search_method = "IP"

    if not details:
        details = {'status': 'Empty/Invalid', **_empty_details()}

    return idx, name_val, details, search_method


def process_sheet(df, sheet_name, qradar_host, username, password, logsource_column, ip_column, in_qradar_col):
    """
    Optimized & Threaded Sheet Processing.
    Filters out 'Pending-Maintenance' natively before hitting the API thread pool.
    """
    print(f"\n{'='*60}")
    print(f"📋 Processing Sheet: {sheet_name} (Multi-threaded execution)")
    print(f"{'='*60}")
    
    if not df.empty:
        df.columns = df.columns.str.strip()

    if in_qradar_col not in df.columns:
        print(f"❌ ERROR: Column '{in_qradar_col}' not found. Skipping sheet.")
        return df

    # ─── HARD RESET / PRE-RUN CLEANSE ───
    cols_to_init = {
        'status': 'object', 
        'qradar_id': 'object', 
        'enabled': 'object',
        'last_seen': 'object', 
        'activity_status': 'object',
        'days_since_last_event': 'float64', 
        'remarks': 'object',
        'QRadar Actual Name': 'object',
        'Log Source Type': 'object',
        'Is Older Expected': 'bool'  # Hidden logic column for the Excel Highlighter
    }
    
    for col, dtype in cols_to_init.items():
        if col in df.columns:
            df[col] = None 
        else:
            df[col] = pd.Series(dtype=dtype)

    # ─── EFFICIENT FILTERING MASKS ───
    in_qradar_series = df[in_qradar_col].astype(str).str.lower()
    process_mask = in_qradar_series.str.contains("yes", na=False)
    pending_mask = in_qradar_series.str.contains("pending-maintenance", na=False)
    
    rows_to_process = df[process_mask]
    total_rows = len(df)
    target_count = len(rows_to_process)
    pending_count = pending_mask.sum()
    
    print(f"📊 Total Rows: {total_rows} | 🎯 'Yes' (To Scan): {target_count} | ⏳ Pending: {pending_count}")
    
    skipped_mask = ~(process_mask | pending_mask)
    df.loc[skipped_mask, 'remarks'] = "Skipped (Not Yes or Pending)"
    
    # Pre-fill Pending-Maintenance rows so they bypass the thread pool entirely
    if pending_count > 0:
        df.loc[pending_mask, 'status'] = 'Pending-Maintenance'
        df.loc[pending_mask, 'activity_status'] = 'Pending-Maintenance'
        df.loc[pending_mask, 'remarks'] = 'Pending Maintenance (Not Scanned)'
        df.loc[pending_mask, 'last_seen'] = 'N/A'
    
    if target_count == 0:
        return df

    # ─── MULTI-THREADING EXECUTION ───
    processed_count = 0
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        
        futures = {
            executor.submit(
                process_single_row, 
                idx, 
                str(row[logsource_column]).strip(), 
                str(row[ip_column]).strip(), 
                qradar_host, 
                username, 
                password
            ): idx for idx, row in rows_to_process.iterrows()
        }
        
        for future in concurrent.futures.as_completed(futures):
            processed_count += 1
            idx, name_val, details, search_method = future.result()
            
            print(f"\n🔹 [{processed_count}/{target_count}] Resolving: {name_val or 'Unknown'} -> {details['status']}")
            
            # Always map the audit columns
            df.at[idx, 'QRadar Actual Name'] = details['actual_name']
            df.at[idx, 'Log Source Type'] = details['ls_type']
            df.at[idx, 'Is Older Expected'] = details.get('is_older_expected', False)
            
            if details['status'] == 'Found':
                df.at[idx, 'qradar_id'] = details['qradar_id']
                df.at[idx, 'enabled'] = details['enabled']
                df.at[idx, 'last_seen'] = details['last_seen']
                df.at[idx, 'days_since_last_event'] = details['days_since_last_event']
                
                base_remark = f"Found by {search_method}"
                
                # Append a warning if we had to bypass a newer, unexpected source
                if details.get('is_older_expected'):
                    base_remark += " | ⚠️ Bypassed newer unexpected source"
                    print(f"      🚨 WARNING: Bypassed a newer unexpected log source to lock onto this expected one!")

                if details['enabled'] == 'No':
                    df.at[idx, 'status'] = 'Found'
                    df.at[idx, 'remarks'] = "Disabled on QRadar"
                    df.at[idx, 'activity_status'] = "Disabled"
                    print(f"      📌 Log Source: {details['actual_name']} [{details['ls_type']}]")
                    print(f"      ⚪ Status:     Disabled (Ignored for Inactivity)")
                else:
                    df.at[idx, 'status'] = 'Found'
                    df.at[idx, 'remarks'] = base_remark
                    df.at[idx, 'activity_status'] = details['activity_status']
                    print(f"      📌 Log Source: {details['actual_name']} [{details['ls_type']}]")
                    print(f"      📊 Activity:   {details['activity_status']}")
                    print(f"      📅 Last Event: {details['last_seen']} ({details['days_since_last_event']} days ago)")
                    
            else:
                status_val = details['status'] 
                remark_val = f"❌ {status_val}"
                
                if "Error" in status_val:
                    act_val = "Error"
                else:
                    act_val = "Not Found"
                
                # ─── INFERENCE LOGIC (WLC & FortiGate) ───
                if name_val and "AP" in name_val:
                    status_val = "Inferred"
                    remark_val = "Under WLC (Inferred)"
                    act_val = "Inferred"
                    print("      ℹ️  Result: Inferred as WLC (AP in name)")
                    
                elif name_val and "FW" in name_val:
                    status_val = "Inferred"
                    remark_val = "Under Forti (Inferred)"
                    act_val = "Inferred"
                    print("      ℹ️  Result: Inferred as FortiGate (FW in name)")
                    
                else:
                    print(f"      ❌ Result: {status_val.upper()}")

                df.at[idx, 'status'] = status_val
                df.at[idx, 'remarks'] = remark_val
                df.at[idx, 'activity_status'] = act_val
                df.at[idx, 'last_seen'] = "N/A"
                df.at[idx, 'days_since_last_event'] = None

    return df


def generate_pie_chart(data_dict, title, filename):
    """
    Generates a localized pie chart, saves it as a PNG, and returns the file path.
    """
    filtered_data = {k: v for k, v in data_dict.items() if v > 0}
    if not filtered_data: return None

    labels = list(filtered_data.keys())
    sizes = list(filtered_data.values())
    
    color_map = {
        'Active': '#28a745',                 # Green
        'Inactive': '#dc3545',               # Red
        'Not Found': '#6c757d',              # Grey
        'API Errors': '#fd7e14',             # Orange
        'Disabled': '#17a2b8',               # Teal
        'Inferred': '#6f42c1',               # Purple
        'Pending-Maintenance': '#007bff'     # Blue
    }
    
    colors = [color_map.get(label, '#cccccc') for label in labels]

    fig, ax = plt.subplots(figsize=(4.5, 3.5))
    ax.pie(
        sizes, labels=labels, colors=colors, autopct='%1.1f%%', 
        startangle=140, textprops={'fontsize': 9}, wedgeprops={'edgecolor': 'white'}
    )
    ax.axis('equal') 
    
    plt.title(title, pad=15, fontsize=11, fontweight='bold')
    
    filepath = os.path.join(os.path.dirname(INPUT_EXCEL_PATH), filename)
    plt.savefig(filepath, bbox_inches='tight', dpi=100)
    plt.close()
    
    return filepath


def create_html_outlook_draft(attachment_path, subject, html_body, image_paths):
    """
    Interfaces with Outlook via win32com to create an HTML draft,
    embeds images via Content-ID, and performs ephemeral file cleanup.
    """
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        
        if os.path.exists(attachment_path):
            mail.Attachments.Add(attachment_path)

        for cid, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                attachment = mail.Attachments.Add(img_path)
                attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
        
        mail.HTMLBody = html_body
        mail.Display()
        print(f"\n✉️  Email draft created successfully.")
        
        for cid, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                try: os.remove(img_path)
                except Exception as cleanup_err: print(f"⚠️ Could not delete temporary image {img_path}: {cleanup_err}")
                
    except Exception as e:
        print(f"\n❌ Failed to create Outlook draft: {e}")


def filter_and_email(processed_sheets_only, draft_path):
    """
    Calculates final SOC metrics across 7 distinct categories (including Pending-Maintenance), 
    generates charts, saves the strictly ACTIONABLE inventory as an Excel attachment, 
    and structures the HTML email.
    """
    report_frames = {}
    sheet_stats = {}
    images_to_embed = {}
    
    global_stats = { 
        'Active': 0, 'Inactive': 0, 'Not Found': 0, 
        'API Errors': 0, 'Disabled': 0, 'Inferred': 0,
        'Pending-Maintenance': 0
    }

    for name, df in processed_sheets_only.items():
        if 'status' not in df.columns: continue
        processed_df = df[df['status'].notna()].copy()
        if len(processed_df) == 0: continue

        active_count = len(processed_df[(processed_df['status'] == 'Found') & (processed_df['activity_status'] == 'Active')])
        mask_inactive = (processed_df['status'] == 'Found') & ((processed_df['activity_status'] == 'Inactive') | (processed_df['activity_status'] == 'No Activity'))
        inactive_count = mask_inactive.sum()
        disabled_count = len(processed_df[(processed_df['status'] == 'Found') & (processed_df['activity_status'] == 'Disabled')])
        inferred_count = len(processed_df[processed_df['status'] == 'Inferred'])
        mask_not_found = (processed_df['status'] == 'Not Found')
        not_found_count = mask_not_found.sum()
        mask_error = processed_df['status'].astype(str).str.startswith('API Error', na=False)
        error_count = mask_error.sum()
        pending_count = len(processed_df[processed_df['status'] == 'Pending-Maintenance'])

        sheet_counts = {
            'Active': active_count, 'Inactive': inactive_count, 'Not Found': not_found_count,
            'API Errors': error_count, 'Disabled': disabled_count, 'Inferred': inferred_count,
            'Pending-Maintenance': pending_count
        }
        
        sheet_stats[name] = sheet_counts
        for k in global_stats: global_stats[k] += sheet_counts[k]

        mask_report = mask_inactive | mask_not_found | mask_error
        
        if mask_report.any():
            sub = processed_df[mask_report].copy()
            sub.loc[mask_inactive, 'remarks'] = f'Inactive - No events in last {ACTIVITY_THRESHOLD_DAYS} days'
            report_frames[name] = sub 

    if not report_frames:
        print("✅ No Actionable Issues detected; skipping email.")
        return

    try:
        with pd.ExcelWriter(draft_path, engine='openpyxl') as writer:
            for sheet_name, df in report_frames.items():
                if 'Is Older Expected' in df.columns:
                    df = df.drop(columns=['Is Older Expected'])
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except PermissionError:
        print(f"❌ ERROR: Could not save report to '{draft_path}'. Is the file open?")
        return

    overall_cid = "overall_chart"
    overall_path = generate_pie_chart(global_stats, "Overall Inventory Status", "temp_overall.png")
    if overall_path: images_to_embed[overall_cid] = overall_path
    
    for name, counts in sheet_stats.items():
        cid = f"chart_{name.replace(' ', '_')}"
        chart_path = generate_pie_chart(counts, f"{name} Status", f"temp_{name}.png")
        if chart_path: images_to_embed[cid] = chart_path

    total_issues = global_stats['Inactive'] + global_stats['Not Found'] + global_stats['API Errors']

    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333;">
        <h2 style="color: #0056b3;">QRadar Action Report - {total_issues} Issues Require Attention</h2>
        <p>Attached is the automated QRadar log source status report generated on <b>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</b>.</p>
        
        <p>⚠️ <b>ACTION REQUIRED: {total_issues} ISSUES</b><br>
        (Count includes Inactive sources and Missing assets only. The attached Excel file contains <b>only the actionable items</b> that require your review.)</p>
        
        <table style="width: 100%; max-width: 600px; border-collapse: collapse; margin-bottom: 20px;">
            <tr style="background-color: #f8f9fa;">
                <td style="padding: 10px; border: 1px solid #dee2e6;"><b>Total Assets Scanned:</b></td>
                <td style="padding: 10px; border: 1px solid #dee2e6;">{sum(global_stats.values()) - global_stats['Pending-Maintenance']}</td>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #dee2e6; color: #dc3545;"><b>🔴 Actionable Issues:</b></td>
                <td style="padding: 10px; border: 1px solid #dee2e6;"><b>{total_issues}</b></td>
            </tr>
        </table>
        
        <h3>Overall System Health</h3>
    """
    
    if overall_cid in images_to_embed: html_body += f'<img src="cid:{overall_cid}"><br>'
    html_body += """<hr style="border: 1px solid #eee; margin: 30px 0;"><h3>Breakdown by Processed Sheets</h3>"""
    
    for name, counts in sheet_stats.items():
        cid = f"chart_{name.replace(' ', '_')}"
        html_body += f"""
        <div style="margin-bottom: 30px;">
            <h4 style="margin-bottom: 5px; color: #444;">{name}</h4>
            <ul style="list-style-type: none; padding-left: 0; margin-bottom: 10px;">
                <li><span style="color: #dc3545; font-weight: bold;">Inactive:</span> {counts['Inactive']}</li>
                <li><span style="color: #6c757d; font-weight: bold;">Not Found:</span> {counts['Not Found']}</li>
                <li><span style="color: #fd7e14; font-weight: bold;">API Errors:</span> {counts['API Errors']}</li>
                <li><span style="color: #28a745; font-weight: bold;">Active:</span> {counts['Active']}</li>
                <li><span style="color: #17a2b8; font-weight: bold;">Disabled:</span> {counts['Disabled']}</li>
                <li><span style="color: #6f42c1; font-weight: bold;">Inferred:</span> {counts['Inferred']}</li>
                <li><span style="color: #007bff; font-weight: bold;">Pending-Maintenance:</span> {counts['Pending-Maintenance']}</li>
            </ul>
        """
        if cid in images_to_embed: html_body += f'<img src="cid:{cid}">'
        html_body += "</div>"

    html_body += """<br><p style="font-size: 12px; color: #777;">Automated Cyber Defense Reporting</p></body></html>"""
    
    subject = f"QRadar Action Report - {total_issues} Issues Require Attention"
    create_html_outlook_draft(draft_path, subject, html_body, images_to_embed)


def save_surgical_updates_to_excel(filepath, processed_dataframes):
    """
    SURGICAL SAVE & HIGHLIGHTER: Opens the original workbook with openpyxl and updates 
    only the specific status columns. Evaluates conditional highlighting for expected/rogue Types.
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        
        red_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid') # Soft Red
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Bright Yellow
        
        cols_to_update = [
            'status', 'qradar_id', 'enabled', 'last_seen', 'activity_status', 
            'days_since_last_event', 'remarks', 'QRadar Actual Name', 'Log Source Type'
        ]
        
        for sheet_name, df in processed_dataframes.items():
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            header_row = 1
            col_map = {}
            
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_row, column=col_idx).value
                if cell_val is not None:
                    col_map[str(cell_val).strip()] = col_idx
            
            for c in cols_to_update:
                if c not in col_map:
                    new_col_idx = ws.max_column + 1
                    col_map[c] = new_col_idx
                    ws.cell(row=header_row, column=new_col_idx).value = c
            
            for idx, row in df.iterrows():
                excel_row = idx + 2  
                
                is_older_expected = row.get('Is Older Expected', False)
                
                for c in cols_to_update:
                    val = row[c]
                    if pd.isna(val):
                        val = ""
                        
                    target_cell = ws.cell(row=excel_row, column=col_map[c])
                    target_cell.value = val
                    
                    # ─── FUZZY MATCH HIGHLIGHTER LOGIC ───
                    if c == 'Log Source Type' and val != "" and val != "N/A":
                        
                        # Use the exact same fuzzy matching logic to determine yellow highlights
                        is_match = False
                        api_name_clean = str(val).lower()
                        for exp in EXPECTED_LS_TYPES:
                            exp_words = str(exp).lower().split()
                            if all(w in api_name_clean for w in exp_words):
                                is_match = True
                                break

                        if is_older_expected:
                            target_cell.fill = red_fill
                        elif not is_match:
                            target_cell.fill = yellow_fill
                    
        wb.save(filepath)
        print("✅ Original Excel file updated surgically (Formulas, original data, and Conditional Formatting applied!).")
        
    except PermissionError:
        print(f"❌ ERROR: Permission Denied! The file '{filepath}' is OPEN. Close it and re-run to save changes.")
    except Exception as e:
        print(f"❌ Failed to surgically save Excel file: {e}")


def main():
    if not VERIFY_SSL:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print("🚀 Starting QRadar Log Source Checker (Multi-threaded & Surgical)...")
    
    if not test_qradar_connection(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD):
        return
        
    # Pre-fetch the Log Source Types dictionary to speed up processing
    fetch_log_source_types(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD)

    print(f"\n📖 Reading Excel file: {INPUT_EXCEL_PATH}")
    try:
        all_sheets = pd.read_excel(INPUT_EXCEL_PATH, sheet_name=None)
    except Exception as e:
        print(f"❌ Failed to read Excel file: {e}")
        return

    if SHEETS_TO_PROCESS == ['all']:
        to_process = list(all_sheets.keys())
    else:
        to_process = SHEETS_TO_PROCESS
    
    for sheet in to_process:
        if sheet in all_sheets:
            all_sheets[sheet] = process_sheet(
                all_sheets[sheet], sheet, QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD,
                LOGSOURCE_COLUMN, IP_COLUMN, IN_QRADAR_COLUMN
            )

    print(f"\n💾 Saving updates to original Excel...")
    processed_sheets_only = {k: v for k, v in all_sheets.items() if k in to_process}
    
    save_surgical_updates_to_excel(INPUT_EXCEL_PATH, processed_sheets_only)
    filter_and_email(processed_sheets_only, DRAFT_OUTPUT_PATH)
    
    print(f"\n✅ Completed!")

if __name__ == '__main__':
    main()

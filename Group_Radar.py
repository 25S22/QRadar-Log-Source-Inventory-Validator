import re
import pandas as pd
import requests
import urllib3
from datetime import datetime, timedelta
import time
import os
import tempfile
import win32com.client
import openpyxl
from openpyxl.styles import PatternFill
import concurrent.futures
import logging
import traceback

# Ensure charts generate in the background without opening UI windows
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ─── LOGGING SETUP ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
INPUT_EXCEL_PATH    = r'C:\path\to\your\input.xlsx'
SHEETS_TO_PROCESS   = ['Sheet1', 'Sheet2']   # or ['all'] for all sheets
LOGSOURCE_COLUMN    = 'log source name'
IP_COLUMN           = 'IP'
IN_QRADAR_COLUMN    = 'In Qradar?'

QRADAR_HOST         = 'https://your-qradar-host'
QRADAR_USERNAME     = 'your-username'
QRADAR_PASSWORD     = 'your-password'
VERIFY_SSL          = False

DRAFT_OUTPUT_PATH        = os.path.join(os.path.dirname(INPUT_EXCEL_PATH), 'inactive_and_errors.xlsx')
ACTIVITY_THRESHOLD_DAYS  = 7
REQUEST_TIMEOUT          = 30
MAX_WORKERS              = 10

# ─── RETRY CONFIGURATION ───────────────────────────────────────────────────────
MAX_RETRIES      = 3     # total attempts (1 = no retry)
RETRY_DELAY_BASE = 1.5   # seconds — attempt 0→1.5 s, 1→3 s, 2→6 s

# ─── PAGINATION CONFIGURATION ──────────────────────────────────────────────────
LS_RANGE_MAX = 9999

EXPECTED_LS_TYPES = ['Microsoft Security', 'Linux OS']

# ─── GROUP THRESHOLD CONFIGURATION ────────────────────────────────────────────
# NOTE: These Excel-column-based overrides still apply as a fallback when
# the QRadar group description does NOT contain a parseable time value.
# Priority order: QRadar Group Description > GROUP_COLUMN > ACTIVITY_THRESHOLD_DAYS
GROUP_COLUMN     = None
GROUP_THRESHOLDS = {}

# ─── END CONFIGURATION ─────────────────────────────────────────────────────────

MIN_TIMESTAMP = 0
MAX_TIMESTAMP = 2147483647

# Cache populated once at startup; both are read-only during parallel scans.
LOG_SOURCE_TYPES_CACHE  = {}   # {type_id: type_name}
LOG_SOURCE_GROUPS_CACHE = {}   # {group_id: {'name': str, 'description': str}}

_MAPI_PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"

# ─── STATUS STRINGS THAT REPRESENT RETRIEVAL FAILURES ─────────────────────────
_ERROR_STATUS_PREFIXES = ('api error', 'timeout', 'connection error',
                          'worker error', 'error:')


# ─── GROUP THRESHOLD RESOLVER (Excel-column fallback) ─────────────────────────
def resolve_threshold(group_name=None):
    """
    Resolves threshold from the Excel GROUP_COLUMN mapping.
    This is the fallback used when QRadar group descriptions carry no
    parseable time value.  Priority: GROUP_THRESHOLDS dict > global default.
    """
    if not GROUP_COLUMN:
        return ACTIVITY_THRESHOLD_DAYS
    if not group_name or str(group_name).strip().lower() in ('nan', 'none', '', 'null'):
        return ACTIVITY_THRESHOLD_DAYS
    group_clean = str(group_name).strip().lower()
    for key, value in GROUP_THRESHOLDS.items():
        if str(key).strip().lower() == group_clean:
            return value
    return ACTIVITY_THRESHOLD_DAYS


# ─── DYNAMIC THRESHOLD PARSER ──────────────────────────────────────────────────
# Regex patterns ordered from finest to coarsest resolution so the first
# match wins against the most specific time unit present in the description.
_TIME_PATTERNS = [
    (r'(\d+(?:\.\d+)?)\s*hour',  1 / 24),   # hours  → fractional days
    (r'(\d+(?:\.\d+)?)\s*day',   1.0),       # days   → days
    (r'(\d+(?:\.\d+)?)\s*week',  7.0),       # weeks  → days
    (r'(\d+(?:\.\d+)?)\s*month', 30.0),      # months → days (approximate)
]


def parse_threshold_from_description(description):
    """
    Scans a QRadar group description for a human-readable time token and
    converts it to a threshold expressed in fractional days.

    Examples:
      "Inactivity window: 1 hour"   → 0.0417
      "Alert after 7 days silence"  → 7.0
      "Review every 2 weeks"        → 14.0
      "No SLA defined"              → None   ← falls back to global default

    Returns float (days) or None if no numeric time reference is found.
    """
    if not description:
        return None
    text = str(description).lower()
    for pattern, multiplier in _TIME_PATTERNS:
        m = re.search(pattern, text)
        if m:
            return float(m.group(1)) * multiplier
    return None


# ─── CHART STATS HELPER ────────────────────────────────────────────────────────
def _chart_stats(stats_dict):
    """
    Merges 'Inferred' into 'Active' and keeps 'Maintenance-Active' as its
    own visible wedge for chart display.
    """
    d = dict(stats_dict)
    d['Active'] = d.get('Active', 0) + d.pop('Inferred', 0)
    return d


def _is_error_status(status_str):
    s = str(status_str).strip().lower()
    return any(s.startswith(p) for p in _ERROR_STATUS_PREFIXES)


def _sanitise_identifier(raw):
    return (str(raw)
            .replace('\\', '')
            .replace('"', '')
            .replace("'", "")
            .replace('%', '')
            .replace('_', r'\_')
            .strip())


def test_qradar_connection(qradar_host, username, password):
    print("🔗 Testing QRadar connection...")
    qradar_host = qradar_host.rstrip('/')
    endpoint    = f"{qradar_host}/api/help/versions"
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0'}
        )
        if resp.status_code == 200:
            print("✅ QRadar connection successful!")
            return True
        elif resp.status_code == 401:
            print("❌ Authentication failed! Check username/password.")
            return False
        else:
            print(f"⚠️ Unexpected response: {resp.status_code}")
            return False
    except requests.exceptions.Timeout:
        print("❌ Connection timed out.")
        return False
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Connection error: {e}")
        return False
    except Exception as e:
        logger.error("Unexpected error during connection test:\n%s", traceback.format_exc())
        print(f"❌ Connection failed: {e}")
        return False


def fetch_log_source_types(qradar_host, username, password):
    print("📥 Fetching Log Source Types Dictionary into memory...")
    qradar_host = qradar_host.rstrip('/')
    endpoint    = (f"{qradar_host}/api/config/event_sources"
                   f"/log_source_management/log_source_types")
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0',
                     'Range': f'items=0-{LS_RANGE_MAX}'}
        )
        if resp.status_code == 200:
            for t in resp.json():
                ls_id   = t.get('id')
                ls_name = t.get('name')
                if ls_id is not None and ls_name is not None:
                    LOG_SOURCE_TYPES_CACHE[ls_id] = ls_name
            print(f"✅ Cached {len(LOG_SOURCE_TYPES_CACHE)} Log Source Types.")
        else:
            print(f"⚠️ Failed to fetch Log Source Types. API returned {resp.status_code}.")
    except requests.exceptions.Timeout:
        print("❌ Timed out while fetching Log Source Types.")
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Connection error while fetching Log Source Types: {e}")
    except Exception:
        logger.error("Unexpected error in fetch_log_source_types:\n%s", traceback.format_exc())
        print("❌ Error fetching Log Source Types.")


# ─── NEW: Log Source Groups fetch ─────────────────────────────────────────────
def fetch_log_source_groups(qradar_host, username, password):
    """
    Fetches all QRadar log source groups and caches them in
    LOG_SOURCE_GROUPS_CACHE keyed by group ID.

    The 'description' field is what we mine for dynamic threshold values
    (e.g. "Inactivity window: 7 days").  Groups without a description or
    without a parseable time token are cached as-is; the parser will simply
    return None for them and the global default will apply.
    """
    print("📥 Fetching Log Source Groups into memory...")
    qradar_host = qradar_host.rstrip('/')
    endpoint    = (f"{qradar_host}/api/config/event_sources"
                   f"/log_source_management/log_source_groups")
    try:
        resp = requests.get(
            endpoint,
            auth=(username, password),
            verify=VERIFY_SSL,
            timeout=REQUEST_TIMEOUT,
            headers={'Accept': 'application/json', 'Version': '14.0',
                     'Range': f'items=0-{LS_RANGE_MAX}'}
        )
        if resp.status_code == 200:
            for g in resp.json():
                gid = g.get('id')
                if gid is not None:
                    LOG_SOURCE_GROUPS_CACHE[gid] = {
                        'name':        g.get('name', '') or '',
                        'description': g.get('description', '') or '',
                    }
            print(f"✅ Cached {len(LOG_SOURCE_GROUPS_CACHE)} Log Source Groups.")
        else:
            print(f"⚠️ Failed to fetch Log Source Groups. API returned {resp.status_code}.")
    except requests.exceptions.Timeout:
        print("❌ Timed out while fetching Log Source Groups.")
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Connection error while fetching Log Source Groups: {e}")
    except Exception:
        logger.error("Unexpected error in fetch_log_source_groups:\n%s",
                     traceback.format_exc())
        print("❌ Error fetching Log Source Groups.")


def _empty_details():
    """
    Returns a fully-populated default dict so every caller can safely use
    dict.get() or direct key access without risking KeyError.
    """
    return {
        'qradar_id':            'N/A',
        'enabled':              'Unknown',
        'last_seen':            'N/A',
        'activity_status':      'Not Found',
        'days_since_last_event': None,
        'actual_name':          'N/A',
        'ls_type':              'N/A',
        'is_older_expected':    False,
        # ── Group fields (new) ────────────────────────────────────────────────
        'group_description':    'N/A',
        'dynamic_threshold':    None,
        'effective_threshold':  ACTIVITY_THRESHOLD_DAYS,
    }


def safe_timestamp_conversion(timestamp_ms, threshold=None):
    effective_threshold = threshold if threshold is not None else ACTIVITY_THRESHOLD_DAYS
    if not timestamp_ms:
        return 'No events recorded', 'No Activity', None
    try:
        if isinstance(timestamp_ms, float):
            timestamp_ms = int(timestamp_ms)
        if timestamp_ms > 4102444800:
            timestamp_seconds = timestamp_ms / 1000.0
        else:
            timestamp_seconds = timestamp_ms
        if timestamp_seconds <= MIN_TIMESTAMP or timestamp_seconds > MAX_TIMESTAMP:
            return f'Invalid timestamp: {timestamp_ms}', 'Unknown', None
        last_event_datetime   = datetime.fromtimestamp(timestamp_seconds)
        last_seen             = last_event_datetime.strftime('%Y-%m-%d %H:%M:%S')
        time_diff             = datetime.now() - last_event_datetime
        days_since_last_event = time_diff.days
        # timedelta() accepts fractional days via total_seconds — convert here.
        threshold_time        = datetime.now() - timedelta(days=effective_threshold)
        activity_status       = 'Active' if last_event_datetime > threshold_time else 'Inactive'
        return last_seen, activity_status, days_since_last_event
    except Exception:
        logger.error("Timestamp conversion error:\n%s", traceback.format_exc())
        return f'Invalid timestamp: {timestamp_ms}', 'Unknown', None


def get_log_source_details(qradar_host, username, password, identifier,
                           is_ip=False, threshold=None):
    """
    Queries the QRadar log-source management API for a single identifier.

    ═══════════════════════════════════════════════════════════════════════
    THRESHOLD RESOLUTION (priority order, highest to lowest):
      1. QRadar Group Description  — parsed from the group's description
                                     field at the API level (dynamic_threshold)
      2. Caller-supplied threshold — from GROUP_COLUMN / resolve_threshold()
      3. ACTIVITY_THRESHOLD_DAYS   — global fallback constant

    This means the team only has to put the SLA text in the QRadar group
    description once, and every log source in that group picks it up
    automatically — no Excel columns need to be maintained.
    ═══════════════════════════════════════════════════════════════════════

    ═══════════════════════════════════════════════════════════════════════
    ENABLED-SOURCE GUARANTEE (unchanged from prior version):
    A source is ONLY reported as "Disabled" when there is NO enabled source
    anywhere in the full result set for that identifier.
    Selection order:
      1. Enabled + expected type   → highest last_event_time wins
      2. Enabled + unexpected type → fallback when no enabled-expected exists
      3. Disabled (any type)       → only when zero enabled sources found
    ═══════════════════════════════════════════════════════════════════════
    """
    clean_identifier = _sanitise_identifier(identifier)

    if is_ip:
        query_filter = (
            f'protocol_parameters contains value="{clean_identifier}" '
            f'or name ilike "%{clean_identifier}%"'
        )
    else:
        query_filter = f'name ilike "%{clean_identifier}%"'

    ls_endpoint = (
        f"{qradar_host.rstrip('/')}/api/config/event_sources"
        f"/log_source_management/log_sources"
    )

    request_headers = {
        'Accept':  'application/json',
        'Version': '14.0',
        'Range':   f'items=0-{LS_RANGE_MAX}',
    }

    last_retriable_error = None

    for attempt in range(MAX_RETRIES):

        if attempt > 0:
            wait = RETRY_DELAY_BASE * (2 ** (attempt - 1))
            logger.warning(
                "Retry %d/%d for '%s' — waiting %.1fs after %s",
                attempt, MAX_RETRIES - 1, identifier, wait,
                type(last_retriable_error).__name__
            )
            time.sleep(wait)

        try:
            resp = requests.get(
                ls_endpoint,
                params={'filter': query_filter},
                auth=(username, password),
                verify=VERIFY_SSL,
                timeout=REQUEST_TIMEOUT,
                headers=request_headers
            )

            if resp.status_code != 200:
                return {'status': f'API Error {resp.status_code}', **_empty_details()}

            content_range = resp.headers.get('Content-Range', '')
            if content_range:
                try:
                    total_str = content_range.split('/')[-1].strip()
                    if total_str.isdigit() and int(total_str) > LS_RANGE_MAX + 1:
                        logger.warning(
                            "Pagination cap hit for '%s': QRadar reports %s total "
                            "matching sources but LS_RANGE_MAX is %d.  Raise "
                            "LS_RANGE_MAX or narrow EXPECTED_LS_TYPES.",
                            identifier, total_str, LS_RANGE_MAX
                        )
                except Exception:
                    pass

            ls_data = resp.json()
            if not ls_data:
                return {'status': 'Not Found', **_empty_details()}

            # ── Filter by IP / name ───────────────────────────────────────────
            valid_sources = []
            if is_ip:
                for src in ls_data:
                    params    = src.get('protocol_parameters', [])
                    in_params = any(p.get('value') == clean_identifier for p in params)
                    in_name   = clean_identifier.lower() in str(src.get('name', '')).lower()
                    if in_params or in_name:
                        valid_sources.append(src)
            else:
                valid_sources = ls_data

            if not valid_sources:
                return {'status': 'Not Found', **_empty_details()}

            # ── Classify each source as expected / unexpected ─────────────────
            def _is_expected_type(src):
                type_name      = LOG_SOURCE_TYPES_CACHE.get(src.get('type_id'), "")
                api_name_clean = str(type_name).lower()
                return any(
                    all(w in api_name_clean for w in str(exp).lower().split())
                    for exp in EXPECTED_LS_TYPES
                )

            # ── Split by enabled state FIRST, then by type ────────────────────
            enabled_expected    = []
            enabled_unexpected  = []
            disabled_expected   = []
            disabled_unexpected = []

            for src in valid_sources:
                is_en  = src.get('enabled') is True
                is_exp = _is_expected_type(src)
                if is_en and is_exp:
                    enabled_expected.append(src)
                elif is_en and not is_exp:
                    enabled_unexpected.append(src)
                elif not is_en and is_exp:
                    disabled_expected.append(src)
                else:
                    disabled_unexpected.append(src)

            def _best(lst):
                return max(lst, key=lambda x: x.get('last_event_time') or 0)

            is_older_expected = False

            if enabled_expected:
                found_source = _best(enabled_expected)
                if enabled_unexpected:
                    max_unexp_time = max(s.get('last_event_time') or 0
                                        for s in enabled_unexpected)
                    if (found_source.get('last_event_time') or 0) < max_unexp_time:
                        is_older_expected = True
            elif enabled_unexpected:
                found_source = _best(enabled_unexpected)
            elif disabled_expected:
                found_source = _best(disabled_expected)
            else:
                found_source = _best(disabled_unexpected)

            ls_id        = found_source.get('id')
            ls_name      = found_source.get('name', identifier)
            type_id      = found_source.get('type_id')
            ls_type_name = LOG_SOURCE_TYPES_CACHE.get(type_id, f"Unknown Type ID: {type_id}")

            # ── NEW: Resolve group description & dynamic threshold ─────────────
            # Each log source carries a list of group_ids.  We look every one up
            # in LOG_SOURCE_GROUPS_CACHE, build a human-readable summary string,
            # and attempt to parse a time-based threshold from the description.
            # The FIRST group whose description yields a numeric time value wins.
            group_ids         = found_source.get('group_ids') or []
            group_parts       = []   # display labels, one per group
            dynamic_threshold = None # float (days) from the first matching group

            for gid in group_ids:
                g_meta = LOG_SOURCE_GROUPS_CACHE.get(gid)
                if not g_meta:
                    group_parts.append(f"[Group ID {gid} — not in cache]")
                    continue

                g_name = g_meta['name']
                g_desc = g_meta['description']

                # Try to parse a time value only until we've found one
                parsed = None
                if dynamic_threshold is None:
                    parsed = parse_threshold_from_description(g_desc)
                    if parsed is not None:
                        dynamic_threshold = parsed

                # Build a readable label:  "GroupName (desc) → Xd threshold"
                label = g_name
                if g_desc:
                    label += f" — {g_desc}"
                if parsed is not None:
                    label += f" [→ {parsed:.4g}d threshold]"
                group_parts.append(label)

            group_description = " | ".join(group_parts) if group_parts else "N/A"

            # ── Resolve effective threshold (priority cascade) ─────────────────
            # 1. QRadar group description  (dynamic_threshold)
            # 2. Caller-supplied threshold (from Excel GROUP_COLUMN mapping)
            # 3. Global constant           (ACTIVITY_THRESHOLD_DAYS)
            if dynamic_threshold is not None:
                effective_threshold = dynamic_threshold
                threshold_source    = "QRadar Group"
            elif threshold is not None:
                effective_threshold = threshold
                threshold_source    = "GROUP_COLUMN" if GROUP_COLUMN else "Global Default"
            else:
                effective_threshold = ACTIVITY_THRESHOLD_DAYS
                threshold_source    = "Global Default"

            last_event_time_ms = found_source.get('last_event_time')
            last_seen, activity_status, days_since_last_event = safe_timestamp_conversion(
                last_event_time_ms, threshold=effective_threshold
            )

            enabled_str = 'Yes' if found_source.get('enabled', False) else 'No'

            total_enabled  = len(enabled_expected) + len(enabled_unexpected)
            total_disabled = len(disabled_expected) + len(disabled_unexpected)

            return {
                'status':                'Found',
                'qradar_id':             str(ls_id) if ls_id is not None else '',
                'actual_name':           ls_name,
                'ls_type':               ls_type_name,
                'enabled':               enabled_str,
                'last_seen':             last_seen,
                'activity_status':       activity_status,
                'days_since_last_event': days_since_last_event,
                'is_older_expected':     is_older_expected,
                'total_enabled':         total_enabled,
                'total_disabled':        total_disabled,
                # ── Group / threshold fields ───────────────────────────────────
                'group_description':    group_description,
                'dynamic_threshold':    dynamic_threshold,
                'effective_threshold':  effective_threshold,
                'threshold_source':     threshold_source,
            }

        except requests.exceptions.Timeout as exc:
            last_retriable_error = exc
            logger.warning("Timeout on attempt %d for '%s'", attempt + 1, identifier)

        except requests.exceptions.ConnectionError as exc:
            last_retriable_error = exc
            logger.warning("ConnectionError on attempt %d for '%s': %s",
                           attempt + 1, identifier, exc)

        except Exception as exc:
            logger.error("Unexpected error for identifier %s:\n%s",
                         identifier, traceback.format_exc())
            return {'status': f'Error: {str(exc)[:50]}', **_empty_details()}

    if isinstance(last_retriable_error, requests.exceptions.Timeout):
        logger.error("All %d attempts timed out for '%s'", MAX_RETRIES, identifier)
        return {'status': 'Timeout', **_empty_details()}
    else:
        logger.error("All %d attempts failed (ConnectionError) for '%s': %s",
                     MAX_RETRIES, identifier, last_retriable_error)
        return {'status': 'Connection Error', **_empty_details()}


def process_single_row(idx, name_val, ip_val, qradar_host, username, password,
                       group_val=None):
    if name_val and str(name_val).lower() in ['nan', 'none', '', 'null']: name_val = None
    if ip_val   and str(ip_val).lower()   in ['nan', 'none', '', 'null']: ip_val   = None

    # ── Resolve Excel-based threshold (GROUP_COLUMN fallback) ─────────────────
    # This value is passed into get_log_source_details() as the *fallback*
    # threshold.  If QRadar's own group description yields a parseable time,
    # that takes priority inside get_log_source_details() and this value is
    # effectively overridden there.
    excel_threshold = resolve_threshold(group_val)
    details         = None
    search_method   = "None"

    if name_val:
        details = get_log_source_details(
            qradar_host, username, password, name_val,
            is_ip=False, threshold=excel_threshold
        )
        if details['status'] == 'Found':
            search_method = "Name"

    if (not details or details['status'] != 'Found') and ip_val:
        details = get_log_source_details(
            qradar_host, username, password, ip_val,
            is_ip=True, threshold=excel_threshold
        )
        if details['status'] == 'Found':
            search_method = "IP"

    if not details:
        details = {'status': 'Empty/Invalid', **_empty_details()}

    return idx, name_val, details, search_method


def process_sheet(df, sheet_name, qradar_host, username, password,
                  logsource_column, ip_column, in_qradar_col):
    print(f"\n{'='*60}")
    print(f"📋 Processing Sheet: {sheet_name}")
    print(f"{'='*60}")

    if not df.empty:
        df.columns = df.columns.str.strip()

    required_columns = [in_qradar_col, logsource_column, ip_column]
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing columns {missing} in sheet '{sheet_name}'. Skipping.")
        return df

    cols_to_init = {
        'status':                'object',
        'qradar_id':             'object',
        'enabled':               'object',
        'last_seen':             'object',
        'activity_status':       'object',
        'days_since_last_event': 'float64',
        'remarks':               'object',
        'QRadar Actual Name':    'object',
        'Log Source Type':       'object',
        'Is Older Expected':     'bool',
        # ── NEW: Group Description column ─────────────────────────────────────
        # Populated from the QRadar group API; contains group name + description
        # text + the parsed threshold token so the analyst has full context.
        'Group Description':     'object',
    }
    for col, dtype in cols_to_init.items():
        if col in df.columns:
            df[col] = None
        else:
            df[col] = pd.Series(dtype=dtype)

    in_qradar_series = df[in_qradar_col].astype(str).str.lower()
    process_mask     = in_qradar_series.str.contains("yes",                 na=False)
    pending_mask     = in_qradar_series.str.contains("pending-maintenance", na=False)

    rows_to_process  = df[process_mask]
    rows_pending     = df[pending_mask]

    total_rows    = len(df)
    target_count  = len(rows_to_process)
    pending_count = len(rows_pending)

    group_feature_active = bool(GROUP_COLUMN and GROUP_COLUMN in df.columns)
    if GROUP_COLUMN and not group_feature_active:
        print(f"⚠️  GROUP_COLUMN '{GROUP_COLUMN}' not found — using global threshold.")

    print(f"📊 Total: {total_rows} | To Scan: {target_count} | "
          f"Pending (will also scan): {pending_count}")

    skipped_mask = ~(process_mask | pending_mask)
    df.loc[skipped_mask, 'remarks'] = "Skipped (Not Yes or Pending)"

    if target_count > 0:
        _run_scan(df, rows_to_process, target_count, qradar_host, username, password,
                  logsource_column, ip_column, in_qradar_col,
                  group_feature_active, is_pending=False)

    if pending_count > 0:
        print(f"\n  🔧 Scanning {pending_count} Pending-Maintenance source(s)...")
        _run_scan(df, rows_pending, pending_count, qradar_host, username, password,
                  logsource_column, ip_column, in_qradar_col,
                  group_feature_active, is_pending=True)

    return df


def _run_scan(df, rows_to_scan, total_count, qradar_host, username, password,
              logsource_column, ip_column, in_qradar_col,
              group_feature_active, is_pending=False):
    """
    Shared worker-dispatch loop used for both normal and pending rows.
    is_pending=True enables the maintenance-aware status path.
    """
    processed_count = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(
                process_single_row,
                idx,
                str(row[logsource_column]).strip(),
                str(row[ip_column]).strip(),
                qradar_host,
                username,
                password,
                str(row[GROUP_COLUMN]).strip() if group_feature_active else None
            ): idx for idx, row in rows_to_scan.iterrows()
        }

        for future in concurrent.futures.as_completed(futures):
            processed_count += 1
            try:
                idx, name_val, details, search_method = future.result()
            except Exception as worker_exc:
                original_idx = futures[future]
                logger.error("Worker crashed for row %s:\n%s", original_idx, traceback.format_exc())
                tag = "[PENDING]" if is_pending else ""
                print(f"\n⚠️  {tag}[{processed_count}/{total_count}] Worker crashed: {worker_exc}")
                df.at[original_idx, 'status']            = 'Worker Error'
                df.at[original_idx, 'remarks']           = f'Thread exception: {str(worker_exc)[:80]}'
                df.at[original_idx, 'activity_status']   = 'Error'
                df.at[original_idx, 'last_seen']         = 'N/A'
                df.at[original_idx, 'Group Description'] = 'N/A'
                continue

            _sep = '─' * 56
            tag  = ' [PENDING]' if is_pending else ''
            print(f"\n  {_sep}")
            print(f"  [{processed_count}/{total_count}]{tag}  {name_val or 'Unknown'}")
            print(f"  {_sep}")

            # ── Write QRadar-sourced metadata ──────────────────────────────────
            df.at[idx, 'QRadar Actual Name'] = details['actual_name']
            df.at[idx, 'Log Source Type']    = details['ls_type']
            df.at[idx, 'Is Older Expected']  = details.get('is_older_expected', False)
            # NEW: Group Description written to the dataframe here so
            # save_surgical_updates_to_excel() will persist it to Excel.
            df.at[idx, 'Group Description']  = details.get('group_description', 'N/A')

            if details['status'] == 'Found':
                df.at[idx, 'qradar_id']             = details['qradar_id']
                df.at[idx, 'enabled']               = details['enabled']
                df.at[idx, 'last_seen']             = details['last_seen']
                df.at[idx, 'days_since_last_event'] = details['days_since_last_event']

                t_en = details.get('total_enabled', '?')
                t_di = details.get('total_disabled', '?')
                print(f"  🔍 Match Via    : {search_method}")
                print(f"  📛 QRadar Name  : {details['actual_name']}")
                print(f"  🏷️  LS Type      : {details['ls_type']}")
                print(f"  🆔 QRadar ID    : {details['qradar_id']}")
                print(f"  📊 Sources      : {t_en} enabled / {t_di} disabled in result set")

                # ── NEW: Group / threshold terminal block ──────────────────────
                grp_desc  = details.get('group_description', 'N/A')
                eff_thr   = details.get('effective_threshold', ACTIVITY_THRESHOLD_DAYS)
                thr_src   = details.get('threshold_source', 'Global Default')
                dyn_thr   = details.get('dynamic_threshold')

                print(f"  🗂️  Group(s)     : {grp_desc}")
                if dyn_thr is not None:
                    # Convert fractional days back to a human label for readability
                    if eff_thr < 1:
                        thr_human = f"{eff_thr * 24:.4g}h ({eff_thr:.4g}d)"
                    else:
                        thr_human = f"{eff_thr:.4g}d"
                    print(f"  ⏱️  Threshold    : {thr_human}  ← parsed from {thr_src} ✅")
                else:
                    print(f"  ⏱️  Threshold    : {eff_thr:.4g}d  ← {thr_src} (no time in group desc)")

                if details.get('is_older_expected'):
                    print(f"  🚨 WARNING      : Bypassed a newer unexpected enabled log source!")

                base_remark = f"Found by {search_method}"
                if details.get('is_older_expected'):
                    base_remark += " | ⚠️ Bypassed newer unexpected source"

                if details['enabled'] == 'No':
                    if is_pending:
                        df.at[idx, 'status']          = 'Pending-Maintenance'
                        df.at[idx, 'activity_status'] = 'Pending-Maintenance'
                        df.at[idx, 'remarks']         = "Pending Maintenance — Disabled on QRadar"
                        print(f"  ⚪ Status       : DISABLED (maintenance — staying Pending)")
                    else:
                        df.at[idx, 'status']          = 'Found'
                        df.at[idx, 'remarks']         = "Disabled on QRadar (no enabled source found)"
                        df.at[idx, 'activity_status'] = "Disabled"
                        print(f"  ⚪ Status       : DISABLED (confirmed — no enabled source in result)")

                elif details['activity_status'] == 'No Activity':
                    if is_pending:
                        df.at[idx, 'status']          = 'Pending-Maintenance'
                        df.at[idx, 'activity_status'] = 'Pending-Maintenance'
                        df.at[idx, 'remarks']         = "Pending Maintenance — No events ever recorded"
                        print(f"  🟠 Activity     : NO ACTIVITY (maintenance — staying Pending)")
                    else:
                        df.at[idx, 'status']          = 'Found'
                        df.at[idx, 'remarks']         = "No events ever recorded on QRadar"
                        df.at[idx, 'activity_status'] = 'No Activity'
                        print(f"  🟠 Activity     : NO ACTIVITY (zero events ever recorded)")

                else:
                    act = details['activity_status']

                    if is_pending and act == 'Active':
                        df.at[idx, 'status']          = 'Found'
                        df.at[idx, 'activity_status'] = 'Maintenance-Active'
                        df.at[idx, 'remarks']         = (
                            f"🚨 MAINTENANCE SOURCE IS ACTIVE — {base_remark} — "
                            f"Last event {details['last_seen']} "
                            f"({details['days_since_last_event']}d ago). "
                            f"Update In Qradar? column to Yes!"
                        )
                        df.at[idx, in_qradar_col] = 'Yes'
                        print(f"  🚨 MAINTENANCE ALERT: Source is ACTIVE! "
                              f"In Qradar? updated to 'Yes'.")
                        print(f"  📅 Last Event   : {details['last_seen']}  "
                              f"({details['days_since_last_event']} days ago)")

                    elif is_pending:
                        df.at[idx, 'status']          = 'Pending-Maintenance'
                        df.at[idx, 'activity_status'] = 'Pending-Maintenance'
                        df.at[idx, 'remarks']         = (
                            f"Pending Maintenance — Found on QRadar ({act}) — {base_remark}"
                        )
                        print(f"  🔧 Maint Status : {act} — kept as Pending-Maintenance")
                        print(f"  📅 Last Event   : {details['last_seen']}  "
                              f"({details['days_since_last_event']} days ago)")

                    else:
                        df.at[idx, 'status']          = 'Found'
                        df.at[idx, 'remarks']         = base_remark
                        df.at[idx, 'activity_status'] = act
                        _icon = '✅' if act == 'Active' else '🔴'
                        print(f"  {_icon} Activity    : {act}")
                        print(f"  📅 Last Event   : {details['last_seen']}  "
                              f"({details['days_since_last_event']} days ago)")

            else:
                status_val = details['status']
                remark_val = f"❌ {status_val}"
                act_val    = "Error" if _is_error_status(status_val) else "Not Found"

                if name_val and "AP" in name_val:
                    status_val = "Inferred"
                    remark_val = "Under WLC (Inferred)"
                    act_val    = "Inferred"
                    print(f"  ℹ️  Status       : INFERRED — grouped under WLC")
                elif name_val and "FW" in name_val:
                    status_val = "Inferred"
                    remark_val = "Under Forti (Inferred)"
                    act_val    = "Inferred"
                    print(f"  ℹ️  Status       : INFERRED — grouped under FortiGate")
                else:
                    print(f"  ❌ Status       : {status_val.upper()}")

                # Still print group info even when not found — shows N/A clearly
                print(f"  🗂️  Group(s)     : {details.get('group_description', 'N/A')}")
                print(f"  ⏱️  Threshold    : {ACTIVITY_THRESHOLD_DAYS}d  ← Global Default")

                df.at[idx, 'status']                = status_val
                df.at[idx, 'remarks']               = remark_val
                df.at[idx, 'activity_status']       = act_val
                df.at[idx, 'last_seen']             = "N/A"
                df.at[idx, 'days_since_last_event'] = None

                if is_pending:
                    df.at[idx, 'status']          = 'Pending-Maintenance'
                    df.at[idx, 'activity_status'] = 'Pending-Maintenance'
                    df.at[idx, 'remarks']         = 'Pending Maintenance — Not found on QRadar'
                    df.at[idx, 'last_seen']       = 'N/A'


# ══════════════════════════════════════════════════════════════════════════════
#  VISUAL / AESTHETIC SECTION
# ══════════════════════════════════════════════════════════════════════════════

_C = {
    'purple':   '#9b72f5',
    'violet':   '#c4b5fd',
    'lavender': '#a78bfa',
    'dim':      '#7c6fa0',
    'green':    '#10b981',
    'red':      '#ef4444',
    'orange':   '#f97316',
    'amber':    '#f59e0b',
    'gray':     '#8b9ab0',
    'cyan':     '#06b6d4',
    'blue':     '#3b82f6',
    'magenta':  '#e879f9',
    'badge_red':     '#7f1d1d',
    'badge_gray':    '#2d3748',
    'badge_orange':  '#7c2d12',
    'badge_amber':   '#78350f',
    'badge_green':   '#065f46',
    'badge_cyan':    '#164e63',
    'badge_blue':    '#1e3a5f',
    'badge_magenta': '#6b21a8',
}

_STATUS_META = {
    'Inactive':           {'accent': _C['red'],     'badge_bg': _C['badge_red'],     'label': 'INACTIVE',       'icon': '●'},
    'No Activity':        {'accent': _C['orange'],  'badge_bg': _C['badge_orange'],  'label': 'NO ACTIVITY',    'icon': '◌'},
    'Not Found':          {'accent': _C['gray'],    'badge_bg': _C['badge_gray'],    'label': 'NOT FOUND',      'icon': '◌'},
    'Error':              {'accent': _C['amber'],   'badge_bg': _C['badge_amber'],   'label': 'API ERROR',      'icon': '▲'},
    'Maintenance-Active': {'accent': _C['magenta'], 'badge_bg': _C['badge_magenta'], 'label': 'MAINT ACTIVE 🚨','icon': '★'},
}

def _get_status_meta(activity_status):
    s = str(activity_status).strip()
    for key, meta in _STATUS_META.items():
        if key.lower() in s.lower():
            return meta
    return {'accent': _C['gray'], 'badge_bg': _C['badge_gray'],
            'label': s.upper()[:14], 'icon': '◌'}


def generate_pie_chart(data_dict, title, prefix='qradar_chart'):
    filtered_data = {k: v for k, v in data_dict.items() if v > 0}
    if not filtered_data:
        return None

    labels = list(filtered_data.keys())
    sizes  = list(filtered_data.values())

    color_map = {
        'Active':              '#10b981',
        'Inactive':            '#ef4444',
        'No Activity':         '#f97316',
        'Not Found':           '#6b7280',
        'API Errors':          '#f59e0b',
        'Disabled':            '#06b6d4',
        'Inferred':            '#8b5cf6',
        'Pending-Maintenance': '#3b82f6',
        'Maintenance-Active':  '#e879f9',
    }
    colors   = [color_map.get(lbl, '#a78bfa') for lbl in labels]
    bg_color = '#0a0618'

    fig, ax = plt.subplots(figsize=(5.2, 3.8), facecolor=bg_color)
    ax.set_facecolor(bg_color)

    wedges, _, autotexts = ax.pie(
        sizes,
        labels=None,
        colors=colors,
        autopct=lambda pct: f'{pct:.0f}%' if pct > 5 else '',
        startangle=140,
        wedgeprops={'edgecolor': bg_color, 'linewidth': 3, 'width': 0.54},
        pctdistance=0.76,
    )
    for at in autotexts:
        at.set_color('#f0eaff')
        at.set_fontweight('bold')
        at.set_fontsize(8)

    total = sum(sizes)
    ax.text(0, 0.10, str(total),  ha='center', va='center',
            fontsize=20, fontweight='bold', color='#c4b5fd')
    ax.text(0, -0.22, 'TOTAL', ha='center', va='center',
            fontsize=6, color='#6d5a9a', fontweight='700',
            fontfamily='monospace')

    legend_patches = [
        mpatches.Patch(color=color_map.get(lbl, '#a78bfa'), label=f'{lbl}  {v}')
        for lbl, v in zip(labels, sizes)
    ]
    leg = ax.legend(
        handles=legend_patches,
        loc='center left', bbox_to_anchor=(1.0, 0.5),
        fontsize=8, frameon=False,
        handlelength=1.1, handleheight=0.9,
        borderpad=0.5, labelspacing=0.7,
    )
    for t in leg.get_texts():
        t.set_color('#c4b5fd')

    ax.set_title(title, color='#a78bfa', fontsize=9, fontweight='700',
                 pad=10, fontfamily='monospace')
    ax.axis('equal')
    plt.tight_layout(pad=0.5)

    tmp      = tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix=f'{prefix}_')
    filepath = tmp.name
    tmp.close()
    plt.savefig(filepath, bbox_inches='tight', dpi=110,
                facecolor=bg_color, edgecolor='none')
    plt.close()
    return filepath


_SF  = "'Segoe UI', 'Segoe UI Historic', Helvetica, Arial, sans-serif"
_MON = "Consolas, 'Courier New', monospace"

_E = {
    'hdr_bg':     '#3b0764',
    'hdr_text':   '#f5f3ff',
    'hdr_sub':    '#c4b5fd',
    'card_bg':    '#ffffff',
    'card_border':'#e5e7eb',
    'row_alt':    '#faf9ff',
    'divider':    '#ede9fe',
    'txt_primary':'#1e1b4b',
    'txt_body':   '#374151',
    'txt_muted':  '#6b7280',
    'txt_mono':   '#4c1d95',
    'purple':     '#7c3aed',
    'purple_lt':  '#ede9fe',
    'green':      '#059669',
    'red':        '#dc2626',
    'orange':     '#ea580c',
    'amber':      '#d97706',
    'gray':       '#6b7280',
    'cyan':       '#0891b2',
    'blue':       '#2563eb',
    'magenta':    '#a21caf',
    'bdg_green':  '#065f46',
    'bdg_red':    '#7f1d1d',
    'bdg_orange': '#7c2d12',
    'bdg_amber':  '#78350f',
    'bdg_gray':   '#374151',
    'bdg_cyan':   '#164e63',
    'bdg_blue':   '#1e3a5f',
    'bdg_magenta':'#701a75',
    'bdg_purple': '#4c1d95',
}

_STATUS_META_LIGHT = {
    'Inactive':           {'badge_bg': _E['bdg_red'],    'label': 'INACTIVE',      'icon': '●'},
    'No Activity':        {'badge_bg': _E['bdg_orange'], 'label': 'NO ACTIVITY',   'icon': '◌'},
    'Not Found':          {'badge_bg': _E['bdg_gray'],   'label': 'NOT FOUND',     'icon': '◌'},
    'Error':              {'badge_bg': _E['bdg_amber'],  'label': 'API ERROR',     'icon': '▲'},
    'Maintenance-Active': {'badge_bg': _E['bdg_magenta'],'label': 'MAINT ACTIVE ★','icon': '★'},
}

def _get_status_meta_light(activity_status):
    s = str(activity_status).strip()
    for key, meta in _STATUS_META_LIGHT.items():
        if key.lower() in s.lower():
            return meta
    return {'badge_bg': _E['bdg_gray'], 'label': s.upper()[:14], 'icon': '◌'}


def create_html_outlook_draft(attachment_path, subject, html_body, image_paths):
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail    = outlook.CreateItem(0)
        mail.Subject = subject

        if os.path.exists(attachment_path):
            mail.Attachments.Add(attachment_path)

        for cid, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                attachment = mail.Attachments.Add(img_path)
                attachment.PropertyAccessor.SetProperty(_MAPI_PR_ATTACH_CONTENT_ID, cid)

        mail.HTMLBody = html_body
        mail.Display()
        print(f"\n✉️  Email draft created successfully.")

    except Exception as e:
        logger.error("Failed to create Outlook draft:\n%s", traceback.format_exc())
        print(f"\n❌ Failed to create Outlook draft: {e}")

    finally:
        for cid, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception as cleanup_err:
                    print(f"⚠️ Could not delete temp image {img_path}: {cleanup_err}")


def _build_actionable_table(report_df, logsource_col, ip_col):
    E = _E
    if report_df is None or len(report_df) == 0:
        return (
            f'<p style="color:{E["txt_muted"]};font-size:13px;font-style:italic;'
            f'font-family:{_SF};padding:10px 0 4px;margin:0;">'
            f'No actionable items for this sheet.</p>'
        )

    rows_html = ''
    for i, (_, row) in enumerate(report_df.iterrows()):
        hostname   = str(row.get(logsource_col, 'N/A') or 'N/A')
        ip_val     = str(row.get(ip_col,        'N/A') or 'N/A')
        qradar_id  = str(row.get('qradar_id',   'N/A') or 'N/A')
        last_seen  = str(row.get('last_seen',   'N/A') or 'N/A')
        act_status = str(row.get('activity_status', 'Unknown') or 'Unknown')

        meta   = _get_status_meta_light(act_status)
        row_bg = E['row_alt'] if i % 2 == 1 else E['card_bg']

        days = row.get('days_since_last_event')
        days_str = ''
        if pd.notna(days) and days is not None:
            try:
                d = int(days)
                days_str = (
                    f'<span style="color:{E["txt_muted"]};font-size:10px;'
                    f'display:block;margin-top:2px;font-family:{_SF};">'
                    f'{"today" if d == 0 else f"{d}d ago"}</span>'
                )
            except Exception:
                pass

        hostname_display = hostname if len(hostname) <= 40 else hostname[:38] + '…'

        rows_html += f"""
        <tr style="background:{row_bg};">
          <td style="padding:9px 12px;border-bottom:1px solid {E['divider']};
                     font-size:12px;color:{E['txt_mono']};font-family:{_MON};
                     max-width:210px;line-height:1.4;">
            <span title="{hostname}">{hostname_display}</span>
          </td>
          <td style="padding:9px 12px;border-bottom:1px solid {E['divider']};
                     font-size:12px;color:{E['txt_muted']};text-align:center;
                     white-space:nowrap;font-family:{_MON};">{ip_val}</td>
          <td style="padding:9px 12px;border-bottom:1px solid {E['divider']};
                     font-size:12px;color:{E['purple']};text-align:center;
                     white-space:nowrap;font-family:{_MON};font-weight:600;">{qradar_id}</td>
          <td style="padding:9px 12px;border-bottom:1px solid {E['divider']};
                     font-size:12px;color:{E['txt_body']};text-align:center;
                     white-space:nowrap;font-family:{_MON};">{last_seen}{days_str}</td>
          <td style="padding:9px 12px;border-bottom:1px solid {E['divider']};
                     text-align:center;">
            <span style="background:{meta['badge_bg']};color:#ffffff;
                         font-size:10px;font-weight:700;padding:3px 9px;
                         border-radius:3px;letter-spacing:0.3px;
                         white-space:nowrap;font-family:{_SF};">
              {meta['icon']}&nbsp;{meta['label']}
            </span>
          </td>
        </tr>"""

    th = (f'padding:8px 12px;font-size:10px;font-weight:700;'
          f'text-transform:uppercase;letter-spacing:0.8px;font-family:{_SF};'
          f'color:{E["txt_muted"]};background:{E["purple_lt"]};'
          f'border-top:2px solid {E["purple"]};border-bottom:1px solid {E["divider"]};')
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;margin-top:10px;
                  border:1px solid {E['divider']};border-radius:4px;">
      <thead>
        <tr>
          <th style="{th}text-align:left;">Hostname / Log Source</th>
          <th style="{th}text-align:center;">IP Address</th>
          <th style="{th}text-align:center;">QRadar ID</th>
          <th style="{th}text-align:center;">Last Event</th>
          <th style="{th}text-align:center;">Status</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>"""


def _build_email_html(global_stats, sheet_stats, total_issues,
                      images_to_embed, report_frames,
                      logsource_col, ip_col):
    E        = _E
    run_time = datetime.now().strftime('%d %b %Y  ·  %H:%M:%S')

    active_display = global_stats['Active'] + global_stats['Inferred']
    total_scanned  = sum(global_stats.values()) - global_stats['Pending-Maintenance']
    maint_active   = global_stats.get('Maintenance-Active', 0)

    if maint_active > 0:
        bdg_bg  = E['bdg_magenta']
        bdg_txt = f'🚨 {maint_active} MAINTENANCE ACTIVE'
    elif total_issues == 0:
        bdg_bg  = E['bdg_green']
        bdg_txt = '✔  ALL CLEAR'
    elif total_issues <= 10:
        bdg_bg  = E['bdg_amber']
        bdg_txt = f'⚠  {total_issues} ISSUES'
    else:
        bdg_bg  = E['bdg_red']
        bdg_txt = f'⚠  {total_issues} ISSUES'

    badge_html = (
        f'<span style="background:{bdg_bg};color:#ffffff;font-size:11px;'
        f'font-weight:700;padding:6px 16px;border-radius:20px;'
        f'letter-spacing:0.4px;font-family:{_SF};white-space:nowrap;">'
        f'{bdg_txt}</span>'
    )

    def metric_cell(label, value, color, note=''):
        note_html = (
            f'<div style="font-size:10px;color:{E["txt_muted"]};margin-top:3px;'
            f'font-family:{_SF};">{note}</div>'
        ) if note else ''
        return f"""
        <td style="padding:0 20px 0 0;text-align:center;vertical-align:top;">
          <div style="font-size:30px;font-weight:800;color:{color};
                      line-height:1;font-family:{_MON};">{value}</div>
          <div style="font-size:10px;color:{E['txt_muted']};margin-top:5px;
                      text-transform:uppercase;letter-spacing:0.8px;
                      font-family:{_SF};font-weight:600;">{label}</div>
          {note_html}
        </td>"""

    metric_row = (
        metric_cell('Active',       active_display,                       E['green'],   'incl. inferred') +
        metric_cell('Inactive',     global_stats['Inactive'],             E['red'])     +
        metric_cell('No Activity',  global_stats['No Activity'],          E['orange'],  'zero events ever') +
        metric_cell('Not Found',    global_stats['Not Found'],            E['gray'])    +
        metric_cell('API Errors',   global_stats['API Errors'],           E['amber'])   +
        metric_cell('Disabled',     global_stats['Disabled'],             E['cyan'])    +
        metric_cell('Maint Active', maint_active,                         E['magenta'], 'back online!') +
        metric_cell('Pending',      global_stats['Pending-Maintenance'],  E['blue'])
    )

    def stat_pill(label, value, color):
        if value == 0:
            return ''
        return (
            f'<span style="display:inline-block;background:{color}15;'
            f'border:1px solid {color}50;color:{color};'
            f'padding:3px 10px;border-radius:20px;margin:3px 5px 3px 0;'
            f'font-size:11px;font-family:{_SF};font-weight:700;">'
            f'{value} {label}</span>'
        )

    overall_chart_html = (
        f'<img src="cid:overall_chart" alt="Overall health chart" '
        f'style="display:block;max-width:420px;margin:12px auto 0;border-radius:8px;">'
    ) if 'overall_chart' in images_to_embed else ''

    maint_banner_html = ''
    if maint_active > 0:
        maint_banner_html = f"""
    <tr>
      <td style="padding:0 0 0 0;">
        <table width="100%" cellpadding="0" cellspacing="0">
          <tr>
            <td style="padding:14px 20px;background:#fdf4ff;
                       border-left:4px solid {E['magenta']};
                       border-top:1px solid {E['divider']};
                       border-bottom:1px solid {E['divider']};">
              <span style="font-size:13px;font-weight:700;color:{E['magenta']};
                           font-family:{_SF};">🚨 Maintenance Alert &nbsp;—&nbsp;</span>
              <span style="font-size:13px;color:{E['txt_body']};font-family:{_SF};">
                {maint_active} source{"s" if maint_active != 1 else ""} under maintenance
                {"are" if maint_active != 1 else "is"} now
                <strong>actively sending events</strong> to QRadar.
                The <em>In Qradar?</em> column has been updated to
                <strong>Yes</strong> in Excel. Please review and close the
                maintenance window.
              </span>
            </td>
          </tr>
        </table>
      </td>
    </tr>"""

    sheet_blocks = ''
    for sheet_name, counts in sheet_stats.items():
        cid         = f"chart_{sheet_name.replace(' ', '_')}"
        sheet_total = sum(counts.values()) - counts['Pending-Maintenance']
        sh_maint    = counts.get('Maintenance-Active', 0)
        issue_count = (counts['Inactive'] + counts['No Activity'] +
                       counts['Not Found'] + counts['API Errors'] + sh_maint)

        pills = (
            stat_pill('Active',       counts['Active'] + counts['Inferred'], E['green'])   +
            stat_pill('Inactive',     counts['Inactive'],                    E['red'])     +
            stat_pill('No Activity',  counts['No Activity'],                 E['orange'])  +
            stat_pill('Not Found',    counts['Not Found'],                   E['gray'])    +
            stat_pill('API Errors',   counts['API Errors'],                  E['amber'])   +
            stat_pill('Disabled',     counts['Disabled'],                    E['cyan'])    +
            stat_pill('Maint Active', sh_maint,                              E['magenta']) +
            stat_pill('Pending',      counts['Pending-Maintenance'],         E['blue'])
        )

        if sh_maint > 0:
            accent         = E['magenta']
            hdr_badge_html = (
                f'<span style="background:{E["bdg_magenta"]};color:#fff;'
                f'font-size:10px;font-weight:700;padding:3px 12px;border-radius:20px;'
                f'font-family:{_SF};">🚨 {sh_maint} MAINT ACTIVE</span>'
            )
        elif issue_count > 0:
            accent         = E['red']
            hdr_badge_html = (
                f'<span style="background:{E["bdg_red"]};color:#fff;'
                f'font-size:10px;font-weight:700;padding:3px 12px;border-radius:20px;'
                f'font-family:{_SF};">{issue_count} ISSUE{"S" if issue_count!=1 else ""}</span>'
            )
        else:
            accent         = E['green']
            hdr_badge_html = (
                f'<span style="color:{E["green"]};font-size:11px;font-weight:700;'
                f'font-family:{_SF};">✔ All Clear</span>'
            )

        chart_html = (
            f'<img src="cid:{cid}" alt="{sheet_name} chart" '
            f'style="display:block;max-width:340px;margin:14px auto 0;border-radius:8px;">'
        ) if cid in images_to_embed else ''

        actionable_html = _build_actionable_table(
            report_frames.get(sheet_name), logsource_col, ip_col
        )

        sheet_blocks += f"""
        <tr>
          <td style="padding:22px 0 0;">
            <table width="100%" cellpadding="0" cellspacing="0"
                   style="border-left:3px solid {accent};padding-left:14px;">
              <tr>
                <td>
                  <span style="font-size:14px;font-weight:700;color:{E['txt_primary']};
                               font-family:{_SF};">{sheet_name}</span>
                  <span style="font-size:11px;color:{E['txt_muted']};margin-left:10px;
                               font-family:{_SF};">{sheet_total} sources scanned</span>
                </td>
                <td align="right">{hdr_badge_html}</td>
              </tr>
            </table>
          </td>
        </tr>
        <tr>
          <td style="padding:10px 0 6px;border-bottom:1px solid {E['divider']};">
            {pills}
          </td>
        </tr>
        <tr><td style="padding:0;text-align:center;">{chart_html}</td></tr>
        <tr>
          <td style="padding:16px 0 4px;">
            <span style="font-size:11px;font-weight:700;color:{E['purple']};
                         text-transform:uppercase;letter-spacing:0.8px;
                         font-family:{_SF};">Requires Attention</span>
            <span style="font-size:11px;color:{E['txt_muted']};margin-left:8px;
                         font-family:{_SF};">
              Inactive · No Activity · Not Found · API Errors · Maintenance Active
              — full dataset in Excel
            </span>
          </td>
        </tr>
        <tr><td style="padding:0 0 28px;">{actionable_html}</td></tr>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
</head>
<body style="margin:0;padding:0;font-family:{_SF};">

<table width="100%" cellpadding="0" cellspacing="0" style="padding:20px 0;">
<tr>
  <td align="center" style="padding:0;">
  <table width="640" cellpadding="0" cellspacing="0"
         style="max-width:640px;width:100%;background:{E['card_bg']};
                border:1px solid {E['card_border']};border-radius:6px;
                border-collapse:separate;">

    <tr>
      <td style="padding:28px 28px 22px;background:{E['hdr_bg']};
                 border-radius:6px 6px 0 0;">
        <table width="100%" cellpadding="0" cellspacing="0"><tr>
          <td style="padding:0;">
            <div style="font-size:10px;color:{E['hdr_sub']};letter-spacing:3px;
                        text-transform:uppercase;font-family:{_SF};margin-bottom:10px;">
              QRadar &nbsp;·&nbsp; Inventory Validation
            </div>
            <div style="font-size:24px;font-weight:800;color:{E['hdr_text']};
                        line-height:1.2;font-family:{_SF};">
              Log Source Validation Report
            </div>
            <div style="margin-top:10px;font-size:12px;color:{E['hdr_sub']};
                        font-family:{_MON};">{run_time}</div>
          </td>
          <td align="right" valign="middle" style="padding-left:20px;
                                                    white-space:nowrap;">
            {badge_html}
          </td>
        </tr></table>
      </td>
    </tr>

    <tr>
      <td style="padding:14px 28px;background:{E['purple_lt']};
                 border-bottom:1px solid {E['divider']};">
        <span style="color:{E['red']};font-size:13px;font-weight:700;
                     font-family:{_SF};">ACTION REQUIRED &nbsp;·&nbsp;</span>
        <span style="color:{E['txt_body']};font-size:13px;font-family:{_SF};">
          {total_issues} source{"s" if total_issues != 1 else ""} need{"s" if total_issues == 1 else ""} attention.
          Full dataset in attached Excel.
        </span>
      </td>
    </tr>

    {maint_banner_html}

    <tr>
      <td style="padding:24px 28px 8px;">
        <div style="font-size:10px;color:{E['purple']};text-transform:uppercase;
                    letter-spacing:1.5px;margin-bottom:18px;font-family:{_SF};
                    font-weight:700;">
          Overall &nbsp;·&nbsp; {total_scanned} Sources Validated
        </div>
        <table cellpadding="0" cellspacing="0"><tr>{metric_row}</tr></table>
      </td>
    </tr>

    <tr>
      <td style="padding:8px 28px 20px;text-align:center;
                 border-bottom:1px solid {E['divider']};">
        <div style="font-size:10px;color:{E['txt_muted']};text-transform:uppercase;
                    letter-spacing:1px;margin-bottom:4px;font-family:{_SF};">
          Inventory Health Distribution
        </div>
        {overall_chart_html}
      </td>
    </tr>

    <tr>
      <td style="padding:22px 28px 2px;">
        <div style="font-size:10px;color:{E['purple']};text-transform:uppercase;
                    letter-spacing:1.5px;font-family:{_SF};font-weight:700;
                    margin-bottom:2px;">
          Breakdown by Sheet
        </div>
        <div style="font-size:11px;color:{E['txt_muted']};font-family:{_SF};">
          Metrics · chart · sources requiring action
        </div>
      </td>
    </tr>
    <tr>
      <td style="padding:0 28px 4px;">
        <table width="100%" cellpadding="0" cellspacing="0">
          {sheet_blocks}
        </table>
      </td>
    </tr>

    <tr>
      <td style="padding:16px 28px 20px;border-top:1px solid {E['divider']};
                 background:{E['purple_lt']};border-radius:0 0 6px 6px;">
        <div style="font-size:10px;color:{E['txt_muted']};font-family:{_SF};">
          QRadar Inventory Validation &nbsp;·&nbsp; Auto-generated {run_time}
        </div>
      </td>
    </tr>

  </table>
  </td>
</tr>
</table>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  END OF VISUAL SECTION
# ══════════════════════════════════════════════════════════════════════════════

def filter_and_email(processed_sheets_only, draft_path,
                     logsource_col=LOGSOURCE_COLUMN, ip_col=IP_COLUMN):
    report_frames   = {}
    sheet_stats     = {}
    images_to_embed = {}

    global_stats = {
        'Active': 0, 'Inactive': 0, 'No Activity': 0,
        'Not Found': 0, 'API Errors': 0, 'Disabled': 0,
        'Inferred': 0, 'Pending-Maintenance': 0,
        'Maintenance-Active': 0,
    }

    for name, df in processed_sheets_only.items():
        if 'status' not in df.columns:
            continue
        processed_df = df[df['status'].notna()].copy()
        if len(processed_df) == 0:
            continue

        active_count = len(processed_df[
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'Active')
        ])

        mask_inactive = (
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'Inactive')
        )
        inactive_count = mask_inactive.sum()

        mask_no_activity = (
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'No Activity')
        )
        no_activity_count = mask_no_activity.sum()

        disabled_count = len(processed_df[
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'Disabled')
        ])

        inferred_count  = len(processed_df[processed_df['status'] == 'Inferred'])
        mask_not_found  = processed_df['status'] == 'Not Found'
        not_found_count = mask_not_found.sum()
        mask_error      = processed_df['status'].apply(_is_error_status)
        error_count     = mask_error.sum()
        pending_count   = len(processed_df[processed_df['status'] == 'Pending-Maintenance'])

        mask_maint_active = (
            (processed_df['status'] == 'Found') &
            (processed_df['activity_status'] == 'Maintenance-Active')
        )
        maint_active_count = mask_maint_active.sum()

        sheet_counts = {
            'Active': active_count, 'Inactive': inactive_count,
            'No Activity': no_activity_count,
            'Not Found': not_found_count, 'API Errors': error_count,
            'Disabled': disabled_count, 'Inferred': inferred_count,
            'Pending-Maintenance': pending_count,
            'Maintenance-Active': maint_active_count,
        }

        sheet_stats[name] = sheet_counts
        for k in global_stats:
            global_stats[k] += sheet_counts.get(k, 0)

        mask_report = (mask_inactive | mask_no_activity |
                       mask_not_found | mask_error | mask_maint_active)

        if mask_report.any():
            sub = processed_df[mask_report].copy()

            for idx in sub[mask_inactive.loc[sub.index]].index:
                days = sub.at[idx, 'days_since_last_event']
                if pd.notna(days):
                    sub.at[idx, 'remarks'] = f'Inactive — No events in last {int(days)} days'
                else:
                    sub.at[idx, 'remarks'] = 'Inactive — No events recorded'

            for idx in sub[mask_no_activity.loc[sub.index]].index:
                sub.at[idx, 'remarks'] = 'No Activity — Zero events ever forwarded to QRadar'

            for idx in sub[mask_maint_active.loc[sub.index]].index:
                sub.at[idx, 'remarks'] = (
                    '🚨 MAINTENANCE SOURCE IS ACTIVE — update In Qradar? column'
                )

            report_frames[name] = sub

    if not report_frames:
        print("✅ No Actionable Issues detected; skipping email.")
        return

    try:
        with pd.ExcelWriter(draft_path, engine='openpyxl') as writer:
            for sheet_name, df in report_frames.items():
                out_df = df.copy()
                if 'Is Older Expected' in out_df.columns:
                    out_df = out_df.drop(columns=['Is Older Expected'])
                out_df.to_excel(writer, sheet_name=sheet_name, index=False)
    except PermissionError:
        print(f"❌ Could not save report to '{draft_path}'. Is the file open?")
        return

    overall_path = generate_pie_chart(
        _chart_stats(global_stats), "Overall Inventory Status",
        prefix='qradar_overall'
    )
    if overall_path:
        images_to_embed["overall_chart"] = overall_path

    for name, counts in sheet_stats.items():
        cid        = f"chart_{name.replace(' ', '_')}"
        chart_path = generate_pie_chart(
            _chart_stats(counts), f"{name} — Status",
            prefix=f'qradar_{name}'
        )
        if chart_path:
            images_to_embed[cid] = chart_path

    total_issues = (global_stats['Inactive'] +
                    global_stats['No Activity'] +
                    global_stats['Not Found'] +
                    global_stats['API Errors'] +
                    global_stats['Maintenance-Active'])

    html_body = _build_email_html(
        global_stats, sheet_stats, total_issues,
        images_to_embed, report_frames,
        logsource_col, ip_col
    )

    maint_count = global_stats.get('Maintenance-Active', 0)
    if maint_count > 0:
        subject = (f"🚨 QRadar — {maint_count} Maintenance Source"
                   f"{'s' if maint_count != 1 else ''} Now Active + "
                   f"{total_issues - maint_count} Other Issue"
                   f"{'s' if (total_issues - maint_count) != 1 else ''}")
    else:
        subject = (f"QRadar Inventory Validation — "
                   f"{total_issues} Source{'s' if total_issues != 1 else ''} Require Attention")

    create_html_outlook_draft(draft_path, subject, html_body, images_to_embed)


def save_surgical_updates_to_excel(filepath, processed_dataframes,
                                   in_qradar_col=IN_QRADAR_COLUMN):
    try:
        wb = openpyxl.load_workbook(filepath)

        red_fill     = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
        yellow_fill  = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        magenta_fill = PatternFill(start_color='F0ABFC', end_color='F0ABFC', fill_type='solid')

        cols_to_update = [
            'status', 'qradar_id', 'enabled', 'last_seen', 'activity_status',
            'days_since_last_event', 'remarks', 'QRadar Actual Name',
            'Log Source Type',
            # NEW: Group Description persisted back to Excel so the analyst can
            # see each source's QRadar group membership and the description text
            # that drove (or didn't drive) the dynamic threshold decision.
            'Group Description',
            in_qradar_col,
        ]

        for sheet_name, df in processed_dataframes.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws         = wb[sheet_name]
            header_row = 1
            col_map    = {}

            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_row, column=col_idx).value
                if cell_val is not None:
                    col_map[str(cell_val).strip()] = col_idx

            for c in cols_to_update:
                if c not in col_map:
                    new_col_idx = ws.max_column + 1
                    col_map[c]  = new_col_idx
                    ws.cell(row=header_row, column=new_col_idx).value = c

            row_index_map = {
                df_idx: excel_row
                for excel_row, df_idx in enumerate(df.index, start=2)
            }

            for idx, row in df.iterrows():
                excel_row         = row_index_map[idx]
                is_older_expected = row.get('Is Older Expected', False)
                is_maint_active   = str(row.get('activity_status', '')).strip() == 'Maintenance-Active'

                for c in cols_to_update:
                    if c not in df.columns:
                        continue
                    val = row[c]
                    if pd.isna(val):
                        val = ""

                    target_cell       = ws.cell(row=excel_row, column=col_map[c])
                    target_cell.value = val

                    if c == 'Log Source Type' and val not in ("", "N/A"):
                        is_match       = False
                        api_name_clean = str(val).lower()
                        for exp in EXPECTED_LS_TYPES:
                            exp_words = str(exp).lower().split()
                            if all(w in api_name_clean for w in exp_words):
                                is_match = True
                                break

                        if is_older_expected:
                            target_cell.fill = red_fill
                        elif not is_match:
                            target_cell.fill = yellow_fill

                    if c == in_qradar_col and is_maint_active:
                        target_cell.fill = magenta_fill

                    if c == 'activity_status' and is_maint_active:
                        target_cell.fill = magenta_fill

        wb.save(filepath)
        print("✅ Original Excel file updated surgically.")

    except PermissionError:
        print(f"❌ Permission Denied! '{filepath}' is open. Close it and re-run.")
    except Exception as e:
        logger.error("Failed to surgically save Excel:\n%s", traceback.format_exc())
        print(f"❌ Failed to surgically save Excel: {e}")


def main():
    if not VERIFY_SSL:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print("🚀 Starting QRadar Log Source Checker (Multi-threaded & Surgical)...")
    print(f"🔁 Retry config    : {MAX_RETRIES} attempts, "
          f"{RETRY_DELAY_BASE}s base backoff (exponential)")
    print(f"📄 Pagination range: items=0-{LS_RANGE_MAX} per API call")
    print(f"🔧 Pending-Maint   : SCANNED (sources going active will be flagged)")

    if GROUP_COLUMN:
        print(f"🏷️  Group Threshold: ENABLED  "
              f"(column: '{GROUP_COLUMN}', {len(GROUP_THRESHOLDS)} threshold(s))")
    else:
        print(f"🏷️  Group Threshold: DISABLED  (global: {ACTIVITY_THRESHOLD_DAYS}d)")

    print(f"🗂️  QRadar Groups  : ENABLED  "
          f"(descriptions parsed for dynamic thresholds — highest priority)")

    if not test_qradar_connection(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD):
        return

    fetch_log_source_types(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD)
    # NEW: Fetch groups immediately after types, before any scanning begins.
    # Both caches are read-only during the multi-threaded scan phase, so no
    # locking is required.
    fetch_log_source_groups(QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD)

    print(f"\n📖 Reading Excel: {INPUT_EXCEL_PATH}")
    try:
        all_sheets = pd.read_excel(INPUT_EXCEL_PATH, sheet_name=None)
    except Exception as e:
        logger.error("Failed to read Excel:\n%s", traceback.format_exc())
        print(f"❌ Failed to read Excel: {e}")
        return

    if [s.lower() for s in SHEETS_TO_PROCESS] == ['all']:
        to_process = list(all_sheets.keys())
    else:
        to_process = SHEETS_TO_PROCESS

    for sheet in to_process:
        if sheet not in all_sheets:
            print(f"⚠️  Sheet '{sheet}' not found. Available: {list(all_sheets.keys())}")

    for sheet in to_process:
        if sheet in all_sheets:
            all_sheets[sheet] = process_sheet(
                all_sheets[sheet], sheet,
                QRADAR_HOST, QRADAR_USERNAME, QRADAR_PASSWORD,
                LOGSOURCE_COLUMN, IP_COLUMN, IN_QRADAR_COLUMN
            )

    print(f"\n💾 Saving updates to original Excel...")
    processed_sheets_only = {k: v for k, v in all_sheets.items() if k in to_process}

    save_surgical_updates_to_excel(INPUT_EXCEL_PATH, processed_sheets_only,
                                   in_qradar_col=IN_QRADAR_COLUMN)
    filter_and_email(
        processed_sheets_only, DRAFT_OUTPUT_PATH,
        logsource_col=LOGSOURCE_COLUMN, ip_col=IP_COLUMN
    )

    print(f"\n✅ Completed!")


if __name__ == '__main__':
    main()

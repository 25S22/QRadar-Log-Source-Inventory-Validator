"""
QRadar EPS Burn Rate Monitor + Rule Effectiveness Auditor
=========================================================
Card 1  —  Ranks every log source by EPS consumption against your license cap.
Card 2  —  Cross-refs all enabled rules against 30-day offense history to find
           dead rules (0 contributions) and noise generators (too many).

Zero AQL.  All data from REST API:
  GET /api/config/event_sources/log_source_management/log_sources
  GET /api/config/event_sources/log_source_management/log_source_types
  GET /api/analytics/rules
  GET /api/siem/offenses

Output → two-sheet Excel  +  Outlook HTML email draft (same style as the
         log source validation script you already use).
"""

import os
import time
import tempfile
import traceback
import logging
from datetime import datetime, timedelta

import requests
import urllib3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import win32com.client

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ─── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  ←  edit this block only
# ══════════════════════════════════════════════════════════════════════════════

QRADAR_HOST     = 'https://your-qradar-host'
QRADAR_USERNAME = 'your-username'
QRADAR_PASSWORD = 'your-password'
VERIFY_SSL      = False          # locally hosted → safe to leave False

# ── EPS (Card 1) ──────────────────────────────────────────────────────────────
# Your licensed EPS ceiling.
# Find it in QRadar Admin → System Information → License or ask your IBM rep.
LICENSE_EPS_CAP = 5000           # example: 5 000 EPS

# A single source consuming more than this % of the license is flagged WARNING.
EPS_SOURCE_WARN_PCT = 10         # default: flag any source > 10 % of cap

# If the TOTAL EPS across all sources exceeds this % of cap → email badge = CRITICAL.
EPS_TOTAL_CRIT_PCT  = 80

# ── Rules (Card 2) ────────────────────────────────────────────────────────────
# How many days of offense history to pull for rule contribution counting.
OFFENSE_LOOKBACK_DAYS = 30

# Rule categories (based on offense contributions in the lookback window):
RULE_DEAD_THRESHOLD  = 0         # <= this  → Dead   (0 = never fired)
RULE_NOISE_THRESHOLD = 50        # >= this  → Noise generator

# ── Output ────────────────────────────────────────────────────────────────────
OUTPUT_DIR = r'C:\path\to\your\output'   # folder where Excel file is saved

# ── API / networking ──────────────────────────────────────────────────────────
REQUEST_TIMEOUT  = 30    # seconds per HTTP call
MAX_RETRIES      = 3     # attempts before giving up (exponential backoff)
RETRY_DELAY_BASE = 1.5   # seconds — waits: 1.5s → 3s → 6s
LS_RANGE_MAX     = 9999  # max items per paginated API call

# ══════════════════════════════════════════════════════════════════════════════
#  END CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

_MAPI_PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"

OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, 'qradar_eps_and_rules.xlsx')


# ─── SHARED HTTP HELPER ────────────────────────────────────────────────────────

def _api_get(path, params=None, label='request'):
    """
    Single GET helper used by every data-fetch function.

    Handles:
      · Exponential-backoff retry on Timeout / ConnectionError
      · Pagination cap via Range header (warns if QRadar has more items)
      · Clean propagation of 401 / other HTTP errors
    """
    url     = f"{QRADAR_HOST.rstrip('/')}{path}"
    headers = {
        'Accept':  'application/json',
        'Version': '14.0',
        'Range':   f'items=0-{LS_RANGE_MAX}',
    }
    last_err = None

    for attempt in range(MAX_RETRIES):
        if attempt > 0:
            wait = RETRY_DELAY_BASE * (2 ** (attempt - 1))
            logger.warning("Retry %d/%d for '%s' — waiting %.1fs after %s",
                           attempt, MAX_RETRIES - 1, label, wait,
                           type(last_err).__name__)
            time.sleep(wait)
        try:
            resp = requests.get(
                url,
                params=params,
                auth=(QRADAR_USERNAME, QRADAR_PASSWORD),
                verify=VERIFY_SSL,
                timeout=REQUEST_TIMEOUT,
                headers=headers,
            )
            if resp.status_code == 200:
                # ── Pagination overflow guard ─────────────────────────────
                cr = resp.headers.get('Content-Range', '')
                if cr:
                    try:
                        total = int(cr.split('/')[-1].strip())
                        if total > LS_RANGE_MAX + 1:
                            logger.warning(
                                "Pagination cap hit for '%s': %d items on server, "
                                "only %d fetched.  Raise LS_RANGE_MAX.",
                                label, total, LS_RANGE_MAX + 1
                            )
                    except Exception:
                        pass
                return resp.json()

            elif resp.status_code == 401:
                raise RuntimeError("Authentication failed — check QRADAR_USERNAME / PASSWORD")
            else:
                last_err = RuntimeError(f"HTTP {resp.status_code}")
                logger.warning("HTTP %d for '%s' (attempt %d/%d)",
                               resp.status_code, label, attempt + 1, MAX_RETRIES)

        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as exc:
            last_err = exc
            logger.warning("%s on attempt %d for '%s'",
                           type(exc).__name__, attempt + 1, label)
        except RuntimeError:
            raise
        except Exception:
            logger.error("Non-retriable error for '%s':\n%s", label, traceback.format_exc())
            raise

    raise RuntimeError(f"All {MAX_RETRIES} attempts failed for '{label}': {last_err}")


# ─── CONNECTION TEST ──────────────────────────────────────────────────────────

def test_connection():
    print("🔗 Testing QRadar connection...")
    try:
        result = _api_get('/api/help/versions', label='connection test')
        if result:
            print("✅ QRadar connection successful!")
            return True
        print("⚠️  Unexpected empty response from /api/help/versions")
        return False
    except RuntimeError as e:
        print(f"❌ {e}")
        return False
    except Exception as e:
        print(f"❌ Connection failed: {e}")
        return False


# ─── DATA FETCHERS ─────────────────────────────────────────────────────────────

def fetch_log_source_types():
    """Returns {type_id: type_name} dict."""
    print("📥 Fetching log source types...")
    try:
        data = _api_get(
            '/api/config/event_sources/log_source_management/log_source_types',
            label='log source types'
        )
        mapping = {t['id']: t['name'] for t in data if 'id' in t and 'name' in t}
        print(f"   ✅ {len(mapping)} types cached.")
        return mapping
    except Exception as e:
        print(f"   ⚠️  Could not fetch log source types: {e} — type names will show as ID.")
        return {}


def fetch_log_sources():
    """Returns list of log source objects (includes average_eps)."""
    print("📥 Fetching log sources with EPS data...")
    try:
        data = _api_get(
            '/api/config/event_sources/log_source_management/log_sources',
            label='log sources'
        )
        print(f"   ✅ {len(data)} log sources fetched.")
        return data
    except Exception as e:
        print(f"   ❌ Failed to fetch log sources: {e}")
        return []


def fetch_all_rules():
    """
    Returns list of correlation rule objects.
    Building blocks are excluded — they never generate offenses and would
    pollute the dead-rule analysis with hundreds of false positives.
    """
    print("📥 Fetching correlation rules...")
    try:
        # Pull all, then filter: enabled only, exclude building blocks
        data = _api_get(
            '/api/analytics/rules',
            label='analytics rules'
        )
        # Building blocks: QRadar marks them with type containing 'BB' or
        # with a 'building_block' flag.  We keep only rules where type does
        # not contain 'BB' and the rule has a proper type string.
        rules = [
            r for r in data
            if r.get('enabled', False) is True
            and 'BB' not in str(r.get('type', '')).upper()
            and 'BUILDING_BLOCK' not in str(r.get('type', '')).upper()
        ]
        print(f"   ✅ {len(rules)} enabled correlation rules (of {len(data)} total).")
        return rules
    except Exception as e:
        print(f"   ❌ Failed to fetch rules: {e}")
        return []


def fetch_recent_offenses():
    """
    Returns list of offense objects from the last OFFENSE_LOOKBACK_DAYS days.
    Uses start_time filter so we don't pull the entire offense history.
    Each offense object contains a 'rules' list showing contributing rule IDs.
    """
    print(f"📥 Fetching offenses from last {OFFENSE_LOOKBACK_DAYS} days...")
    cutoff_ms   = int(
        (datetime.now() - timedelta(days=OFFENSE_LOOKBACK_DAYS)).timestamp() * 1000
    )
    filter_str  = f'start_time >= {cutoff_ms}'
    # Request only the fields we need — reduces payload on large deployments
    fields_str  = 'id,rules,start_time,magnitude,status,severity'
    try:
        data = _api_get(
            '/api/siem/offenses',
            params={'filter': filter_str, 'fields': fields_str},
            label='recent offenses'
        )
        print(f"   ✅ {len(data)} offenses fetched.")
        return data
    except Exception as e:
        print(f"   ⚠️  Filtered offense fetch failed ({e}). Trying unfiltered...")
        try:
            data = _api_get('/api/siem/offenses', label='offenses (unfiltered)')
            # Filter in Python as fallback
            cutoff_s  = cutoff_ms / 1000
            data      = [o for o in data
                         if (o.get('start_time') or 0) >= cutoff_ms]
            print(f"   ✅ {len(data)} offenses after in-memory filter.")
            return data
        except Exception as e2:
            print(f"   ❌ Offense fetch failed entirely: {e2} — rule analysis will be empty.")
            return []


# ─── EPS ANALYSIS ─────────────────────────────────────────────────────────────

def _get_eps(source):
    """
    Reads average_eps from a log source object.
    Different QRadar versions nest this field differently — tries both locations.
    """
    # Direct field (most common)
    val = source.get('average_eps')
    if val is None:
        # Nested inside 'status' object on some versions
        status = source.get('status', {})
        if isinstance(status, dict):
            val = status.get('average_eps')
    try:
        return max(0.0, float(val)) if val is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def analyze_eps(log_sources, type_map):
    """
    Builds a ranked DataFrame of log sources by EPS consumption.
    Columns: rank, name, ls_type, enabled, average_eps, pct_of_total,
             pct_of_license, eps_status
    """
    rows = []
    for src in log_sources:
        eps = _get_eps(src)
        type_id   = src.get('type_id')
        rows.append({
            'name':        src.get('name', 'Unknown'),
            'ls_type':     type_map.get(type_id, f"Type {type_id}"),
            'enabled':     'Yes' if src.get('enabled') else 'No',
            'average_eps': eps,
            'qradar_id':   src.get('id', ''),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    total_eps = df['average_eps'].sum()

    df['pct_of_total']   = (
        df['average_eps'] / total_eps * 100 if total_eps > 0 else 0.0
    )
    df['pct_of_license'] = (
        df['average_eps'] / LICENSE_EPS_CAP * 100 if LICENSE_EPS_CAP > 0 else 0.0
    )

    # ── Status label ──────────────────────────────────────────────────────────
    def _eps_status(row):
        if row['enabled'] == 'No':
            return 'Disabled'
        if row['pct_of_license'] >= EPS_SOURCE_WARN_PCT:
            return 'Warning'
        if row['average_eps'] == 0:
            return 'Idle'
        return 'OK'

    df['eps_status'] = df.apply(_eps_status, axis=1)

    # Sort: highest EPS first
    df = df.sort_values('average_eps', ascending=False).reset_index(drop=True)
    df.insert(0, 'rank', df.index + 1)
    df['average_eps']    = df['average_eps'].round(2)
    df['pct_of_total']   = df['pct_of_total'].round(2)
    df['pct_of_license'] = df['pct_of_license'].round(2)
    return df


# ─── RULE ANALYSIS ────────────────────────────────────────────────────────────

def _extract_rule_ids(offense):
    """
    Extracts contributing rule IDs from an offense object.
    Handles: list of dicts ({'id': X, 'type': Y}), list of ints, or missing field.
    """
    ids = []
    for r in offense.get('rules', []):
        if isinstance(r, dict):
            rid = r.get('id')
        elif isinstance(r, (int, str)):
            rid = r
        else:
            rid = None
        if rid is not None:
            try:
                ids.append(int(rid))
            except (TypeError, ValueError):
                pass
    return ids


def analyze_rules(rules, offenses):
    """
    Builds a DataFrame of all enabled correlation rules annotated with:
      - offense_count: how many offenses in the lookback window they contributed to
      - category: Dead / Active / Noise Generator
      - recommendation: human-readable action
    """
    # ── Build rule-ID → offense count map ─────────────────────────────────────
    rule_offense_counts = {}
    for offense in offenses:
        for rid in _extract_rule_ids(offense):
            rule_offense_counts[rid] = rule_offense_counts.get(rid, 0) + 1

    rows = []
    for rule in rules:
        rid           = rule.get('id')
        offense_count = rule_offense_counts.get(rid, 0)
        origin        = rule.get('origin', 'UNKNOWN')
        rule_type     = rule.get('type', 'UNKNOWN')
        name          = rule.get('name', f"Rule {rid}")

        # ── Category ──────────────────────────────────────────────────────────
        if offense_count <= RULE_DEAD_THRESHOLD:
            category = 'Dead'
            recommendation = (
                f"Zero offense contributions in {OFFENSE_LOOKBACK_DAYS}d. "
                "Review — disable if no longer required to reduce rule engine load."
            )
        elif offense_count >= RULE_NOISE_THRESHOLD:
            category = 'Noise Generator'
            recommendation = (
                f"{offense_count} offenses in {OFFENSE_LOOKBACK_DAYS}d. "
                "Review thresholds — analysts may be ignoring or bulk-closing these."
            )
        else:
            category = 'Active'
            recommendation = 'Rule is contributing normally. No action required.'

        rows.append({
            'rule_id':        rid,
            'rule_name':      name,
            'rule_type':      rule_type,
            'origin':         origin,
            'offense_count':  offense_count,
            'category':       category,
            'recommendation': recommendation,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # Sort: Dead rules first (most actionable), then Noise, then Active
    cat_order = {'Dead': 0, 'Noise Generator': 1, 'Active': 2}
    df['_sort'] = df['category'].map(cat_order).fillna(3)
    df = df.sort_values(['_sort', 'offense_count'], ascending=[True, False]) \
           .drop(columns=['_sort']) \
           .reset_index(drop=True)
    return df


# ─── CHARTS ───────────────────────────────────────────────────────────────────

_BG = '#0a0618'   # dark background baked into every chart PNG


def generate_eps_bar_chart(df, top_n=20):
    """
    Horizontal bar chart: top N log sources by average EPS.
    Bars are coloured by individual % of license cap.
    Returns filepath to temp PNG or None on failure.
    """
    if df.empty:
        return None

    plot_df = df[df['average_eps'] > 0].head(top_n).copy()
    if plot_df.empty:
        return None

    # Reverse so highest EPS is at the top
    plot_df = plot_df.iloc[::-1].reset_index(drop=True)

    labels = plot_df['name'].apply(lambda x: x[:40] + '…' if len(x) > 40 else x)
    values = plot_df['average_eps'].tolist()
    pcts   = plot_df['pct_of_license'].tolist()

    # Colour by % of license
    def _bar_colour(pct):
        if pct >= EPS_SOURCE_WARN_PCT:
            return '#ef4444'
        if pct >= EPS_SOURCE_WARN_PCT / 2:
            return '#f59e0b'
        return '#10b981'

    colours = [_bar_colour(p) for p in pcts]

    fig, ax = plt.subplots(figsize=(7, max(3.5, len(labels) * 0.36)), facecolor=_BG)
    ax.set_facecolor(_BG)

    bars = ax.barh(labels, values, color=colours, height=0.62, edgecolor=_BG, linewidth=1.5)

    # Value annotations on bars
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f'{val:.1f}', va='center', color='#c4b5fd', fontsize=7.5,
                fontfamily='monospace')

    # License cap line
    ax.axvline(x=LICENSE_EPS_CAP, color='#9b72f5', linestyle='--',
               linewidth=1, alpha=0.6, label=f'License cap  ({LICENSE_EPS_CAP} EPS)')

    ax.set_xlabel('Average EPS', color='#7c6fa0', fontsize=8, fontfamily='monospace')
    ax.set_title(f'Top {len(labels)} Log Sources by EPS Consumption',
                 color='#a78bfa', fontsize=9, fontweight='700',
                 fontfamily='monospace', pad=10)
    ax.tick_params(colors='#c4b5fd', labelsize=7.5)
    ax.xaxis.label.set_color('#7c6fa0')
    for spine in ax.spines.values():
        spine.set_edgecolor('#3a2d6a')
        spine.set_linewidth(0.5)
    ax.legend(fontsize=7, frameon=False, labelcolor='#c4b5fd')

    # Colour legend patches
    legend_handles = [
        mpatches.Patch(color='#ef4444', label=f'Warning (>{EPS_SOURCE_WARN_PCT}% of cap)'),
        mpatches.Patch(color='#f59e0b', label='Elevated'),
        mpatches.Patch(color='#10b981', label='Normal'),
    ]
    ax.legend(handles=legend_handles, loc='lower right', fontsize=7,
              frameon=False, labelcolor='#c4b5fd')

    plt.tight_layout(pad=0.8)
    tmp  = tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix='qr_eps_')
    path = tmp.name
    tmp.close()
    plt.savefig(path, bbox_inches='tight', dpi=110, facecolor=_BG, edgecolor='none')
    plt.close()
    return path


def generate_rule_donut(stats):
    """
    Donut chart: Dead / Noise Generator / Active rule distribution.
    stats = {'Dead': int, 'Noise Generator': int, 'Active': int}
    Returns filepath to temp PNG or None.
    """
    filtered = {k: v for k, v in stats.items() if v > 0}
    if not filtered:
        return None

    labels  = list(filtered.keys())
    sizes   = list(filtered.values())
    colours = {'Dead': '#ef4444', 'Noise Generator': '#f59e0b', 'Active': '#10b981'}
    clrs    = [colours.get(l, '#a78bfa') for l in labels]

    fig, ax = plt.subplots(figsize=(5.2, 3.8), facecolor=_BG)
    ax.set_facecolor(_BG)

    _, _, autotexts = ax.pie(
        sizes, labels=None, colors=clrs,
        autopct=lambda p: f'{p:.0f}%' if p > 5 else '',
        startangle=140,
        wedgeprops={'edgecolor': _BG, 'linewidth': 3, 'width': 0.54},
        pctdistance=0.76,
    )
    for at in autotexts:
        at.set_color('#f0eaff')
        at.set_fontweight('bold')
        at.set_fontsize(8)

    total = sum(sizes)
    ax.text(0, 0.10, str(total), ha='center', va='center',
            fontsize=20, fontweight='bold', color='#c4b5fd')
    ax.text(0, -0.22, 'RULES', ha='center', va='center',
            fontsize=6, color='#6d5a9a', fontweight='700', fontfamily='monospace')

    patches = [mpatches.Patch(color=colours.get(l, '#a78bfa'), label=f'{l}  {v}')
               for l, v in zip(labels, sizes)]
    leg = ax.legend(handles=patches, loc='center left', bbox_to_anchor=(1.0, 0.5),
                    fontsize=8, frameon=False, handlelength=1.1, handleheight=0.9,
                    borderpad=0.5, labelspacing=0.7)
    for t in leg.get_texts():
        t.set_color('#c4b5fd')

    ax.set_title(f'Rule Effectiveness  ·  last {OFFENSE_LOOKBACK_DAYS}d',
                 color='#a78bfa', fontsize=9, fontweight='700',
                 pad=10, fontfamily='monospace')
    ax.axis('equal')
    plt.tight_layout(pad=0.5)

    tmp  = tempfile.NamedTemporaryFile(suffix='.png', delete=False, prefix='qr_rules_')
    path = tmp.name
    tmp.close()
    plt.savefig(path, bbox_inches='tight', dpi=110, facecolor=_BG, edgecolor='none')
    plt.close()
    return path


# ─── EXCEL REPORT ─────────────────────────────────────────────────────────────

_FILLS = {
    'red':    PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
    'orange': PatternFill(start_color='FFBF47', end_color='FFBF47', fill_type='solid'),
    'green':  PatternFill(start_color='A8E6CF', end_color='A8E6CF', fill_type='solid'),
    'blue':   PatternFill(start_color='74B9FF', end_color='74B9FF', fill_type='solid'),
    'gray':   PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
    'header': PatternFill(start_color='2D2257', end_color='2D2257', fill_type='solid'),
}
_HDR_FONT  = Font(bold=True, color='E8E0FF', size=10)
_BOLD      = Font(bold=True, size=10)
_CENTRE    = Alignment(horizontal='center', vertical='center')
_LEFT      = Alignment(horizontal='left',   vertical='center')


def _write_sheet_header(ws, columns, col_widths):
    ws.append(columns)
    for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
        cell.fill    = _FILLS['header']
        cell.font    = _HDR_FONT
        cell.alignment = _CENTRE
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20


def save_excel_report(eps_df, rule_df, path):
    """Saves two-sheet Excel: 'EPS Analysis' and 'Rule Effectiveness'."""
    try:
        wb = openpyxl.Workbook()

        # ── Sheet 1: EPS ──────────────────────────────────────────────────────
        ws_eps = wb.active
        ws_eps.title = 'EPS Analysis'

        eps_cols  = ['Rank', 'Log Source Name', 'Type', 'Enabled',
                     'Avg EPS', '% of Total EPS', '% of License Cap', 'Status']
        eps_widths = [6, 44, 28, 9, 10, 16, 18, 12]
        _write_sheet_header(ws_eps, eps_cols, eps_widths)

        if not eps_df.empty:
            for _, row in eps_df.iterrows():
                ws_eps.append([
                    row['rank'],
                    row['name'],
                    row['ls_type'],
                    row['enabled'],
                    row['average_eps'],
                    row['pct_of_total'],
                    row['pct_of_license'],
                    row['eps_status'],
                ])
                r_idx = ws_eps.max_row
                status_cell = ws_eps.cell(row=r_idx, column=8)
                if row['eps_status'] == 'Warning':
                    status_cell.fill = _FILLS['orange']
                    status_cell.font = _BOLD
                elif row['eps_status'] == 'Disabled':
                    status_cell.fill = _FILLS['gray']
                elif row['eps_status'] == 'Idle':
                    status_cell.fill = _FILLS['blue']
                else:
                    status_cell.fill = _FILLS['green']

                if row['pct_of_license'] >= EPS_SOURCE_WARN_PCT:
                    ws_eps.cell(row=r_idx, column=7).fill = _FILLS['orange']
                    ws_eps.cell(row=r_idx, column=7).font = _BOLD

        ws_eps.freeze_panes = 'A2'

        # ── Sheet 2: Rules ────────────────────────────────────────────────────
        ws_rules = wb.create_sheet('Rule Effectiveness')

        rule_cols   = ['Rule ID', 'Rule Name', 'Type', 'Origin',
                       f'Offenses ({OFFENSE_LOOKBACK_DAYS}d)', 'Category', 'Recommendation']
        rule_widths = [10, 52, 18, 16, 16, 18, 64]
        _write_sheet_header(ws_rules, rule_cols, rule_widths)

        if not rule_df.empty:
            for _, row in rule_df.iterrows():
                ws_rules.append([
                    row['rule_id'],
                    row['rule_name'],
                    row['rule_type'],
                    row['origin'],
                    row['offense_count'],
                    row['category'],
                    row['recommendation'],
                ])
                r_idx  = ws_rules.max_row
                cat_cell = ws_rules.cell(row=r_idx, column=6)
                if row['category'] == 'Dead':
                    cat_cell.fill = _FILLS['red']
                    cat_cell.font = _BOLD
                elif row['category'] == 'Noise Generator':
                    cat_cell.fill = _FILLS['orange']
                    cat_cell.font = _BOLD
                else:
                    cat_cell.fill = _FILLS['green']
                ws_rules.cell(row=r_idx, column=7).alignment = _LEFT

        ws_rules.freeze_panes = 'A2'

        wb.save(path)
        print(f"✅ Excel saved → {path}")
    except PermissionError:
        print(f"❌ Cannot save Excel — is '{path}' already open?")
    except Exception as e:
        logger.error("Excel save failed:\n%s", traceback.format_exc())
        print(f"❌ Excel save error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL HTML
# ══════════════════════════════════════════════════════════════════════════════

_C = {
    'purple':   '#9b72f5',
    'violet':   '#c4b5fd',
    'lavender': '#a78bfa',
    'dim':      '#7c6fa0',
    'green':    '#10b981',
    'red':      '#ef4444',
    'orange':   '#f97316',
    'amber':    '#f59e0b',
    'gray':     '#8b9ab0',
    'cyan':     '#06b6d4',
    'blue':     '#3b82f6',
    'badge_red':    '#7f1d1d',
    'badge_amber':  '#78350f',
    'badge_green':  '#065f46',
    'badge_gray':   '#2d3748',
}


def _eps_status_badge(status):
    meta = {
        'Warning':  {'bg': _C['badge_amber'], 'label': 'WARNING',  'icon': '▲'},
        'Disabled': {'bg': _C['badge_gray'],  'label': 'DISABLED', 'icon': '◌'},
        'Idle':     {'bg': _C['badge_gray'],  'label': 'IDLE',     'icon': '●'},
        'OK':       {'bg': _C['badge_green'], 'label': 'OK',       'icon': '✔'},
    }.get(status, {'bg': _C['badge_gray'], 'label': status.upper()[:10], 'icon': '◌'})

    return (
        f'<span style="background:{meta["bg"]};color:#f0eaff;font-size:9px;'
        f'font-weight:700;padding:2px 8px;border-radius:3px;'
        f'letter-spacing:0.5px;white-space:nowrap;font-family:monospace;">'
        f'{meta["icon"]}&nbsp;{meta["label"]}</span>'
    )


def _rule_category_badge(category):
    meta = {
        'Dead':           {'bg': _C['badge_red'],   'label': 'DEAD',   'icon': '●'},
        'Noise Generator':{'bg': _C['badge_amber'],  'label': 'NOISE',  'icon': '▲'},
        'Active':         {'bg': _C['badge_green'],  'label': 'ACTIVE', 'icon': '✔'},
    }.get(category, {'bg': _C['badge_gray'], 'label': category[:8].upper(), 'icon': '◌'})

    return (
        f'<span style="background:{meta["bg"]};color:#f0eaff;font-size:9px;'
        f'font-weight:700;padding:2px 8px;border-radius:3px;'
        f'letter-spacing:0.5px;white-space:nowrap;font-family:monospace;">'
        f'{meta["icon"]}&nbsp;{meta["label"]}</span>'
    )


def _build_eps_table(eps_df, top_n=20):
    C = _C
    rows_html = ''
    _rb = f'border-bottom:1px solid {C["dim"]}30;'

    for _, row in eps_df.head(top_n).iterrows():
        name_short  = str(row['name'])[:44] + '…' if len(str(row['name'])) > 44 else str(row['name'])
        pct_lic     = f"{row['pct_of_license']:.1f}%"
        pct_tot     = f"{row['pct_of_total']:.1f}%"
        avg_eps_str = f"{row['average_eps']:.2f}"
        badge       = _eps_status_badge(row['eps_status'])

        # Highlight the % of license cell colour
        pct_color = C['red'] if row['pct_of_license'] >= EPS_SOURCE_WARN_PCT else \
                    C['amber'] if row['pct_of_license'] >= EPS_SOURCE_WARN_PCT / 2 else \
                    C['green']

        rows_html += f"""
        <tr>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['violet']};
                     font-family:monospace;" title="{row['name']}">{name_short}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['dim']};
                     font-family:monospace;text-align:right;">{avg_eps_str}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{pct_color};
                     font-family:monospace;text-align:right;font-weight:700;">{pct_lic}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['dim']};
                     font-family:monospace;text-align:right;">{pct_tot}</td>
          <td style="padding:7px 10px;{_rb}text-align:center;">{badge}</td>
        </tr>"""

    _hdr = f'border-top:2px solid {C["purple"]};border-bottom:1px solid {C["purple"]}50;'
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;margin-top:8px;">
      <thead><tr>
        <th style="padding:6px 10px;text-align:left;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Log Source</th>
        <th style="padding:6px 10px;text-align:right;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Avg EPS</th>
        <th style="padding:6px 10px;text-align:right;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">% of License</th>
        <th style="padding:6px 10px;text-align:right;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">% of Total</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Status</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>"""


def _build_rules_table(rule_df):
    """
    Shows only Dead + Noise Generator rules (the actionable ones).
    Active rules are healthy — no point listing them in the email body.
    """
    C = _C
    actionable = rule_df[rule_df['category'].isin(['Dead', 'Noise Generator'])]

    if actionable.empty:
        return (
            f'<p style="color:{C["green"]};font-size:11px;font-weight:700;'
            f'font-family:monospace;padding:8px 0;">✔ No dead or noisy rules found.</p>'
        )

    rows_html = ''
    _rb = f'border-bottom:1px solid {C["dim"]}30;'

    for _, row in actionable.iterrows():
        name_short   = str(row['rule_name'])[:50] + '…' if len(str(row['rule_name'])) > 50 else str(row['rule_name'])
        count_color  = C['red'] if row['category'] == 'Dead' else C['amber']
        count_str    = str(row['offense_count'])
        badge        = _rule_category_badge(row['category'])
        rec_short    = str(row['recommendation'])[:80] + '…' if len(str(row['recommendation'])) > 80 else str(row['recommendation'])

        rows_html += f"""
        <tr>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['violet']};
                     font-family:monospace;" title="{row['rule_name']}">{name_short}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['dim']};
                     font-family:monospace;text-align:center;">{row['rule_type']}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{count_color};
                     font-family:monospace;text-align:center;font-weight:700;">{count_str}</td>
          <td style="padding:7px 10px;{_rb}text-align:center;">{badge}</td>
          <td style="padding:7px 10px;{_rb}font-size:10px;color:{C['dim']};
                     max-width:220px;">{rec_short}</td>
        </tr>"""

    _hdr = f'border-top:2px solid {C["purple"]};border-bottom:1px solid {C["purple"]}50;'
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;margin-top:8px;">
      <thead><tr>
        <th style="padding:6px 10px;text-align:left;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Rule Name</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Type</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Offenses</th>
        <th style="padding:6px 10px;text-align:center;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Category</th>
        <th style="padding:6px 10px;text-align:left;font-size:9px;color:{C['purple']};
                   font-weight:700;text-transform:uppercase;letter-spacing:1.4px;
                   font-family:monospace;{_hdr}">Recommendation</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>"""


def build_email_html(eps_df, rule_df, eps_cid, rule_cid):
    """Assembles the full HTML email body."""
    C        = _C
    run_time = datetime.now().strftime('%d %b %Y  ·  %H:%M:%S')

    # ── EPS summary numbers ───────────────────────────────────────────────────
    total_eps     = round(eps_df['average_eps'].sum(), 1) if not eps_df.empty else 0
    pct_used      = round(total_eps / LICENSE_EPS_CAP * 100, 1) if LICENSE_EPS_CAP > 0 else 0
    eps_headroom  = max(0.0, round(LICENSE_EPS_CAP - total_eps, 1))
    eps_warn_srcs = len(eps_df[eps_df['eps_status'] == 'Warning']) if not eps_df.empty else 0
    eps_idle_srcs = len(eps_df[eps_df['average_eps'] == 0])         if not eps_df.empty else 0

    if pct_used >= EPS_TOTAL_CRIT_PCT:
        eps_badge_bg, eps_badge_txt = C['badge_red'],   f'⚠  {pct_used}% OF LICENSE'
    elif pct_used >= 50:
        eps_badge_bg, eps_badge_txt = C['badge_amber'],  f'⚠  {pct_used}% OF LICENSE'
    else:
        eps_badge_bg, eps_badge_txt = C['badge_green'], f'✔  {pct_used}% OF LICENSE'

    # ── Rule summary numbers ──────────────────────────────────────────────────
    total_rules = len(rule_df)  if not rule_df.empty else 0
    dead_rules  = len(rule_df[rule_df['category'] == 'Dead'])           if not rule_df.empty else 0
    noise_rules = len(rule_df[rule_df['category'] == 'Noise Generator']) if not rule_df.empty else 0
    ok_rules    = total_rules - dead_rules - noise_rules

    rule_issues = dead_rules + noise_rules
    if dead_rules > 10 or noise_rules > 5:
        rule_badge_bg, rule_badge_txt = C['badge_red'],   f'⚠  {rule_issues} RULES NEED ATTENTION'
    elif rule_issues > 0:
        rule_badge_bg, rule_badge_txt = C['badge_amber'],  f'⚠  {rule_issues} RULES NEED ATTENTION'
    else:
        rule_badge_bg, rule_badge_txt = C['badge_green'], '✔  RULES HEALTHY'

    def badge(bg, txt):
        return (
            f'<span style="background:{bg};color:#f0eaff;font-size:10px;'
            f'font-weight:700;padding:4px 12px;border-radius:3px;'
            f'letter-spacing:0.8px;font-family:monospace;white-space:nowrap;">'
            f'{txt}</span>'
        )

    def metric(label, value, color, note=''):
        note_html = (
            f'<div style="font-size:9px;color:{C["dim"]};margin-top:3px;'
            f'font-family:monospace;">{note}</div>'
        ) if note else ''
        return (
            f'<td style="padding:0 22px 0 0;text-align:center;vertical-align:top;">'
            f'<div style="font-size:30px;font-weight:800;color:{color};'
            f'line-height:1;font-family:monospace;letter-spacing:-1px;">{value}</div>'
            f'<div style="font-size:9px;color:{C["dim"]};margin-top:4px;'
            f'text-transform:uppercase;letter-spacing:1.2px;">{label}</div>'
            f'{note_html}</td>'
        )

    def chip(label, value, color):
        if value == 0:
            return ''
        return (
            f'<span style="display:inline-block;border-left:3px solid {color};'
            f'padding:1px 10px 1px 7px;margin:3px 8px 3px 0;font-size:10px;'
            f'font-family:monospace;color:{color};font-weight:700;letter-spacing:0.3px;">'
            f'{value}&nbsp;{label}</span>'
        )

    eps_chart_html  = (f'<img src="cid:{eps_cid}"  alt="EPS chart"  '
                       f'style="display:block;max-width:100%;margin:14px auto 0;">'
                       ) if eps_cid else ''
    rule_chart_html = (f'<img src="cid:{rule_cid}" alt="Rule chart" '
                       f'style="display:block;max-width:420px;margin:14px auto 0;">'
                       ) if rule_cid else ''

    eps_table  = _build_eps_table(eps_df)
    rule_table = _build_rules_table(rule_df)

    eps_metrics = (
        metric('Total EPS',     f'{total_eps:.0f}',   C['purple'])  +
        metric('License Cap',   LICENSE_EPS_CAP,       C['dim'])     +
        metric('% Used',        f'{pct_used:.0f}%',   C['amber'] if pct_used > 50 else C['green']) +
        metric('Headroom',      f'{eps_headroom:.0f}', C['cyan'],  'EPS remaining') +
        metric('At-Risk Srcs',  eps_warn_srcs,         C['red']   if eps_warn_srcs > 0 else C['green']) +
        metric('Idle Sources',  eps_idle_srcs,         C['gray'],  '0 EPS avg')
    )
    rule_metrics = (
        metric('Total Rules',   total_rules,  C['purple']) +
        metric('Dead',          dead_rules,   C['red']   if dead_rules > 0 else C['green'],  '0 offenses') +
        metric('Noise Gens',    noise_rules,  C['amber'] if noise_rules > 0 else C['green'], f'≥{RULE_NOISE_THRESHOLD} offenses') +
        metric('Active',        ok_rules,     C['green'])
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="color-scheme" content="light dark">
<meta name="supported-color-schemes" content="light dark">
</head>
<body style="margin:0;padding:0;font-family:'Segoe UI',Helvetica,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="padding:24px 0;">
<tr><td align="center">
<table width="660" cellpadding="0" cellspacing="0" style="max-width:660px;width:100%;">

  <!-- ══ MASTHEAD ══ -->
  <tr><td style="padding:0 0 10px;border-bottom:3px solid {C['purple']};">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>
      <td>
        <div style="font-size:9px;color:{C['dim']};letter-spacing:3px;
                    text-transform:uppercase;font-family:monospace;margin-bottom:8px;">
          QRadar &nbsp;·&nbsp; Weekly Intelligence Report
        </div>
        <div style="font-size:23px;font-weight:800;color:{C['violet']};
                    letter-spacing:-0.5px;line-height:1.2;">
          EPS Burn Rate Monitor  +  Rule Effectiveness Auditor
        </div>
        <div style="margin-top:8px;font-size:11px;color:{C['dim']};
                    font-family:monospace;">{run_time}</div>
      </td>
    </tr></table>
  </td></tr>

  <!-- ══ SECTION 1: EPS ══ -->
  <tr><td style="padding:28px 0 4px;border-top:2px solid {C['purple']};">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>
      <td>
        <span style="font-size:13px;font-weight:700;color:{C['purple']};
                     font-family:monospace;">Card 1  ·  EPS Burn Rate Monitor</span>
        <span style="font-size:10px;color:{C['dim']};margin-left:12px;
                     font-family:monospace;">license: {LICENSE_EPS_CAP} EPS cap</span>
      </td>
      <td align="right">{badge(eps_badge_bg, eps_badge_txt)}</td>
    </tr></table>
  </td></tr>

  <tr><td style="padding:16px 0 8px;">
    <table cellpadding="0" cellspacing="0"><tr>{eps_metrics}</tr></table>
  </td></tr>

  <tr><td style="padding:4px 0 20px;text-align:center;">
    <div style="font-size:9px;color:{C['dim']};text-transform:uppercase;
                letter-spacing:2px;margin-bottom:4px;font-family:monospace;">
      Top log sources by EPS consumption
    </div>
    {eps_chart_html}
  </td></tr>

  <tr><td style="padding:2px 0 4px;">
    <span style="font-size:9px;color:{C['purple']};text-transform:uppercase;
                 letter-spacing:2px;font-family:monospace;font-weight:700;">
      Ranked table — top 20 sources
    </span>
    <span style="font-size:10px;color:{C['dim']};margin-left:10px;">
      Full data in Excel attachment &nbsp;·&nbsp;
      Warning threshold: single source &gt; {EPS_SOURCE_WARN_PCT}% of license cap
    </span>
  </td></tr>
  <tr><td style="padding:0 0 32px;">{eps_table}</td></tr>

  <!-- ══ SECTION 2: RULES ══ -->
  <tr><td style="padding:28px 0 4px;border-top:2px solid {C['purple']};">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>
      <td>
        <span style="font-size:13px;font-weight:700;color:{C['purple']};
                     font-family:monospace;">Card 2  ·  Rule Effectiveness Auditor</span>
        <span style="font-size:10px;color:{C['dim']};margin-left:12px;
                     font-family:monospace;">lookback: {OFFENSE_LOOKBACK_DAYS} days</span>
      </td>
      <td align="right">{badge(rule_badge_bg, rule_badge_txt)}</td>
    </tr></table>
  </td></tr>

  <tr><td style="padding:16px 0 8px;">
    <table cellpadding="0" cellspacing="0"><tr>{rule_metrics}</tr></table>
  </td></tr>

  <tr><td style="padding:4px 0 20px;text-align:center;">
    {chip('Dead rules',       dead_rules,  C['red'])}
    {chip('Noise generators', noise_rules, C['amber'])}
    {chip('Active rules',     ok_rules,    C['green'])}
    {rule_chart_html}
  </td></tr>

  <tr><td style="padding:2px 0 4px;">
    <span style="font-size:9px;color:{C['purple']};text-transform:uppercase;
                 letter-spacing:2px;font-family:monospace;font-weight:700;">
      Actionable rules only
    </span>
    <span style="font-size:10px;color:{C['dim']};margin-left:10px;">
      Dead · Noise generators — all {total_rules} rules in Excel attachment
    </span>
  </td></tr>
  <tr><td style="padding:0 0 30px;">{rule_table}</td></tr>

  <!-- ══ FOOTER ══ -->
  <tr><td style="padding:16px 0 20px;border-top:1px solid {C['purple']}30;">
    <div style="font-size:9px;color:{C['dim']};font-family:monospace;
                letter-spacing:0.5px;">
      QRadar Weekly Intelligence &nbsp;·&nbsp; Auto-generated {run_time} &nbsp;·&nbsp;
      Source: QRadar REST API (no AQL)
    </div>
  </td></tr>

</table>
</td></tr>
</table>
</body></html>"""


# ─── OUTLOOK DRAFT ────────────────────────────────────────────────────────────

def create_outlook_draft(excel_path, subject, html_body, images):
    """
    Creates an Outlook draft with embedded PNG charts and Excel attachment.
    images = {'cid_key': '/path/to/file.png', ...}
    """
    try:
        outlook = win32com.client.Dispatch('Outlook.Application')
        mail    = outlook.CreateItem(0)
        mail.Subject = subject

        if excel_path and os.path.exists(excel_path):
            mail.Attachments.Add(excel_path)

        for cid, img_path in images.items():
            if img_path and os.path.exists(img_path):
                att = mail.Attachments.Add(img_path)
                att.PropertyAccessor.SetProperty(_MAPI_PR_ATTACH_CONTENT_ID, cid)

        mail.HTMLBody = html_body
        mail.Display()
        print("\n✉️  Outlook draft created.")

    except Exception as e:
        logger.error("Outlook draft failed:\n%s", traceback.format_exc())
        print(f"\n❌ Outlook draft error: {e}")

    finally:
        # Clean up temp PNG files
        for cid, img_path in images.items():
            if img_path and os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except Exception:
                    pass


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if not VERIFY_SSL:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print("=" * 62)
    print("  QRadar EPS Burn Rate Monitor + Rule Effectiveness Auditor")
    print("=" * 62)
    print(f"  Host         : {QRADAR_HOST}")
    print(f"  License cap  : {LICENSE_EPS_CAP} EPS")
    print(f"  Rule lookback: {OFFENSE_LOOKBACK_DAYS} days")
    print(f"  Retry config : {MAX_RETRIES} attempts, {RETRY_DELAY_BASE}s base backoff")
    print("=" * 62)

    if not test_connection():
        return

    # ── Fetch all data ────────────────────────────────────────────────────────
    type_map     = fetch_log_source_types()
    log_sources  = fetch_log_sources()
    rules        = fetch_all_rules()
    offenses     = fetch_recent_offenses()

    if not log_sources and not rules:
        print("\n❌ No data returned from QRadar — check credentials and host URL.")
        return

    # ── Analyse ───────────────────────────────────────────────────────────────
    print("\n🔍 Running EPS analysis...")
    eps_df = analyze_eps(log_sources, type_map)
    if not eps_df.empty:
        total_eps = eps_df['average_eps'].sum()
        pct_used  = total_eps / LICENSE_EPS_CAP * 100 if LICENSE_EPS_CAP > 0 else 0
        warn_srcs = len(eps_df[eps_df['eps_status'] == 'Warning'])
        print(f"   Total EPS   : {total_eps:.1f} / {LICENSE_EPS_CAP}  ({pct_used:.1f}% of license)")
        print(f"   At-risk srcs: {warn_srcs} source(s) > {EPS_SOURCE_WARN_PCT}% of cap")

    print("\n🔍 Running rule effectiveness analysis...")
    rule_df = analyze_rules(rules, offenses)
    if not rule_df.empty:
        dead    = len(rule_df[rule_df['category'] == 'Dead'])
        noise   = len(rule_df[rule_df['category'] == 'Noise Generator'])
        active  = len(rule_df[rule_df['category'] == 'Active'])
        print(f"   Total rules : {len(rule_df)}")
        print(f"   Dead        : {dead}")
        print(f"   Noise gens  : {noise}")
        print(f"   Active      : {active}")

    # ── Save Excel ────────────────────────────────────────────────────────────
    print(f"\n💾 Saving Excel report → {OUTPUT_EXCEL}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    save_excel_report(eps_df, rule_df, OUTPUT_EXCEL)

    # ── Generate charts ───────────────────────────────────────────────────────
    print("\n📊 Generating charts...")
    images = {}

    eps_chart_path = generate_eps_bar_chart(eps_df)
    if eps_chart_path:
        images['eps_chart'] = eps_chart_path
        print("   ✅ EPS bar chart generated.")
    else:
        print("   ⚠️  EPS bar chart skipped (no active sources with EPS > 0).")

    rule_stats = {
        'Dead':            len(rule_df[rule_df['category'] == 'Dead'])            if not rule_df.empty else 0,
        'Noise Generator': len(rule_df[rule_df['category'] == 'Noise Generator']) if not rule_df.empty else 0,
        'Active':          len(rule_df[rule_df['category'] == 'Active'])           if not rule_df.empty else 0,
    }
    rule_chart_path = generate_rule_donut(rule_stats)
    if rule_chart_path:
        images['rule_chart'] = rule_chart_path
        print("   ✅ Rule effectiveness donut generated.")

    # ── Build and send email ──────────────────────────────────────────────────
    print("\n✉️  Building email draft...")
    html_body = build_email_html(
        eps_df,
        rule_df,
        eps_cid  = 'eps_chart'  if 'eps_chart'  in images else None,
        rule_cid = 'rule_chart' if 'rule_chart' in images else None,
    )

    total_eps     = round(eps_df['average_eps'].sum(), 1) if not eps_df.empty else 0
    pct_used      = round(total_eps / LICENSE_EPS_CAP * 100, 1) if LICENSE_EPS_CAP > 0 else 0
    total_issues  = rule_stats['Dead'] + rule_stats['Noise Generator']
    subject = (
        f"QRadar Weekly Intelligence — "
        f"EPS: {pct_used:.0f}% of license  |  "
        f"Rules: {total_issues} requiring attention"
    )
    create_outlook_draft(OUTPUT_EXCEL, subject, html_body, images)
    print("\n✅ Done!")


if __name__ == '__main__':
    main()
